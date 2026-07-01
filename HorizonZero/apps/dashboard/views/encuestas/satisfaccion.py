# apps/dashboard/views/encuestas/satisfaccion.py
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from apps.core.decorators import permiso_requerido
from apps.dashboard.mixins import aplicar_restricciones_perfil
from apps.dashboard.services.encuesta_satisfaccion_api import EncuestaSatisfaccionAPIService
from apps.dashboard.tables import EncuestaSatisfaccionTable
from apps.dashboard.views._base import build_timeline_heatmap


DEBUG_DUMP_PATH = Path(__file__).resolve().parents[4] / "encuesta_satisfaccion_debug.json"
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _append_view_debug(section: str, payload: dict):
    try:
        data = {}
        if DEBUG_DUMP_PATH.exists():
            data = json.loads(DEBUG_DUMP_PATH.read_text(encoding="utf-8"))
        data[section] = payload
        DEBUG_DUMP_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _safe_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return ILLEGAL_EXCEL_CHARS_RE.sub("", str(value))


CAMPOS_ENCUESTA_SAT = {
    "agente": "Agente",
    "unidad": "Unidad",
    "sucursal": "Sucursal",
    "gestion": "Gestion",
    "nombre": "Nombre",
}


@login_required
def encuesta_satisfaccion_buscar(request, campo):
    if campo not in CAMPOS_ENCUESTA_SAT:
        return JsonResponse({"results": []})

    term = request.GET.get("term", "").strip()
    if not term:
        return JsonResponse({"results": []})

    try:
        data = EncuestaSatisfaccionAPIService.buscar_valores(campo, term, request=request)
    except Exception:
        data = []

    return JsonResponse({"results": data, "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_kpis(request):
    kpi_sucursales = request.GET.getlist("kpi_sucursal")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio", "").strip() or None
    kpi_fecha_fin = request.GET.get("kpi_fecha_fin", "").strip() or None

    try:
        kpis = EncuestaSatisfaccionAPIService.obtener_kpis_globales(
            request=request,
            sucursales=kpi_sucursales if kpi_sucursales else None,
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse(kpis)


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion(request):
    filtros = {
        "agente": request.GET.get("agente"),
        "sucursal": request.GET.get("sucursal"),
        "unidad": request.GET.get("unidad"),
        "gestion": request.GET.get("gestion"),
        "nombre": request.GET.get("nombre"),
        "cedula": request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin": request.GET.get("fecha_fin"),
        "clasificacion": request.GET.get("clasificacion"),
    }

    kpi_sucursales = request.GET.getlist("kpi_sucursal")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio", "")
    kpi_fecha_fin = request.GET.get("kpi_fecha_fin", "")

    filtros, unidad_forzada, sucursal_forzada = aplicar_restricciones_perfil(request, filtros)

    datos = []
    opciones_sucursal = []
    opciones_unidad = []
    timeline_data = []
    heatmap_anios = []
    heatmap_json = "{}"

    try:
        datos = EncuestaSatisfaccionAPIService.obtener_listado(filtros=filtros, request=request)
    except Exception as e:
        _append_view_debug("view_listado_error", {
            "timestamp": datetime.now().isoformat(),
            "error": str(e),
            "filtros": filtros,
        })
        messages.error(request, "Error al obtener los registros de encuesta.")

    try:
        opciones_sucursal, opciones_unidad = EncuestaSatisfaccionAPIService.obtener_opciones(
            filtros={
                "fecha_inicio": filtros.get("fecha_inicio"),
                "fecha_fin": filtros.get("fecha_fin"),
                "agente": filtros.get("agente"),
                "cedula": filtros.get("cedula"),
                "nombre": filtros.get("nombre"),
            },
            request=request,
            items=datos,
        )
    except Exception as e:
        _append_view_debug("view_opciones_error", {
            "timestamp": datetime.now().isoformat(),
            "error": str(e),
            "filtros": filtros,
        })

    try:
        raw_timeline = EncuestaSatisfaccionAPIService.obtener_timeline(filtros=filtros, request=request)
        timeline_data, heatmap_anios, heatmap_json = build_timeline_heatmap(raw_timeline)
        _append_view_debug("view_timeline_success", {
            "timestamp": datetime.now().isoformat(),
            "raw_timeline": raw_timeline,
            "timeline_data": timeline_data,
            "heatmap_anios": heatmap_anios,
            "heatmap_json": heatmap_json,
        })
    except Exception as e:
        _append_view_debug("view_timeline_error", {
            "timestamp": datetime.now().isoformat(),
            "error": str(e),
            "filtros": filtros,
        })
        messages.error(request, "Error al obtener la evolución mensual.")

    table = EncuestaSatisfaccionTable(datos)

    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion.html", {
        "table": table,
        "filtros": filtros,
        "kpi_sucursales": kpi_sucursales,
        "kpi_fecha_inicio": kpi_fecha_inicio,
        "kpi_fecha_fin": kpi_fecha_fin,
        "opciones_sucursal": opciones_sucursal,
        "opciones_unidad": opciones_unidad,
        "timeline_data": timeline_data,
        "heatmap_anios": heatmap_anios,
        "heatmap_json": heatmap_json,
        "unidad_forzada": unidad_forzada,
        "sucursal_forzada": sucursal_forzada,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_detalle(request, respuesta_id):
    try:
        detalle = EncuestaSatisfaccionAPIService.obtener_detalle(respuesta_id, request=request)
        encabezado = detalle["encabezado"]
        filas = detalle["preguntas"]
    except Exception:
        detalle = {}
        encabezado = {}
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    respuestas_numericas = []
    for fila in filas:
        try:
            respuestas_numericas.append(float(fila.get("Respuesta", 0)))
        except (ValueError, TypeError):
            pass

    satisfechos = sum(1 for r in respuestas_numericas if r >= 4)
    satisfaccion_pct = round((satisfechos / len(respuestas_numericas) * 100)) if respuestas_numericas else 0

    try:
        agente_kpis = EncuestaSatisfaccionAPIService.obtener_kpis_globales(
            request=request,
            agente=encabezado.get("Agente"),
        ) if encabezado.get("Agente") else {}
    except Exception:
        agente_kpis = {}

    promedio_agente_pct = agente_kpis.get("promedio_general", 0)
    promedio_encuesta = detalle.get("promedio_encuesta", 0) or 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    respuesta_satisfaccion = ""
    for fila in filas:
        if "satisfecho" in (fila.get("Pregunta") or "").lower():
            respuesta_satisfaccion = fila.get("Respuesta", "")
            break

    kpis = {
        "promedio_general": promedio_agente_pct,
        "total_encuestas": agente_kpis.get("total_encuestas", 0),
        "promedio_encuesta": promedio_encuesta_pct,
        "respuesta_satisfaccion": respuesta_satisfaccion,
        "satisfaccion_pct": satisfaccion_pct,
        "agente": encabezado.get("Agente", ""),
        "sucursal": encabezado.get("Sucursal", ""),
        "unidad": encabezado.get("Unidad", ""),
    }

    return render(request, "dashboard/encuestas/satisfaccion_detalle.html", {
        "encabezado": encabezado,
        "preguntas": filas,
        "respuesta_id": respuesta_id,
        "kpis": kpis,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_exportar(request):
    filtros = {
        "agente": request.GET.get("agente"),
        "sucursal": request.GET.get("sucursal"),
        "unidad": request.GET.get("unidad"),
        "gestion": request.GET.get("gestion"),
        "nombre": request.GET.get("nombre"),
        "cedula": request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin": request.GET.get("fecha_fin"),
        "clasificacion": request.GET.get("clasificacion"),
    }
    try:
        datos = EncuestaSatisfaccionAPIService.obtener_detalles_para_exportar(filtros=filtros, request=request)

        encuestas, preguntas_orden = {}, []
        for fila in datos:
            rid = fila.get("respuesta_id")
            if not rid:
                continue

            pregunta = _safe_excel_value(fila.get("Pregunta", ""))
            respuesta = _safe_excel_value(fila.get("Respuesta", ""))

            if rid not in encuestas:
                encuestas[rid] = {
                    "respuesta_id": rid,
                    "Fecha": _safe_excel_value(fila.get("Fecha", "")),
                    "Agente": _safe_excel_value(fila.get("Agente", "")),
                    "Unidad": _safe_excel_value(fila.get("Unidad", "")),
                    "Sucursal": _safe_excel_value(fila.get("Sucursal", "")),
                    "Gestion": _safe_excel_value(fila.get("Gestion", "")),
                    "Nombre": _safe_excel_value(fila.get("Nombre", "")),
                    "Cedula": _safe_excel_value(fila.get("Cedula", "")),
                    "clasificacion": _safe_excel_value(fila.get("clasificacion", "")),
                    "promedio": _safe_excel_value(fila.get("promedio_encuesta", "")),
                }
            if pregunta and pregunta not in preguntas_orden:
                preguntas_orden.append(pregunta)

            if pregunta:
                encuestas[rid][pregunta] = respuesta

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Encuesta Satisfaccion"

        cols_fijas = ["ID", "Fecha", "Agente", "Unidad", "Sucursal",
                      "Gestion", "Accionista", "Cedula", "Clasificacion", "Promedio"]
        keys_fijas = ["respuesta_id", "Fecha", "Agente", "Unidad", "Sucursal",
                      "Gestion", "Nombre", "Cedula", "clasificacion", "promedio"]

        hf = PatternFill("solid", fgColor="003FB7")
        hfont = Font(bold=True, color="FFFFFF", name="Arial")
        hfont_preg = Font(bold=True, color="1A1000", name="Arial")

        for col_idx, col_name in enumerate(cols_fijas, 1):
            c = ws.cell(row=1, column=col_idx, value=col_name)
            c.fill = hf
            c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")

        for i, pregunta in enumerate(preguntas_orden, len(cols_fijas) + 1):
            c = ws.cell(row=1, column=i, value=pregunta)
            c.fill = PatternFill("solid", fgColor="FFC900")
            c.font = hfont_preg
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, (_, enc) in enumerate(encuestas.items(), 2):
            for col_idx, key in enumerate(keys_fijas, 1):
                ws.cell(row=row_idx, column=col_idx, value=_safe_excel_value(enc.get(key, "")))
            for i, pregunta in enumerate(preguntas_orden, len(cols_fijas) + 1):
                ws.cell(row=row_idx, column=i, value=_safe_excel_value(enc.get(pregunta, "")))
            if row_idx % 2 == 0:
                ff = PatternFill("solid", fgColor="E8EFFE")
                for c in range(1, len(cols_fijas) + len(preguntas_orden) + 1):
                    ws.cell(row=row_idx, column=c).fill = ff

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        ws.row_dimensions[1].height = 60
        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="encuesta_satisfaccion.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        _append_view_debug("view_exportar_error", {
            "timestamp": datetime.now().isoformat(),
            "error": str(e),
            "filtros": filtros,
        })
        return JsonResponse({"error": str(e) or "Error generando el archivo."}, status=500)


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_detalle_exportar(request, respuesta_id):
    try:
        detalle = EncuestaSatisfaccionAPIService.obtener_detalle(respuesta_id, request=request)
        filas = detalle["preguntas"]
        encabezado = detalle["encabezado"]
    except Exception:
        detalle = {}
        filas = []
        encabezado = {}

    if not filas:
        return HttpResponse("Sin datos", status=404)

    try:
        stats_agente = EncuestaSatisfaccionAPIService.obtener_kpis_globales(
            request=request,
            agente=encabezado.get("Agente"),
        ) if encabezado.get("Agente") else {}
    except Exception:
        stats_agente = {}

    promedio_agente_pct = stats_agente.get("promedio_general", 0)
    promedio_encuesta = detalle.get("promedio_encuesta", 0) or 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Encuesta"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    der = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for col, w in zip("ABCDEFG", [2, 26, 2, 35, 20, 20, 2]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = centro

    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte de Encuesta de Satisfaccion")
    c.font = font(bold=True, color="003FB7", size=15)
    c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9)
    c.alignment = izq

    c = ws.cell(row=2, column=6, value="FECHA DE GENERACION")
    c.font = font(bold=True, color="FFC900", size=8)
    c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11)
    c.alignment = der

    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    kpi_cols = [4, 5, 6]
    kpi_labels = ["PROMEDIO GENERAL AGENTE", "ENCUESTAS DEL AGENTE", "PROMEDIO ESTA ENCUESTA"]
    kpi_vals = [f"{promedio_agente_pct}%", str(stats_agente.get("total_encuestas", 0)), f"{promedio_encuesta_pct}%"]
    kpi_fills = ["003FB7", "E8EEFF", "8B1A0A"]
    kpi_fonts = ["FFFFFF", "003FB7", "FFFFFF"]

    for i, col in enumerate(kpi_cols):
        c = ws.cell(row=5, column=col, value=kpi_labels[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=8)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

        c = ws.cell(row=6, column=col, value=kpi_vals[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=22)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

    panel_items = [
        ("AGENTE RESPONSABLE", encabezado.get("Agente", "")),
        ("UNIDAD INSTITUCIONAL", encabezado.get("Unidad", "")),
        ("SUCURSAL REGIONAL", encabezado.get("Sucursal", "")),
        ("TIPO DE GESTION", encabezado.get("Gestion", "")),
        ("NOMBRE ACCIONISTA", encabezado.get("Nombre", "-")),
        ("CEDULA", encabezado.get("Cedula", "-")),
        ("FECHA", str(encabezado.get("Fecha", ""))),
    ]

    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18
    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = izq

    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = centro

    row_data = 9
    for i in range(max(len(panel_items) * 2, len(filas))):
        ws.row_dimensions[row_data + i].height = 22

    for i, (label, valor) in enumerate(panel_items):
        r_label = row_data + (i * 2)
        r_valor = row_data + (i * 2) + 1
        c = ws.cell(row=r_label, column=2, value=label)
        c.fill = fill("FFF0C0")
        c.font = font(bold=True, color="003FB7", size=8)
        c.alignment = izq
        c = ws.cell(row=r_valor, column=2, value=valor)
        c.fill = fill("FFF8E0")
        c.font = font(color="1A1A1A", size=10)
        c.alignment = izq

    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9)
        c.alignment = izq
        c.border = border_full("E0E0E0")

        respuesta = fila.get("Respuesta", "")
        color_resp = "4CAF50" if respuesta in ["Muy satisfecho", "Muy fácil", "Sí"] else \
                     "F44336" if respuesta in ["No", "Nada satisfecho", "Muy difícil"] else "003FB7"
        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro
        c.border = border_full(color_resp)

    row_pie = row_data + max(len(panel_items) * 2, len(filas)) + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")

    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie + 1, start_column=2, end_row=row_pie + 1, end_column=6)
    c = ws.cell(row=row_pie + 1, column=2, value="© 2026 Caja de Ande · Documento generado automaticamente · Confidencial")
    c.font = font(color="74788A", size=8)
    c.alignment = centro

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="encuesta_detalle_{respuesta_id}.xlsx"'
    wb.save(response)
    return response
