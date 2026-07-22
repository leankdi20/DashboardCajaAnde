# apps/dashboard/views/encuestas/whatsapp_agente.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports import encuesta_whatsapp_agente as _rep
from apps.dashboard.services.db_service import ReportesDBService
from apps.dashboard.tables import EncuestaWhatsappAgenteTable
from apps.dashboard.views._base import build_timeline_heatmap


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsApp_agente(request):
    filtros = {
        "agente":        request.GET.get("agente"),
        "nombre":        request.GET.get("nombre"),
        "cedula":        request.GET.get("cedula"),
        "fecha_inicio":  request.GET.get("fecha_inicio"),
        "fecha_fin":     request.GET.get("fecha_fin"),
        "clasificacion": request.GET.get("clasificacion"),
    }
    kpi_agentes      = request.GET.getlist("kpi_agente")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    datos, opciones_agente, opciones_nombre = [], [], []
    kpis_globales = {}
    timeline_data, heatmap_anios, heatmap_json = [], [], "{}"

    try:
        datos         = _rep.ReporteEncuestaWhatsappAgente.obtener_datos_agrupados(filtros)
        kpis_globales = _rep.ReporteEncuestaWhatsappAgente.obtener_kpis_globales(
            agentes=kpi_agentes if kpi_agentes else None,
            fecha_inicio=kpi_fecha_inicio, fecha_fin=kpi_fecha_fin,
        )
        opciones_agente = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Agente FROM dbo.vw_reporte_encuestas_satisfaccion_whatsapp_agente "
            "WHERE Agente IS NOT NULL ORDER BY Agente"
        )
        opciones_nombre = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_whatsapp_agente "
            "WHERE Nombre IS NOT NULL ORDER BY Nombre"
        )
        raw_timeline = _rep.ReporteEncuestaWhatsappAgente.obtener_timeline()
        timeline_data, heatmap_anios, heatmap_json = build_timeline_heatmap(raw_timeline)
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaWhatsappAgenteTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp_agente.html", {
        "table": table, "filtros": filtros,
        "kpis_globales": kpis_globales, "kpi_agentes": kpi_agentes,
        "kpi_fecha_inicio": kpi_fecha_inicio, "kpi_fecha_fin": kpi_fecha_fin,
        "opciones_agente": opciones_agente, "opciones_nombre": opciones_nombre,
        "timeline_data": timeline_data, "heatmap_anios": heatmap_anios, "heatmap_json": heatmap_json,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_detalle(request, respuesta_id):
    try:
        filas = _rep.ReporteEncuestaWhatsappAgente.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    encabezado        = filas[0] if filas else {}
    stats_agente      = _rep.ReporteEncuestaWhatsappAgente.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta = _rep.ReporteEncuestaWhatsappAgente.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"]  / 5) * 100) if stats_agente["promedio_agente"]  else 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp_agente_detalle.html", {
        "encabezado": encabezado, "preguntas": filas, "respuesta_id": respuesta_id,
        "kpis": {
            "promedio_general":  promedio_agente_pct,
            "total_encuestas":   stats_agente["total_encuestas"],
            "promedio_encuesta": promedio_encuesta_pct,
            "agente":            encabezado.get("Agente", ""),
        },
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_exportar(request):
    filtros = {
        k: request.GET.get(k)
        for k in ("agente", "nombre", "cedula", "fecha_inicio", "fecha_fin", "clasificacion")
    }
    datos   = _rep.ReporteEncuestaWhatsappAgente.obtener_datos(filtros)

    encuestas, preguntas_orden = {}, []
    for fila in datos:
        rid = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        if rid not in encuestas:
            encuestas[rid] = {"respuesta_id": rid, "Fecha": fila.get("Fecha", ""),
                              "Agente": fila.get("Agente", ""), "Nombre": fila.get("Nombre", ""),
                              "Cedula": fila.get("Cedula", "")}
        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)
        encuestas[rid][pregunta] = fila.get("Respuesta", "")

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "WhatsApp Agente"
    cols_fijas = ["ID", "Fecha", "Agente", "Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha", "Agente", "Nombre", "Cedula"]
    hf = PatternFill("solid", fgColor="003FB7"); hfont = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(cols_fijas, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center")
    for i, p in enumerate(preguntas_orden, len(cols_fijas) + 1):
        c = ws.cell(row=1, column=i, value=p)
        c.fill = PatternFill("solid", fgColor="FFC900"); c.font = Font(bold=True, color="1A1000")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, (rid, enc) in enumerate(encuestas.items(), 2):
        for col_idx, key in enumerate(keys_fijas, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(enc.get(key, "") or ""))
        for i, p in enumerate(preguntas_orden, len(cols_fijas) + 1):
            ws.cell(row=row_idx, column=i, value=str(enc.get(p, "") or ""))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 60)
    ws.row_dimensions[1].height = 60

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="whatsapp_agente.xlsx"'
    wb.save(response); return response


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_detalle_exportar(request, respuesta_id):
    try:
        filas = _rep.ReporteEncuestaWhatsappAgente.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado        = filas[0]
    stats_agente      = _rep.ReporteEncuestaWhatsappAgente.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta = _rep.ReporteEncuestaWhatsappAgente.obtener_promedio_encuesta(respuesta_id)
    prom_ag_pct  = round((stats_agente["promedio_agente"]  / 5) * 100) if stats_agente["promedio_agente"]  else 0
    prom_enc_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Reporte WhatsApp Agente"

    def fill(c): return PatternFill("solid", fgColor=c)
    def bfc(c="CCCCCC"):
        s = Side(style="thin", color=c); return Border(left=s, right=s, top=s, bottom=s)
    def font(bold=False, color="1A1A1A", size=10): return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    for col, w in zip("ABCDEFG", [2, 26, 2, 35, 20, 20, 2]):
        ws.column_dimensions[col].width = w
    for r, h in [(1,5),(2,40),(3,18),(4,5),(5,16),(6,40),(7,5)]:
        ws.row_dimensions[r].height = h

    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7"); c.font = font(True,"FFFFFF",10); c.alignment = centro

    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte Encuesta WhatsApp Agente")
    c.font = font(True,"003FB7",15); c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A",size=9); c.alignment = izq

    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(True,"FFC900",8); c.alignment = der
    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(True,"003FB7",11); c.alignment = der

    ws.row_dimensions[4].height = 4
    for col in range(2, 7): ws.cell(row=4, column=col).fill = fill("003FB7")

    for col, label, val, fgc, fc in zip(
        [4,5,6],
        ["PROMEDIO GENERAL AGENTE","ENCUESTAS DEL AGENTE","PROMEDIO ESTA ENCUESTA"],
        [f"{prom_ag_pct}%", str(stats_agente["total_encuestas"]), f"{prom_enc_pct}%"],
        ["003FB7","E8EEFF","8B1A0A"],
        ["FFFFFF","003FB7","FFFFFF"],
    ):
        c = ws.cell(row=5, column=col, value=label)
        c.fill = fill(fgc); c.font = font(True,fc,8); c.alignment = centro; c.border = bfc(fgc)
        c = ws.cell(row=6, column=col, value=val)
        c.fill = fill(fgc); c.font = font(True,fc,22); c.alignment = centro; c.border = bfc(fgc)

    panel = [("AGENTE RESPONSABLE", encabezado.get("Agente","")),
             ("NOMBRE ACCIONISTA",  encabezado.get("Nombre","—")),
             ("CÉDULA",             encabezado.get("Cedula","—")),
             ("FECHA",              str(encabezado.get("Fecha",""))),
             ("HORA",               str(encabezado.get("Hora","")))]

    row_enc = 8; ws.row_dimensions[row_enc].height = 18
    ws.merge_cells(start_row=row_enc, start_column=4, end_row=row_enc, end_column=5)
    c = ws.cell(row=row_enc, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF"); c.font = font(True,"003FB7",9); c.alignment = izq
    c = ws.cell(row=row_enc, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF"); c.font = font(True,"003FB7",9); c.alignment = centro

    row_data = 9
    for i in range(max(len(panel)*2, len(filas))): ws.row_dimensions[row_data+i].height = 22

    for i, (lbl, val) in enumerate(panel):
        c = ws.cell(row=row_data+(i*2),   column=2, value=lbl)
        c.fill = fill("FFF0C0"); c.font = font(True,"003FB7",8); c.alignment = izq
        c = ws.cell(row=row_data+(i*2)+1, column=2, value=val)
        c.fill = fill("FFF8E0"); c.font = font(color="1A1A1A",size=10); c.alignment = izq

    for i, fila in enumerate(filas):
        r = row_data + i; ws.row_dimensions[r].height = 28
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta",""))
        c.fill = fill("FFFFFF") if i%2==0 else fill("F4F7FF")
        c.font = font(color="1A1A1A",size=9); c.alignment = izq; c.border = bfc("E0E0E0")

        resp = fila.get("Respuesta","")
        try:
            val  = float(resp)
            clr  = "4CAF50" if val >= 4 else "FFC107" if val == 3 else "F44336"
        except (ValueError, TypeError):
            clr  = "003FB7"
        c = ws.cell(row=r, column=6, value=resp)
        c.fill = fill(clr); c.font = font(True,"FFFFFF",9); c.alignment = centro; c.border = bfc(clr)

    rp = row_data + max(len(panel)*2, len(filas)) + 1
    ws.row_dimensions[rp].height = 4
    for col in range(2,7): ws.cell(row=rp, column=col).fill = fill("FFC900")
    ws.row_dimensions[rp+1].height = 14
    ws.merge_cells(start_row=rp+1, start_column=2, end_row=rp+1, end_column=6)
    c = ws.cell(row=rp+1, column=2, value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A",size=8); c.alignment = centro

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="whatsapp_agente_detalle_{respuesta_id}.xlsx"'
    wb.save(response); return response
