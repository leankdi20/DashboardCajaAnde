# apps/dashboard/views/encuestas/feria_salud.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import django_tables2 as dt2
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports import encuesta_feria_salud as _rep


def _build_feria_table(datos, preguntas_orden):
    attrs = {"Fecha": dt2.Column(verbose_name="Fecha")}
    for pregunta in preguntas_orden:
        attrs[pregunta] = dt2.Column(verbose_name=pregunta, default="—", orderable=False)

    Meta = type("Meta", (), {
        "template_name": "dashboard/components/table.html",
        "attrs":         {"class": "w-full text-left border-collapse"},
        "row_attrs":     {"class": "hover:bg-surface-container-low transition-colors cursor-pointer"},
        "sequence":      ("Fecha", *preguntas_orden),
    })
    attrs["Meta"] = Meta
    return type("EncuestaFeriaSaludDynamicTable", (dt2.Table,), attrs)(datos)


@login_required
@permiso_requerido("dashboard.view_encuesta_feria_salud")
def encuesta_feria_salud_(request):
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio", ""),
        "fecha_fin":    request.GET.get("fecha_fin", ""),
    }
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio", "")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin", "")

    try:
        datos, preguntas_orden = _rep.ReporteEncuestaFeriaSalud.obtener_datos_agrupados(filtros)
    except Exception:
        datos, preguntas_orden = [], []

    try:
        total = _rep.ReporteEncuestaFeriaSalud.obtener_total(
            filtros.get("fecha_inicio"), filtros.get("fecha_fin")
        )
    except Exception:
        total = 0

    table = _build_feria_table(datos, preguntas_orden)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=25)
    except Exception:
        table.paginate(page=1, per_page=25)

    return render(request, "dashboard/encuestas/satisfaccion_feria_salud.html", {
        "table": table, "filtros": filtros, "total_encuestas": total,
        "kpi_fecha_inicio": kpi_fecha_inicio, "kpi_fecha_fin": kpi_fecha_fin,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_feria_salud")
def encuesta_feria_salud_exportar(request):
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }
    try:
        datos, preguntas_orden = _rep.ReporteEncuestaFeriaSalud.obtener_datos_agrupados(filtros)
    except Exception:
        datos, preguntas_orden = [], []

    if not datos:
        return HttpResponse("Sin datos", status=404)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Feria de la Salud"
    hf    = PatternFill("solid", fgColor="003FB7")
    hfont = Font(bold=True, color="FFFFFF")
    pfill = PatternFill("solid", fgColor="FFC900")
    pfont = Font(bold=True, color="1A1000")

    ws.cell(row=1, column=1, value="ID").fill = hf
    ws.cell(row=1, column=1).font = hfont
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=2, value="Fecha").fill = hf
    ws.cell(row=1, column=2).font = hfont
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    for i, pregunta in enumerate(preguntas_orden, 3):
        c = ws.cell(row=1, column=i, value=pregunta)
        c.fill = pfill; c.font = pfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, enc in enumerate(datos, 2):
        ws.cell(row=row_idx, column=1, value=enc.get("respuesta_id", ""))
        ws.cell(row=row_idx, column=2, value=str(enc.get("Fecha", "")))
        for i, pregunta in enumerate(preguntas_orden, 3):
            valor = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(valor) if valor else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 60)
    ws.row_dimensions[1].height = 60

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="feria_salud.xlsx"'
    wb.save(response); return response
