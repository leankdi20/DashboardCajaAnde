# apps/dashboard/views/encuestas/pagina_web.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports import encuesta_pagina_web as _rep
from apps.dashboard.services.db_service import ReportesDBService
from apps.dashboard.tables import EncuestaPaginaWebTable
from apps.dashboard.views._base import build_timeline_heatmap


@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web(request):
    filtros = {
        "nombre":         request.GET.get("nombre"),
        "fecha_inicio":   request.GET.get("fecha_inicio"),
        "fecha_fin":      request.GET.get("fecha_fin"),
        "sitio_evaluado": request.GET.get("sitio_evaluado"),
        "segmento_nps":   request.GET.get("segmento_nps"),
    }
    kpi_fecha_inicio  = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin     = request.GET.get("kpi_fecha_fin")
    kpi_sitio_evaluado = request.GET.get("kpi_sitio_evaluado")

    datos, opciones_nombre, opciones_sitio = [], [], []
    kpis_globales = {}
    timeline_data, heatmap_anios, heatmap_json = [], [], "{}"

    try:
        datos         = _rep.ReporteEncuestaPaginaWeb.obtener_datos_agrupados(filtros)
        kpis_globales = _rep.ReporteEncuestaPaginaWeb.obtener_kpis_globales(
            fecha_inicio=kpi_fecha_inicio, fecha_fin=kpi_fecha_fin,
            sitio_evaluado=kpi_sitio_evaluado,
        )
        opciones_nombre = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_pagina_web "
            "WHERE Nombre IS NOT NULL ORDER BY Nombre"
        )
        opciones_sitio = _rep.ReporteEncuestaPaginaWeb.obtener_opciones_sitio()
        raw_timeline   = _rep.ReporteEncuestaPaginaWeb.obtener_timeline()
        timeline_data, heatmap_anios, heatmap_json = build_timeline_heatmap(raw_timeline)
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaPaginaWebTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_experiencia_web.html", {
        "table": table, "filtros": filtros,
        "kpis_globales": kpis_globales,
        "kpi_fecha_inicio": kpi_fecha_inicio, "kpi_fecha_fin": kpi_fecha_fin,
        "kpi_sitio_evaluado": kpi_sitio_evaluado,
        "opciones_nombre": opciones_nombre, "opciones_sitio": opciones_sitio,
        "timeline_data": timeline_data, "heatmap_anios": heatmap_anios, "heatmap_json": heatmap_json,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web_detalle(request, respuesta_id):
    try:
        filas = _rep.ReporteEncuestaPaginaWeb.obtener_detalle(respuesta_id)
    except Exception as e:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    for fila in filas:
        pregunta  = (fila.get("Pregunta") or "").lower()
        respuesta = fila.get("Respuesta", "")
        if "encontró" in pregunta or "encontro" in pregunta:
            if respuesta == "1":
                fila["Respuesta"] = "Sí"
            elif respuesta == "0":
                fila["Respuesta"] = "No"

    encabezado = filas[0] if filas else {}
    encabezado["Nombre"] = encabezado.get("Nombre") or "—"
    fecha = encabezado.get("Fecha")
    encabezado["Fecha"] = fecha.strftime("%d/%m/%Y %H:%M") if fecha else "—"

    promedio_encuesta     = _rep.ReporteEncuestaPaginaWeb.obtener_promedio_encuesta(respuesta_id)
    promedio_encuesta_nps = round((promedio_encuesta / 10) * 100) if promedio_encuesta else 0

    return render(request, "dashboard/encuestas/satisfaccion_experiencia_web_detalle.html", {
        "encabezado": encabezado, "preguntas": filas, "respuesta_id": respuesta_id,
        "kpis": {
            "promedio_encuesta": promedio_encuesta,
            "promedio_pct":      promedio_encuesta_nps,
            "nombre":            encabezado.get("Nombre", ""),
            "fecha":             encabezado.get("Fecha", ""),
        },
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web_exportar(request):
    filtros = {k: request.GET.get(k) for k in ("nombre", "fecha_inicio", "fecha_fin", "sitio_evaluado", "segmento_nps")}
    datos   = _rep.ReporteEncuestaPaginaWeb.obtener_datos(filtros)

    encuestas, preguntas_orden = {}, []
    for fila in datos:
        rid      = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        if rid not in encuestas:
            encuestas[rid] = {"respuesta_id": rid, "Fecha": fila.get("Fecha", ""), "Nombre": fila.get("Nombre", "")}
        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)
        encuestas[rid][pregunta] = fila.get("Respuesta", "")

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Experiencia Página Web"
    cols_fijas = ["ID", "Fecha", "Nombre"]
    keys_fijas = ["respuesta_id", "Fecha", "Nombre"]
    hf = PatternFill("solid", fgColor="003FB7"); hfont = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(cols_fijas, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, p in enumerate(preguntas_orden, len(cols_fijas) + 1):
        c = ws.cell(row=1, column=i, value=p)
        c.fill = PatternFill("solid", fgColor="FFC900"); c.font = Font(bold=True, color="1A1000")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, (rid, enc) in enumerate(encuestas.items(), 2):
        for col_idx, key in enumerate(keys_fijas, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(enc.get(key, "") or ""))
        for i, p in enumerate(preguntas_orden, len(cols_fijas) + 1):
            respuesta = enc.get(p, "")
            if "encontró" in p.lower() or "encontro" in p.lower():
                if str(respuesta) == "1": respuesta = "Sí"
                elif str(respuesta) == "0": respuesta = "No"
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 60)
    ws.row_dimensions[1].height = 60

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="experiencia_pagina_web.xlsx"'
    wb.save(response); return response
