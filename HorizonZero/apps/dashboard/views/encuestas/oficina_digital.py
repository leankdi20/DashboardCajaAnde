# apps/dashboard/views/encuestas/oficina_digital.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports import encuesta_oficina_digital
from apps.dashboard.services.db_service import ReportesDBService
from apps.dashboard.tables import EncuestaOficinaDigitalTable
from apps.dashboard.views._base import build_timeline_heatmap



@login_required
def encuesta_oficina_digital_buscar(request, campo):
    CAMPOS = {
        "agente": "Agente",
        "nombre": "Nombre",
    }
    if campo not in CAMPOS:
        return JsonResponse({"results": []})

    col_db = CAMPOS[campo]
    term   = request.GET.get("term", "").strip()

    sql = f"""
        SELECT DISTINCT TOP 30 {col_db}
        FROM dbo.vw_reporte_encuestas_satisfaccion_oficina_digital
        WHERE {col_db} IS NOT NULL
          AND {col_db} LIKE %s
        ORDER BY {col_db}
    """
    filas = ReportesDBService.ejecutar_query(sql, [f"%{term}%"])
    data  = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db)
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})

@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina")
def encuesta_satisfaccion_oficina(request):
    filtros = {
        "agente":       request.GET.get("agente"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
        "clasificacion": request.GET.get("clasificacion"),
    }
    kpi_sucursales   = request.GET.getlist("kpi_sucursal")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    datos, opciones_agente, opciones_nombre = [], [], []
    kpis_globales = {}
    timeline_data, heatmap_anios, heatmap_json = [], [], "{}"

    try:
        datos         = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_datos_agrupados(filtros)
        kpis_globales = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_kpis_globales(
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
        opciones_agente = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Agente FROM dbo.vw_reporte_encuestas_satisfaccion_oficina_digital "
            "WHERE Agente IS NOT NULL ORDER BY Agente"
        )
        try:
            opciones_nombre = ReportesDBService.ejecutar_query(
                "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_oficina_digital "
                "WHERE Nombre IS NOT NULL ORDER BY Nombre"
            )
        except Exception:
            opciones_nombre = []

        raw_timeline = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_timeline()
        timeline_data, heatmap_anios, heatmap_json = build_timeline_heatmap(raw_timeline)
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaOficinaDigitalTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_of_dig.html", {
        "table":             table,
        "filtros":           filtros,
        "kpis_globales":     kpis_globales,
        "kpi_sucursales":    kpi_sucursales,
        "kpi_fecha_inicio":  kpi_fecha_inicio,
        "kpi_fecha_fin":     kpi_fecha_fin,
        "opciones_agente":   opciones_agente,
        "opciones_nombre":   opciones_nombre,
        "timeline_data":     timeline_data,
        "heatmap_anios":     heatmap_anios,
        "heatmap_json":      heatmap_json,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_detalle_of_dig(request, respuesta_id):
    try:
        filas = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    encabezado = filas[0] if filas else {}

    respuestas_numericas = []
    for fila in filas:
        try:
            respuestas_numericas.append(float(fila.get("Respuesta", 0)))
        except (ValueError, TypeError):
            pass

    satisfechos      = sum(1 for r in respuestas_numericas if r >= 4)
    satisfaccion_pct = round((satisfechos / len(respuestas_numericas) * 100)) if respuestas_numericas else 0

    stats_agente      = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"]  / 3) * 100) if stats_agente["promedio_agente"]  else 0
    promedio_encuesta_pct = round((promedio_encuesta / 3) * 100) if promedio_encuesta else 0

    respuesta_satisfaccion = ""
    for fila in filas:
        if "satisfecho" in (fila.get("Pregunta") or "").lower():
            respuesta_satisfaccion = fila.get("Respuesta", "")
            break

    kpis = {
        "promedio_general":       promedio_agente_pct,
        "total_encuestas":        stats_agente["total_encuestas"],
        "promedio_encuesta":      promedio_encuesta_pct,
        "respuesta_satisfaccion": respuesta_satisfaccion,
        "satisfaccion_pct":       satisfaccion_pct,
    }

    return render(request, "dashboard/encuestas/satisfaccion_detalle_of_dig.html", {
        "encabezado":   encabezado,
        "preguntas":    filas,
        "respuesta_id": respuesta_id,
        "kpis":         kpis,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_of_dig_exportar(request):
    filtros = {
        "agente":       request.GET.get("agente"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }
    datos = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_datos(filtros)

    encuestas, preguntas_orden = {}, []
    for fila in datos:
        rid      = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha":  fila.get("Fecha", ""),
                "Agente": fila.get("Agente", ""),
                "Nombre": fila.get("Nombre", ""),
                "Cedula": fila.get("Cedula", ""),
            }
        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)
        encuestas[rid][pregunta] = fila.get("Respuesta", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Encuesta Satisfacción Oficina Digital"

    cols_fijas = ["ID", "Fecha", "Agente", "Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha", "Agente", "Nombre", "Cedula"]
    hf = PatternFill("solid", fgColor="003FB7")
    hfont = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(cols_fijas, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center")

    for i, pregunta in enumerate(preguntas_orden, len(cols_fijas) + 1):
        c = ws.cell(row=1, column=i, value=pregunta)
        c.fill = PatternFill("solid", fgColor="FFC900")
        c.font = Font(bold=True, color="1A1000")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, (rid, enc) in enumerate(encuestas.items(), 2):
        for col_idx, key in enumerate(keys_fijas, 1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")
        for i, pregunta in enumerate(preguntas_orden, len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    ws.row_dimensions[1].height = 60

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="encuesta_satisfaccion_oficina_digital.xlsx"'
    wb.save(response)
    return response


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_detalle_of_dig_exportar(request, respuesta_id):
    # Mismo layout que satisfaccion_detalle_exportar pero con escala /3
    try:
        filas = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_detalle(respuesta_id)
    except Exception:
        filas = []

    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado        = filas[0]
    stats_agente      = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"]  / 3) * 100) if stats_agente["promedio_agente"]  else 0
    promedio_encuesta_pct = round((promedio_encuesta / 3) * 100) if promedio_encuesta else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Encuesta Satisfacción Oficina Digital"

    def fill(color): return PatternFill("solid", fgColor=color)
    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)
    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    for col, w in zip("ABCDEFG", [2, 26, 2, 35, 20, 20, 2]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7"); c.font = font(bold=True, color="FFFFFF", size=10); c.alignment = centro

    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte de Encuesta de Satisfacción")
    c.font = font(bold=True, color="003FB7", size=15); c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9); c.alignment = izq

    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(bold=True, color="FFC900", size=8); c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11); c.alignment = der

    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    for i, (col, label, val, fgc, fc) in enumerate(zip(
        [4, 5, 6],
        ["PROMEDIO GENERAL AGENTE", "ENCUESTAS DEL AGENTE", "PROMEDIO ESTA ENCUESTA"],
        [f"{promedio_agente_pct}%", str(stats_agente["total_encuestas"]), f"{promedio_encuesta_pct}%"],
        ["003FB7", "E8EEFF", "8B1A0A"],
        ["FFFFFF", "003FB7", "FFFFFF"],
    )):
        c = ws.cell(row=5, column=col, value=label)
        c.fill = fill(fgc); c.font = font(bold=True, color=fc, size=8)
        c.alignment = centro; c.border = border_full(fgc)
        c = ws.cell(row=6, column=col, value=val)
        c.fill = fill(fgc); c.font = font(bold=True, color=fc, size=22)
        c.alignment = centro; c.border = border_full(fgc)

    panel_items = [
        ("AGENTE RESPONSABLE", encabezado.get("Agente", "")),
        ("NOMBRE ACCIONISTA",  encabezado.get("Nombre", "—")),
        ("CÉDULA",             encabezado.get("Cedula", "—")),
        ("FECHA",              str(encabezado.get("Fecha", ""))),
    ]

    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18
    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF"); c.font = font(bold=True, color="003FB7", size=9); c.alignment = izq
    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF"); c.font = font(bold=True, color="003FB7", size=9); c.alignment = centro

    row_data = 9
    for i in range(max(len(panel_items) * 2, len(filas))):
        ws.row_dimensions[row_data + i].height = 22

    for i, (label, valor) in enumerate(panel_items):
        c = ws.cell(row=row_data + (i * 2), column=2, value=label)
        c.fill = fill("FFF0C0"); c.font = font(bold=True, color="003FB7", size=8); c.alignment = izq
        c = ws.cell(row=row_data + (i * 2) + 1, column=2, value=valor)
        c.fill = fill("FFF8E0"); c.font = font(color="1A1A1A", size=10); c.alignment = izq

    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9); c.alignment = izq; c.border = border_full("E0E0E0")

        respuesta  = fila.get("Respuesta", "")
        color_resp = "4CAF50" if respuesta in ["Muy satisfecho", "Muy fácil", "Sí"] else "003FB7"
        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp); c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro; c.border = border_full(color_resp)

    row_pie = row_data + max(len(panel_items) * 2, len(filas)) + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")
    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie + 1, start_column=2, end_row=row_pie + 1, end_column=6)
    c = ws.cell(row=row_pie + 1, column=2, value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A", size=8); c.alignment = centro

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="encuesta_detalle_of_dig{respuesta_id}.xlsx"'
    wb.save(response)
    return response
