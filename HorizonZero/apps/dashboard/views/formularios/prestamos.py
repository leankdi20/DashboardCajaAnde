# apps/dashboard/views/formularios/prestamos.py
# ─────────────────────────────────────────────────────────────────
# Contiene:
#   - formulario_prestamos (hub)
#   - Pre-Solicitud Crédito Personal
# ─────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_presolicitud_credito_personal import ReporteSolicitudPresolicitudCreditoPersonal
from apps.dashboard.tables import SolicitudPresolicitudCreditoPersonalTable


# ── Helpers de formato Excel ─────────────────────────────────────
def _fill(h):
    return PatternFill("solid", fgColor=h)

def _borde():
    return Border(bottom=Side(style="thin", color="E2E8F0"))

def _font(bold=False, color="1E293B", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill      = _fill(cf)
    c.font      = _font(True, ct, 11)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28

def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill      = _fill("F1F5F9")
    lc.font      = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2)
    lc.border    = _borde()

    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font      = _font()
    vc.alignment = Alignment(vertical="center", indent=2)
    vc.border    = _borde()
    ws.row_dimensions[row].height = 22

def _urow(ws, row, label, url):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill      = _fill("F1F5F9")
    lc.font      = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2)
    lc.border    = _borde()

    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=url or "No disponible")
    if url:
        vc.hyperlink = url
        vc.font      = Font(color="003FB7", underline="single", size=10, name="Arial")
    else:
        vc.font      = Font(color="94A3B8", size=10, name="Arial", italic=True)
    vc.alignment = Alignment(vertical="center", indent=2)
    vc.border    = _borde()
    ws.row_dimensions[row].height = 22

def _header_inst(ws, titulo, rid):
    """Encabezado institucional estándar. Retorna la fila donde continuar."""
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill      = _fill("003FB7")
    t.font      = _font(True, "FFFFFF", 13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1,
                value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill      = _fill("002A80")
    s.font      = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8
    return 4  # primera fila de contenido

def _pie(ws, r):
    ws.row_dimensions[r].height = 8
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill      = _fill("F1F5F9")
    pie.font      = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18


# ── Campos para búsqueda AJAX ────────────────────────────────────
CAMPOS_CREDITO_PERSONAL = {"cedula": "Cedula", "nombre": "NombreCompleto"}


# ════════════════════════════════════════════════════════════════
# HUB
# ════════════════════════════════════════════════════════════════
@login_required
@permiso_requerido("dashboard.view_formulario_prestamos")
def formulario_prestamos(request):
    return render(request, "dashboard/formularios/prestamos.html", {"active_tab": "prestamos"})


# ════════════════════════════════════════════════════════════════
# PRE-SOLICITUD CRÉDITO PERSONAL
# ════════════════════════════════════════════════════════════════
@login_required
def soli_presolicitud_credito_personal_buscar(request, campo):
    if campo not in CAMPOS_CREDITO_PERSONAL:
        return JsonResponse({"results": []})

    col_db = CAMPOS_CREDITO_PERSONAL[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudPresolicitudCreditoPersonal.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas
        if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_lista(request):
    filtros = {
        k: request.GET.get(k, "").strip()
        for k in ("cedula", "nombre", "tipo_credito", "sucursal", "fecha_inicio", "fecha_fin")
    }
    filtros = {k: v for k, v in filtros.items() if v}

    datos         = ReporteSolicitudPresolicitudCreditoPersonal.obtener_datos(filtros)
    kpis          = ReporteSolicitudPresolicitudCreditoPersonal.obtener_kpis(filtros)
    tipos_credito = ReporteSolicitudPresolicitudCreditoPersonal.obtener_tipos_credito()
    sucursales    = ReporteSolicitudPresolicitudCreditoPersonal.obtener_sucursales()

    table = SolicitudPresolicitudCreditoPersonalTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)

    return render(request, "dashboard/formularios/reportes/soli_presolicitud_credito_personal.html", {
        "table":         table,
        "filtros":       filtros,
        "kpis":          kpis,
        "tipos_credito": tipos_credito,
        "sucursales":    sucursales,
    })


@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPresolicitudCreditoPersonal.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(
        request,
        "dashboard/formularios/reportes/soli_presolicitud_credito_personal_detalle.html",
        {"solicitud": solicitud},
    )


@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_export(request):
    filtros = {
        k: request.GET.get(k, "").strip()
        for k in ("cedula", "nombre", "tipo_credito", "sucursal", "fecha_inicio", "fecha_fin")
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudPresolicitudCreditoPersonal.obtener_datos(filtros)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Solicitud Crédito"

    headers = [
        "ID", "Fecha / Hora", "Cédula", "Nombre", "Teléfono",
        "Tipo de Crédito", "Sucursal", "Monto (₡)",
        "URL Cédula Frente", "URL Cédula Reverso", "URL Desglose Pensión",
    ]
    hf    = _fill("003FB7")
    hfont = Font(bold=True, color="FFFFFF", name="Arial")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill      = hf
        c.font      = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        monto = row.get("Monto")

        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("NombreCompleto") or "").strip())
        ws.cell(row=ri, column=5,  value=(row.get("Telefono") or "").strip())
        ws.cell(row=ri, column=6,  value=row.get("TipoCredito"))
        ws.cell(row=ri, column=7,  value=row.get("SucursalFormalizacion"))
        c_m = ws.cell(row=ri, column=8,  value=int(float(monto)) if monto else None)
        if monto:
            c_m.number_format = "₡#,##0"
        ws.cell(row=ri, column=9,  value=row.get("FrenteCedula_URL"))
        ws.cell(row=ri, column=10, value=row.get("ReversoCedula_URL"))
        ws.cell(row=ri, column=11, value=row.get("DesglosePension_URL"))

        if ri % 2 == 0:
            ff = _fill("E8EFFE")
            for c in range(1, 12):
                ws.cell(row=ri, column=c).fill = ff

    for i, w in enumerate([8, 18, 14, 35, 14, 42, 20, 14, 55, 55, 55], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="presolicitud_credito_{fstr}.xlsx"'
    wb.save(response)
    return response


@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPresolicitudCreditoPersonal.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404

    wb = Workbook()
    ws = wb.active
    ws.title = f"Crédito #{respuesta_id}"

    for col, w in enumerate([5, 28, 50, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    r = _header_inst(ws, "CAJA DE ANDE — Pre-Solicitud Crédito Personal", solicitud["respuesta_id"])

    fecha     = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")

    _hrow(ws, r, "DATOS DEL ACCIONISTA", "003FB7"); r += 1
    _drow(ws, r, "CÉDULA",    (solicitud.get("Cedula") or "").strip()); r += 1
    _drow(ws, r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r += 1
    _drow(ws, r, "TELÉFONO",  (solicitud.get("Telefono") or "").strip()); r += 1
    _drow(ws, r, "FECHA / HORA", fecha_str); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DATOS DEL CRÉDITO", "FFC900", "1E293B"); r += 1
    _drow(ws, r, "TIPO DE CRÉDITO", solicitud.get("TipoCredito")); r += 1
    _drow(ws, r, "SUCURSAL",        solicitud.get("SucursalFormalizacion")); r += 1

    # Monto con formato monetario
    monto = solicitud.get("Monto")
    ws.merge_cells(f"A{r}:B{r}")
    lc = ws.cell(row=r, column=1, value="MONTO")
    lc.fill      = _fill("F1F5F9")
    lc.font      = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2)
    lc.border    = _borde()
    ws.merge_cells(f"C{r}:D{r}")
    vc = ws.cell(row=r, column=3, value=int(float(monto)) if monto else "—")
    if monto:
        vc.number_format = "₡#,##0"
    vc.font      = _font(True)
    vc.alignment = Alignment(vertical="center", indent=2)
    vc.border    = _borde()
    ws.row_dimensions[r].height = 22
    r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DOCUMENTOS ADJUNTOS — SHAREFILE", "FFC900", "1E293B"); r += 1
    _urow(ws, r, "CÉDULA — FRENTE",     solicitud.get("FrenteCedula_URL")); r += 1
    _urow(ws, r, "CÉDULA — REVERSO",    solicitud.get("ReversoCedula_URL")); r += 1
    _urow(ws, r, "DESGLOSE DE PENSIÓN", solicitud.get("DesglosePension_URL")); r += 1

    _pie(ws, r)
    ws.freeze_panes      = "A3"
    ws.sheet_view.zoomScale = 110

    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="credito_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response