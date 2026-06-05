# apps/dashboard/views/formularios/ahorros.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables

from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_deposito_salario import ReporteSolicitudDepositoSalario
from apps.dashboard.reports.soli_ahorro_mod_cuota import ReporteSolicitudAhorroModCuota
from apps.dashboard.reports.soli_reinversion_ahorro import ReporteSolicitudReinversionAhorro
from apps.dashboard.reports.soli_autorizacion_ahorro_nuevo import ReporteSolicitudAutorizacionAhorroNuevo
from apps.dashboard.tables import (
    SolicitudDepositoSalarioTable,
    SolicitudAhorroModCuotaTable,
    SolicitudReinversionAhorroTable,
    SolicitudAutorizacionAhorroNuevoTable,
)


def _fill(h): return PatternFill("solid", fgColor=h)
def _borde(c="E2E8F0"): return Border(bottom=Side(style="thin", color=c))
def _font(bold=False, color="1E293B", size=10): return Font(bold=bold, color=color, size=size, name="Arial")

def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill = _fill(cf); c.font = _font(True, ct, 11)
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28

def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font = _font(); vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22

def _urow(ws, row, label, url):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=url or "No disponible")
    if url:
        vc.hyperlink = url
        vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
    else:
        vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22

def _header_inst(ws, titulo, rid):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = _fill("003FB7"); t.font = _font(True, "FFFFFF", 13)
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1,
                value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8
    return 4

def _pie(ws, r):
    ws.row_dimensions[r].height = 8; r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = _fill("F1F5F9"); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110


# ════════════════════════════════════════════════════════════════
# HUB
# ════════════════════════════════════════════════════════════════
@login_required
@permiso_requerido("dashboard.view_formulario_ahorros")
def formulario_ahorros(request):
    return render(request, "dashboard/formularios/ahorros.html", {"active_tab": "ahorros"})


# ════════════════════════════════════════════════════════════════
# DEPÓSITO DE SALARIO
# ════════════════════════════════════════════════════════════════
@login_required
def soli_deposito_salario_buscar_cedula(request):
    term  = request.GET.get("term", "").strip()
    filas = ReporteSolicitudDepositoSalario.buscar_cedulas(term)
    data  = [
        {"id": (r["Cedula"] or "").strip(), "text": (r["Cedula"] or "").strip()}
        for r in filas if r.get("Cedula") and r["Cedula"].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_lista(request):
    filtros = {k: request.GET.get(k, "").strip() for k in ("cedula", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteSolicitudDepositoSalario.obtener_datos(filtros)
    kpis  = ReporteSolicitudDepositoSalario.obtener_kpis(filtros)
    table = SolicitudDepositoSalarioTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_deposito_salario.html",
                  {"table": table, "filtros": filtros, "kpis": kpis})


@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudDepositoSalario.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request, "dashboard/formularios/reportes/soli_deposito_salario_detalle.html",
                  {"solicitud": solicitud})


@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_export(request):
    filtros = {k: request.GET.get(k, "").strip() for k in ("cedula", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudDepositoSalario.obtener_datos(filtros)

    wb = Workbook(); ws = wb.active; ws.title = "Depósito de Salario"
    headers = ["ID", "Fecha / Hora", "Cédula",
               "URL Boleta Solicitud", "URL Cédula Frente", "URL Cédula Reverso"]
    hf = _fill("003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4, value=row.get("BoletaSolicitud_URL"))
        ws.cell(row=ri, column=5, value=row.get("FrenteCedula_URL"))
        ws.cell(row=ri, column=6, value=row.get("ReversoCedula_URL"))
        if ri % 2 == 0:
            ff = _fill("E8EFFE")
            for c in range(1, 7): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 18, 14, 55, 55, 55], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="deposito_salario_{fstr}.xlsx"'
    wb.save(response); return response


@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudDepositoSalario.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404

    wb = Workbook(); ws = wb.active; ws.title = f"Depósito Salario #{respuesta_id}"
    for col, w in enumerate([5, 25, 55, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    r = _header_inst(ws, "CAJA DE ANDE — Solicitud Depósito de Salario", respuesta_id)
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")

    _hrow(ws, r, "DATOS DEL SOLICITANTE", "003FB7"); r += 1
    _drow(ws, r, "CÉDULA",      (solicitud.get("Cedula") or "").strip()); r += 1
    _drow(ws, r, "FECHA / HORA", fecha_str); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DOCUMENTOS ADJUNTOS — SHAREFILE", "FFC900", "1E293B"); r += 1
    _urow(ws, r, "BOLETA DE SOLICITUD", solicitud.get("BoletaSolicitud_URL")); r += 1
    _urow(ws, r, "CÉDULA — FRENTE",     solicitud.get("FrenteCedula_URL")); r += 1
    _urow(ws, r, "CÉDULA — REVERSO",    solicitud.get("ReversoCedula_URL")); r += 1

    _pie(ws, r)
    cedula = (solicitud.get("Cedula") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="deposito_salario_{respuesta_id}_{cedula}.xlsx"'
    wb.save(response); return response


# ════════════════════════════════════════════════════════════════
# MODIFICACIÓN DE CUOTA
# ════════════════════════════════════════════════════════════════
CAMPOS_AHORRO_MOD = {"cedula": "Cedula", "nombre": "Nombre_Completo"}

@login_required
def soli_ahorro_mod_cuota_buscar(request, campo):
    if campo not in CAMPOS_AHORRO_MOD: return JsonResponse({"results": []})
    col_db = CAMPOS_AHORRO_MOD[campo]; term = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudAhorroModCuota.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
                                     for r in filas if r.get(col_db) and r[col_db].strip()],
                         "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_lista(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "tipo_modificacion", "numero_contrato", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos              = ReporteSolicitudAhorroModCuota.obtener_datos(filtros)
    kpis               = ReporteSolicitudAhorroModCuota.obtener_kpis(filtros)
    tipos_ahorro       = ReporteSolicitudAhorroModCuota.obtener_tipos_ahorro()
    tipos_modificacion = ReporteSolicitudAhorroModCuota.obtener_tipos_modificacion()
    table = SolicitudAhorroModCuotaTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_ahorro_mod_cuota.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
        "tipos_ahorro": tipos_ahorro, "tipos_modificacion": tipos_modificacion,
    })


@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAhorroModCuota.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request, "dashboard/formularios/reportes/soli_ahorro_mod_cuota_detalle.html",
                  {"solicitud": solicitud})


@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_export(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "tipo_modificacion", "numero_contrato", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudAhorroModCuota.obtener_datos(filtros)

    wb = Workbook(); ws = wb.active; ws.title = "Modificación Cuota Ahorro"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Tipo de Ahorro", "N° Contrato", "Modificación", "Monto (₡)"]
    hf = _fill("003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        monto = row.get("MontoCuotaDeducir")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=5, value=row.get("TipoAhorro"))
        ws.cell(row=ri, column=6, value=row.get("NumeroContrato"))
        ws.cell(row=ri, column=7, value=row.get("TipoModificacion"))
        c_m = ws.cell(row=ri, column=8, value=int(float(monto)) if monto else 0)
        c_m.number_format = "₡#,##0"
        if ri % 2 == 0:
            ff = _fill("E8EFFE")
            for c in range(1, 9): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 18, 14, 30, 25, 15, 14, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="ahorro_mod_cuota_{fstr}.xlsx"'
    wb.save(response); return response


@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAhorroModCuota.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404

    wb = Workbook(); ws = wb.active; ws.title = f"Mod. Cuota #{respuesta_id}"
    for col, w in enumerate([5, 25, 35, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    r = _header_inst(ws, "CAJA DE ANDE — Solicitud Modificación de Cuota de Ahorro", respuesta_id)
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")

    _hrow(ws, r, "DATOS DEL ACCIONISTA", "003FB7"); r += 1
    _drow(ws, r, "CÉDULA",      (solicitud.get("Cedula") or "").strip()); r += 1
    _drow(ws, r, "NOMBRE",      (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    _drow(ws, r, "FECHA / HORA", fecha_str); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DATOS DE LA MODIFICACIÓN", "FFC900", "1E293B"); r += 1
    _drow(ws, r, "TIPO DE AHORRO",    solicitud.get("TipoAhorro")); r += 1
    _drow(ws, r, "N° DE CONTRATO",    solicitud.get("NumeroContrato")); r += 1
    _drow(ws, r, "TIPO MODIFICACIÓN", solicitud.get("TipoModificacion")); r += 1

    monto = solicitud.get("MontoCuotaDeducir")
    ws.merge_cells(f"A{r}:B{r}")
    lc = ws.cell(row=r, column=1, value="MONTO CUOTA A DEDUCIR")
    lc.fill = _fill("F1F5F9"); lc.font = _font(True, "64748B", 9)
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{r}:D{r}")
    vc = ws.cell(row=r, column=3, value=int(float(monto)) if monto else 0)
    vc.font = _font(True); vc.number_format = "₡#,##0"
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[r].height = 22; r += 1

    _pie(ws, r)
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="mod_cuota_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ════════════════════════════════════════════════════════════════
# REINVERSIÓN AHORRO
# ════════════════════════════════════════════════════════════════
CAMPOS_REINVERSION = {"cedula": "Cedula", "nombre": "Nombre_Completo"}

@login_required
def soli_reinversion_ahorro_buscar(request, campo):
    if campo not in CAMPOS_REINVERSION: return JsonResponse({"results": []})
    col_db = CAMPOS_REINVERSION[campo]; term = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudReinversionAhorro.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
                                     for r in filas if r.get(col_db) and r[col_db].strip()],
                         "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_lista(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "tipo_reinversion", "numero_contrato", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos             = ReporteSolicitudReinversionAhorro.obtener_datos(filtros)
    kpis              = ReporteSolicitudReinversionAhorro.obtener_kpis(filtros)
    tipos_ahorro      = ReporteSolicitudReinversionAhorro.obtener_tipos_ahorro()
    tipos_reinversion = ReporteSolicitudReinversionAhorro.obtener_tipos_reinversion()
    table = SolicitudReinversionAhorroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_reinversion_ahorro.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
        "tipos_ahorro": tipos_ahorro, "tipos_reinversion": tipos_reinversion,
    })


@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudReinversionAhorro.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request, "dashboard/formularios/reportes/soli_reinversion_ahorro_detalle.html",
                  {"solicitud": solicitud})


@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_export(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "tipo_reinversion", "numero_contrato", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudReinversionAhorro.obtener_datos(filtros)

    wb = Workbook(); ws = wb.active; ws.title = "Reinversión Ahorro"
    headers = ["ID", "Fecha", "Hora", "Cédula", "Nombre",
               "Tipo de Ahorro", "Tipo Reinversión", "N° Contrato", "Reinversión 2"]
    hf = _fill("003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("Fecha"); hora = row.get("Hora")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=str(fecha) if fecha else "")
        ws.cell(row=ri, column=3, value=str(hora) if hora else "")
        ws.cell(row=ri, column=4, value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=6, value=row.get("SolicitoReinversion"))
        ws.cell(row=ri, column=7, value=row.get("TipoReinversion"))
        ws.cell(row=ri, column=8, value=row.get("NumeroContrato"))
        ws.cell(row=ri, column=9, value=row.get("TipoReinversion2"))
        if ri % 2 == 0:
            ff = _fill("E8EFFE")
            for c in range(1, 10): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 12, 10, 14, 30, 20, 30, 15, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reinversion_ahorro_{fstr}.xlsx"'
    wb.save(response); return response


@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudReinversionAhorro.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404

    wb = Workbook(); ws = wb.active; ws.title = f"Reinversión #{respuesta_id}"
    for col, w in enumerate([5, 28, 35, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    r = _header_inst(ws, "CAJA DE ANDE — Solicitud Reinversión Ahorro Existente", respuesta_id)
    fecha = solicitud.get("Fecha"); hora = solicitud.get("Hora")

    _hrow(ws, r, "DATOS DEL ACCIONISTA", "003FB7"); r += 1
    _drow(ws, r, "CÉDULA",  (solicitud.get("Cedula") or "").strip()); r += 1
    _drow(ws, r, "NOMBRE",  (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    _drow(ws, r, "FECHA",   str(fecha) if fecha else "—"); r += 1
    _drow(ws, r, "HORA",    str(hora) if hora else "—"); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DATOS DE LA REINVERSIÓN", "FFC900", "1E293B"); r += 1
    _drow(ws, r, "TIPO DE AHORRO",   solicitud.get("SolicitoReinversion")); r += 1
    _drow(ws, r, "N° DE CONTRATO",   solicitud.get("NumeroContrato")); r += 1
    _drow(ws, r, "TIPO REINVERSIÓN", solicitud.get("TipoReinversion")); r += 1
    if solicitud.get("TipoReinversion2"):
        _drow(ws, r, "REINVERSIÓN CUOTA", solicitud.get("TipoReinversion2")); r += 1

    _pie(ws, r)
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reinversion_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ════════════════════════════════════════════════════════════════
# AUTORIZACIÓN AHORRO NUEVO
# ════════════════════════════════════════════════════════════════
CAMPOS_AUTORIZACION_AHORRO = {"cedula": "Cedula", "nombre": "Nombre_Completo"}

@login_required
def soli_autorizacion_ahorro_nuevo_buscar(request, campo):
    if campo not in CAMPOS_AUTORIZACION_AHORRO: return JsonResponse({"results": []})
    col_db = CAMPOS_AUTORIZACION_AHORRO[campo]; term = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudAutorizacionAhorroNuevo.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
                                     for r in filas if r.get(col_db) and r[col_db].strip()],
                         "pagination": {"more": False}})


@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_lista(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "forma_pago", "tipo_reinversion",
                "retiro_intereses", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos             = ReporteSolicitudAutorizacionAhorroNuevo.obtener_datos(filtros)
    kpis              = ReporteSolicitudAutorizacionAhorroNuevo.obtener_kpis(filtros)
    tipos_ahorro      = ReporteSolicitudAutorizacionAhorroNuevo.obtener_tipos_ahorro()
    formas_pago       = ReporteSolicitudAutorizacionAhorroNuevo.obtener_formas_pago()
    tipos_reinversion = ReporteSolicitudAutorizacionAhorroNuevo.obtener_tipos_reinversion()
    table = SolicitudAutorizacionAhorroNuevoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_autorizacion_ahorro_nuevo.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
        "tipos_ahorro": tipos_ahorro, "formas_pago": formas_pago,
        "tipos_reinversion": tipos_reinversion,
    })


@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAutorizacionAhorroNuevo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request, "dashboard/formularios/reportes/soli_autorizacion_ahorro_nuevo_detalle.html",
                  {"solicitud": solicitud})


@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_export(request):
    filtros = {k: request.GET.get(k, "").strip() for k in
               ("cedula", "nombre", "tipo_ahorro", "forma_pago", "tipo_reinversion",
                "retiro_intereses", "fecha_inicio", "fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudAutorizacionAhorroNuevo.obtener_datos(filtros)

    wb = Workbook(); ws = wb.active; ws.title = "Autorización Ahorro Nuevo"
    headers = ["ID", "Fecha", "Cédula", "Nombre", "Tipo de Ahorro",
               "Forma de Pago", "Tipo Reinversión", "Retiro Intereses",
               "Monto Cuota (₡)", "Monto Débito Ahorro (₡)"]
    hf = _fill("003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("Fecha"); mc = row.get("MontoCuota"); mda = row.get("MontoDebitarAhorro")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=str(fecha) if fecha else "")
        ws.cell(row=ri, column=3,  value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=5,  value=row.get("TipoAhorro"))
        ws.cell(row=ri, column=6,  value=row.get("FormaPago"))
        ws.cell(row=ri, column=7,  value=row.get("TipoReinversion"))
        ws.cell(row=ri, column=8,  value=row.get("RetiroIntereses"))
        c_mc  = ws.cell(row=ri, column=9,  value=int(float(mc))  if mc  else None)
        c_mda = ws.cell(row=ri, column=10, value=int(float(mda)) if mda else None)
        if mc:  c_mc.number_format  = "₡#,##0"
        if mda: c_mda.number_format = "₡#,##0"
        if ri % 2 == 0:
            ff = _fill("E8EFFE")
            for c in range(1, 11): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 12, 14, 30, 22, 15, 30, 15, 16, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="autorizacion_ahorro_nuevo_{fstr}.xlsx"'
    wb.save(response); return response


@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAutorizacionAhorroNuevo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404

    wb = Workbook(); ws = wb.active; ws.title = f"Ahorro Nuevo #{respuesta_id}"
    for col, w in enumerate([5, 28, 35, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    r = _header_inst(ws, "CAJA DE ANDE — Solicitud Autorización Ahorro Nuevo", respuesta_id)
    fecha = solicitud.get("Fecha")

    _hrow(ws, r, "DATOS DEL ACCIONISTA", "003FB7"); r += 1
    _drow(ws, r, "CÉDULA",  (solicitud.get("Cedula") or "").strip()); r += 1
    _drow(ws, r, "NOMBRE",  (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    _drow(ws, r, "FECHA",   str(fecha) if fecha else "—"); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "DATOS DEL AHORRO", "FFC900", "1E293B"); r += 1
    _drow(ws, r, "TIPO DE AHORRO",      solicitud.get("TipoAhorro")); r += 1
    _drow(ws, r, "FORMA DE PAGO",       solicitud.get("FormaPago")); r += 1
    _drow(ws, r, "TIPO REINVERSIÓN",    solicitud.get("TipoReinversion")); r += 1
    _drow(ws, r, "RETIRO DE INTERESES", solicitud.get("RetiroIntereses")); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    _hrow(ws, r, "MONTOS", "003FB7"); r += 1
    for label, key in [("MONTO CUOTA", "MontoCuota"), ("MONTO DÉBITO AHORRO", "MontoDebitarAhorro")]:
        monto = solicitud.get(key)
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = _fill("F1F5F9"); lc.font = _font(True, "64748B", 9)
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=int(float(monto)) if monto else "—")
        if monto: vc.number_format = "₡#,##0"
        vc.font = _font(True); vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
        ws.row_dimensions[r].height = 22; r += 1

    _pie(ws, r)
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="ahorro_nuevo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response