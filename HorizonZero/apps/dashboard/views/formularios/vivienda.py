# apps/dashboard/views/formularios/vivienda.py
# Compra Vehículo · Préstamo Vivienda · Préstamo Desarrollo Económico
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables
from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_compra_vehiculo import ReporteSolicitudCompraVehiculo
from apps.dashboard.reports.soli_prestamo_vivienda import ReporteSolicitudPrestamoVivienda
from apps.dashboard.reports.soli_prestamo_desarrollo import ReporteSolicitudPrestamoDesarrollo
from apps.dashboard.tables import SolicitudCompraVehiculoTable, SolicitudPrestamoViviendaTable, SolicitudPrestamoDesarrolloTable

CAMPOS_VEHICULO = {"cedula": "Cedula", "nombre": "NombreCompleto"}
CAMPOS_PRESTAMO_VIVIENDA = {"cedula": "Cedula", "nombre": "NombreCompleto", "telefono": "Telefono"}
CAMPOS_DESARROLLO = {"cedula": "Cedula", "nombre": "NombreCompleto", "telefono": "Telefono"}


def _fill(h): return PatternFill("solid", fgColor=h)
def _borde(): return Border(bottom=Side(style="thin", color="E2E8F0"))
def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill = _fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font = Font(color="1E293B", size=10, name="Arial")
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _header_inst(ws, titulo, rid):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = _fill("003FB7"); t.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1, value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8; return 4
def _pie(ws, r):
    ws.row_dimensions[r].height = 8; r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = _fill("F1F5F9"); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18


@login_required
@permiso_requerido("dashboard.view_formulario_vivienda")
def formulario_vivienda(request):
    return render(request, "dashboard/formularios/vivienda.html", {"active_tab": "vivienda"})


# ── Compra Vehículo ──────────────────────────────────────────────
@login_required
def soli_compra_vehiculo_buscar(request, campo):
    if campo not in CAMPOS_VEHICULO: return JsonResponse({"results": []})
    col_db = CAMPOS_VEHICULO[campo]; term = request.GET.get("term", "").strip()
    filas = ReporteSolicitudCompraVehiculo.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()], "pagination": {"more": False}})

@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_lista(request):
    filtros = {k: request.GET.get(k, "").strip() for k in ("cedula","nombre","tipo_vehiculo","gastos_formalizacion","provincia_domicilio","fecha_inicio","fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteSolicitudCompraVehiculo.obtener_datos(filtros)
    kpis = ReporteSolicitudCompraVehiculo.obtener_kpis(filtros)
    tipos = ReporteSolicitudCompraVehiculo.obtener_tipos_vehiculo()
    table = SolicitudCompraVehiculoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_compra_vehiculo.html",
                  {"table": table, "filtros": filtros, "kpis": kpis, "tipos_vehiculo": tipos})

@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudCompraVehiculo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request, "dashboard/formularios/reportes/soli_compra_vehiculo_detalle.html", {"solicitud": solicitud})

@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_export(request):
    filtros = {k: request.GET.get(k, "").strip() for k in ("cedula","nombre","tipo_vehiculo","gastos_formalizacion","provincia_domicilio","fecha_inicio","fecha_fin")}
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteSolicitudCompraVehiculo.obtener_datos(filtros)
    wb = Workbook(); ws = wb.active; ws.title = "Solicitudes Vehículo"
    headers = ["ID","Fecha/Hora","Cédula","Nombre","Correo","Fecha Nac.","Estado Civil","Tel. Celular","Tel. Casa","Tel. Trabajo","Nacionalidad","Dir. Domicilio","Prov. Domicilio","Cantón Dom.","Distrito Dom.","Dir. Trabajo","Prov. Trabajo","Cantón Trab.","Distrito Trab.","Monto Solicitado","Plazo","Gastos Formaliz.","Fecha Entrega","Tipo Vehículo","Garantía"]
    hf = PatternFill("solid", fgColor="003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora"); fstr = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        fn = row.get("FechaNacimiento"); monto = row.get("MontoCreditoSolicitado")
        vals = [row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),row.get("Correo"),str(fn) if fn else "",row.get("EstadoCivil"),row.get("TelefonoCelular"),row.get("TelefonoCasa"),row.get("TelefonoTrabajo"),row.get("Nacionalidad"),row.get("DireccionDomicilio"),row.get("ProvinciaDomicilio"),row.get("CantonDomicilio"),row.get("DistritoDomicilio"),row.get("DireccionTrabajo"),row.get("ProvinciaTrabajo"),row.get("CantonTrabajo"),row.get("DistritoTrabajo"),None,row.get("Plazo"),row.get("GastosFormalizacion"),str(row.get("FechaEntrega")) if row.get("FechaEntrega") else "",row.get("TipoVehiculo"),row.get("Garantia")]
        for col, v in enumerate(vals, 1): ws.cell(row=ri, column=col, value=v)
        cm = ws.cell(row=ri, column=20, value=int(float(monto)) if monto else None)
        if monto: cm.number_format = '₡#,##0'
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 26): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8,18,14,30,28,12,14,14,14,14,14,30,16,16,16,30,16,16,16,16,8,12,12,12,30], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="solicitudes_vehiculo_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudCompraVehiculo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb = Workbook(); ws = wb.active; ws.title = f"Vehículo #{respuesta_id}"
    for col, w in enumerate([5,20,15,15,20,15], 1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = w
    # header con 6 columnas
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="CAJA DE ANDE — Solicitud Préstamo Compra de Vehículo")
    t.fill = _fill("003FB7"); t.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:F2")
    s = ws.cell(row=2, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8; r = 4
    fecha = solicitud.get("FechaHora"); fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    def hrow6(row, texto, cf, ct="FFFFFF"):
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = _fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
    def drow6(row, label, valor):
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
        ws.merge_cells(f"D{row}:F{row}")
        vc = ws.cell(row=row, column=4, value=valor if valor is not None else "—")
        vc.font = Font(color="1E293B", size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
        ws.row_dimensions[row].height = 22
    hrow6(r,"DATOS PERSONALES","003FB7"); r+=1
    for lbl,val in [("CÉDULA",(solicitud.get("Cedula") or "").strip()),("NOMBRE",(solicitud.get("NombreCompleto") or "").strip()),("CORREO",solicitud.get("Correo")),("FECHA NAC.",str(solicitud.get("FechaNacimiento")) if solicitud.get("FechaNacimiento") else "—"),("ESTADO CIVIL",solicitud.get("EstadoCivil")),("NACIONALIDAD",solicitud.get("Nacionalidad")),("TEL. CELULAR",solicitud.get("TelefonoCelular")),("TEL. CASA",solicitud.get("TelefonoCasa")),("TEL. TRABAJO",solicitud.get("TelefonoTrabajo"))]:
        drow6(r,lbl,val); r+=1
    ws.row_dimensions[r].height=8; r+=1
    hrow6(r,"DOMICILIO","FFC900","1E293B"); r+=1
    for lbl,val in [("DIRECCIÓN",solicitud.get("DireccionDomicilio")),("PROVINCIA",solicitud.get("ProvinciaDomicilio")),("CANTÓN",solicitud.get("CantonDomicilio")),("DISTRITO",solicitud.get("DistritoDomicilio"))]:
        drow6(r,lbl,val); r+=1
    ws.row_dimensions[r].height=8; r+=1
    hrow6(r,"LUGAR DE TRABAJO","003FB7"); r+=1
    for lbl,val in [("DIRECCIÓN",solicitud.get("DireccionTrabajo")),("PROVINCIA",solicitud.get("ProvinciaTrabajo")),("CANTÓN",solicitud.get("CantonTrabajo")),("DISTRITO",solicitud.get("DistritoTrabajo"))]:
        drow6(r,lbl,val); r+=1
    ws.row_dimensions[r].height=8; r+=1
    hrow6(r,"DATOS DEL PRÉSTAMO","FFC900","1E293B"); r+=1
    monto = solicitud.get("MontoCreditoSolicitado")
    ws.merge_cells(f"A{r}:C{r}"); lc=ws.cell(row=r,column=1,value="MONTO SOLICITADO"); lc.fill=_fill("F1F5F9"); lc.font=Font(bold=True,color="64748B",size=9,name="Arial"); lc.alignment=Alignment(vertical="center",indent=2); lc.border=_borde()
    ws.merge_cells(f"D{r}:F{r}"); vc=ws.cell(row=r,column=4,value=int(float(monto)) if monto else "—")
    if monto: vc.number_format='₡#,##0'
    vc.font=Font(color="1E293B",size=10,name="Arial",bold=True); vc.alignment=Alignment(vertical="center",indent=2); vc.border=_borde(); ws.row_dimensions[r].height=22; r+=1
    for lbl,val in [("PLAZO",solicitud.get("Plazo")),("GASTOS FORMALIZ.",solicitud.get("GastosFormalizacion")),("FECHA ENTREGA",str(solicitud.get("FechaEntrega")) if solicitud.get("FechaEntrega") else "—"),("TIPO DE VEHÍCULO",solicitud.get("TipoVehiculo")),("GARANTÍA",solicitud.get("Garantia"))]:
        drow6(r,lbl,val); r+=1
    ws.row_dimensions[r].height=8; r+=1
    ws.merge_cells(f"A{r}:F{r}"); pie=ws.cell(row=r,column=1,value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill=_fill("F1F5F9"); pie.font=Font(color="94A3B8",size=8,name="Arial",italic=True); pie.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[r].height=18
    ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="vehiculo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ── Préstamo Vivienda ────────────────────────────────────────────
@login_required
def soli_prestamo_vivienda_buscar(request, campo):
    if campo not in CAMPOS_PRESTAMO_VIVIENDA: return JsonResponse({"results": []})
    col_db = CAMPOS_PRESTAMO_VIVIENDA[campo]; term = request.GET.get("term", "").strip()
    filas = ReporteSolicitudPrestamoVivienda.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()], "pagination": {"more": False}})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_lista(request):
    filtros = {k: request.GET.get(k,"").strip() for k in ("cedula","nombre","telefono","tipo_prestamo","fecha_inicio","fecha_fin")}
    filtros = {k:v for k,v in filtros.items() if v}
    datos=ReporteSolicitudPrestamoVivienda.obtener_datos(filtros); kpis=ReporteSolicitudPrestamoVivienda.obtener_kpis(filtros)
    tipos=ReporteSolicitudPrestamoVivienda.obtener_tipos_prestamo(); table=SolicitudPrestamoViviendaTable(datos)
    tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_prestamo_vivienda.html",
                  {"table":table,"filtros":filtros,"kpis":kpis,"tipos_prestamo":tipos})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudPrestamoVivienda.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request,"dashboard/formularios/reportes/soli_prestamo_vivienda_detalle.html",{"solicitud":solicitud})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","telefono","tipo_prestamo","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteSolicitudPrestamoVivienda.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Préstamos Vivienda"
    headers=["ID","Fecha / Hora","Cédula","Nombre","Teléfono","Tipo de Préstamo"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        fecha=row.get("FechaHora"); fstr=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
        for col,v in enumerate([row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),(row.get("Telefono") or "").strip(),row.get("TipoPrestamo")],1): ws.cell(row=ri,column=col,value=v)
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,7): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,18,14,35,14,40],1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="prestamo_vivienda_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_export_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudPrestamoVivienda.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"Vivienda #{respuesta_id}"
    for col,w in enumerate([5,25,40,15],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Solicitud Préstamo para Vivienda",solicitud["respuesta_id"])
    fecha=solicitud.get("FechaHora"); fecha_str=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"TELÉFONO",(solicitud.get("Telefono") or "").strip()); r+=1
    _drow(ws,r,"FECHA / HORA",fecha_str); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL PRÉSTAMO","FFC900","1E293B"); r+=1
    _drow(ws,r,"TIPO DE PRÉSTAMO",solicitud.get("TipoPrestamo")); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="vivienda_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ── Préstamo Desarrollo Económico ────────────────────────────────
@login_required
def soli_prestamo_desarrollo_buscar(request, campo):
    if campo not in CAMPOS_DESARROLLO: return JsonResponse({"results": []})
    col_db=CAMPOS_DESARROLLO[campo]; term=request.GET.get("term","").strip()
    filas=ReporteSolicitudPrestamoDesarrollo.buscar_opciones(col_db,term)
    return JsonResponse({"results":[{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()],"pagination":{"more":False}})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_lista(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","telefono","plan_inversion","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}
    datos=ReporteSolicitudPrestamoDesarrollo.obtener_datos(filtros); kpis=ReporteSolicitudPrestamoDesarrollo.obtener_kpis(filtros)
    planes=ReporteSolicitudPrestamoDesarrollo.obtener_planes(); table=SolicitudPrestamoDesarrolloTable(datos)
    tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_prestamo_desarrollo.html",
                  {"table":table,"filtros":filtros,"kpis":kpis,"planes":planes})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudPrestamoDesarrollo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request,"dashboard/formularios/reportes/soli_prestamo_desarrollo_detalle.html",{"solicitud":solicitud})

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","telefono","plan_inversion","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteSolicitudPrestamoDesarrollo.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Desarrollo Económico"
    headers=["ID","Fecha / Hora","Cédula","Nombre","Teléfono","Plan de Inversión"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        fecha=row.get("FechaHora"); fstr=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
        for col,v in enumerate([row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),(row.get("Telefono") or "").strip(),row.get("PlanInversion")],1): ws.cell(row=ri,column=col,value=v)
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,7): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,18,14,35,14,20],1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="prestamo_desarrollo_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_export_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudPrestamoDesarrollo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"Desarrollo #{respuesta_id}"
    for col,w in enumerate([5,25,35,15],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Solicitud Préstamo Desarrollo Económico",solicitud["respuesta_id"])
    fecha=solicitud.get("FechaHora"); fecha_str=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"TELÉFONO",(solicitud.get("Telefono") or "").strip()); r+=1
    _drow(ws,r,"FECHA / HORA",fecha_str); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL PRÉSTAMO","FFC900","1E293B"); r+=1
    _drow(ws,r,"PLAN DE INVERSIÓN",solicitud.get("PlanInversion")); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="desarrollo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response