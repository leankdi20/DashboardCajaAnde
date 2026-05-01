# apps/dashboard/views/formularios/servicio_accionista.py
# Solicitud Clave Temporal CajaTel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables
from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_clave_temporal_cajatel import ReporteSolicitudClaveTemporalCajaTel
from apps.dashboard.tables import SolicitudClaveTemporalCajaTelTable

CAMPOS_CAJATEL = {"cedula": "Cedula", "nombre": "NombreCompleto"}

def _fill(h): return PatternFill("solid", fgColor=h)
def _borde(): return Border(bottom=Side(style="thin", color="E2E8F0"))
def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill = _fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font = Font(color="1E293B", size=10, name="Arial")
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _header_inst(ws, titulo, rid):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = _fill("003FB7"); t.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1, value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8; return 4
def _pie(ws, r):
    ws.row_dimensions[r].height = 8; r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = _fill("F1F5F9"); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18


@login_required
@permiso_requerido("dashboard.view_formulario_servicio_accionista")
def formulario_servicio_accionista(request):
    return render(request, "dashboard/formularios/servicio_accionista.html", {"active_tab": "servicio_accionista"})


@login_required
def soli_clave_temporal_cajatel_buscar(request, campo):
    if campo not in CAMPOS_CAJATEL: return JsonResponse({"results": []})
    col_db=CAMPOS_CAJATEL[campo]; term=request.GET.get("term","").strip()
    filas=ReporteSolicitudClaveTemporalCajaTel.buscar_opciones(col_db,term)
    return JsonResponse({"results":[{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()],"pagination":{"more":False}})

@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_lista(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}
    datos=ReporteSolicitudClaveTemporalCajaTel.obtener_datos(filtros); kpis=ReporteSolicitudClaveTemporalCajaTel.obtener_kpis(filtros)
    table=SolicitudClaveTemporalCajaTelTable(datos); tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_clave_temporal_cajatel.html",
                  {"table":table,"filtros":filtros,"kpis":kpis})

@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudClaveTemporalCajaTel.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request,"dashboard/formularios/reportes/soli_clave_temporal_cajatel_detalle.html",{"solicitud":solicitud})

@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteSolicitudClaveTemporalCajaTel.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Clave CajaTel"
    headers=["ID","Fecha / Hora","Cédula","Nombre","Correo Personal"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        fecha=row.get("FechaHora"); fstr=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
        for col,v in enumerate([row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),(row.get("CorreoPersonal") or "").strip()],1): ws.cell(row=ri,column=col,value=v)
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,6): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,18,14,35,35],1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="clave_cajatel_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_export_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudClaveTemporalCajaTel.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"CajaTel #{respuesta_id}"
    for col,w in enumerate([5,25,40,15],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Solicitud Clave Temporal CajaTel",solicitud["respuesta_id"])
    fecha=solicitud.get("FechaHora"); fecha_str=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"CORREO PERSONAL",(solicitud.get("CorreoPersonal") or "").strip()); r+=1
    _drow(ws,r,"FECHA / HORA",fecha_str); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="cajatel_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response