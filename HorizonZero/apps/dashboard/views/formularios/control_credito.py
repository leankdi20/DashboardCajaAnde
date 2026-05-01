# apps/dashboard/views/formularios/control_credito.py
# Comprobante Autorización Ahorro · Comprobantes de Pago
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables
from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_comprobante_autorizacion_ahorro import ReporteComprobanteAutorizacionAhorro
from apps.dashboard.reports.soli_comprobantes_pago import ReporteComprobantesPago
from apps.dashboard.tables import ComprobanteAutorizacionAhorroTable, ComprobantesPagoTable

DOCS_COLS = [f"Documento{i}" for i in range(1, 11)]
CAMPOS_COMP_AHORRO = {"cedula": "Cedula", "nombre": "NombreCompleto"}
CAMPOS_COMP_PAGO   = {"cedula": "Cedula", "nombre": "NombreCompleto", "banco": "Banco"}


def _fill(h): return PatternFill("solid", fgColor=h)
def _borde(): return Border(bottom=Side(style="thin", color="E2E8F0"))
def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill = _fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font = Font(color="1E293B", size=10, name="Arial")
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _urow(ws, row, label, url):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=url or "No disponible")
    if url: vc.hyperlink = url; vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
    else:   vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _header_inst(ws, titulo, rid):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = _fill("003FB7"); t.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1, value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8; return 4
def _pie(ws, r):
    ws.row_dimensions[r].height = 8; r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = _fill("F1F5F9"); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18


@login_required
@permiso_requerido("dashboard.view_formulario_control_credito")
def formulario_control_credito(request):
    return render(request, "dashboard/formularios/control_credito.html", {"active_tab": "control_credito"})


# ── COMPROBANTE AUTORIZACIÓN AHORRO ──────────────────────────────
@login_required
def comprobante_autorizacion_ahorro_buscar(request, campo):
    if campo not in CAMPOS_COMP_AHORRO: return JsonResponse({"results": []})
    col_db = CAMPOS_COMP_AHORRO[campo]; term = request.GET.get("term", "").strip()
    filas = ReporteComprobanteAutorizacionAhorro.buscar_opciones(col_db, term)
    return JsonResponse({"results": [{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()], "pagination": {"more": False}})

@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_lista(request):
    filtros = {k: request.GET.get(k,"").strip() for k in ("cedula","nombre","fecha_inicio","fecha_fin")}
    filtros = {k:v for k,v in filtros.items() if v}
    datos = ReporteComprobanteAutorizacionAhorro.obtener_datos(filtros)
    kpis  = ReporteComprobanteAutorizacionAhorro.obtener_kpis(filtros)
    table = ComprobanteAutorizacionAhorroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_comprobante_autorizacion_ahorro.html",
                  {"table": table, "filtros": filtros, "kpis": kpis})

@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_detalle(request, respuesta_id):
    solicitud = ReporteComprobanteAutorizacionAhorro.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    documentos = [(f"Documento {i}", solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    return render(request, "dashboard/formularios/reportes/soli_comprobante_autorizacion_ahorro_detalle.html",
                  {"solicitud": solicitud, "documentos": documentos})

@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_export(request):
    filtros = {k: request.GET.get(k,"").strip() for k in ("cedula","nombre","fecha_inicio","fecha_fin")}
    filtros = {k:v for k,v in filtros.items() if v}
    datos = ReporteComprobanteAutorizacionAhorro.obtener_datos(filtros)
    wb = Workbook(); ws = wb.active; ws.title = "Autorizacion Ahorro"
    headers = ["ID","Fecha","Hora","Cédula","Nombre","Teléfono","Detalle Pago","Doc1","Doc2","Doc3","Doc4","Doc5","Doc6","Doc7","Doc8","Doc9","Doc10"]
    hf = PatternFill("solid", fgColor="003FB7"); hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        ws.cell(row=ri,column=1,value=row.get("respuesta_id")); ws.cell(row=ri,column=2,value=str(row.get("Fecha") or ""))
        ws.cell(row=ri,column=3,value=str(row.get("Hora") or "")); ws.cell(row=ri,column=4,value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri,column=5,value=(row.get("NombreCompleto") or "").strip()); ws.cell(row=ri,column=6,value=(row.get("NumeroTelefonico") or "").strip())
        ws.cell(row=ri,column=7,value=row.get("DetallePago"))
        for i,doc in enumerate(DOCS_COLS,8): ws.cell(row=ri,column=i,value=row.get(doc))
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,18): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,12,10,14,30,14,35]+[50]*10,1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="autorizacion_ahorro_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_export_detalle(request, respuesta_id):
    solicitud = ReporteComprobanteAutorizacionAhorro.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb = Workbook(); ws = wb.active; ws.title = f"Autorizacion #{respuesta_id}"
    for col,w in enumerate([5,25,50,10],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r = _header_inst(ws,"CAJA DE ANDE — Comprobante Autorización Ahorro Voluntario",solicitud["respuesta_id"])
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"TELÉFONO",(solicitud.get("NumeroTelefonico") or "").strip()); r+=1
    _drow(ws,r,"FECHA",str(solicitud.get("Fecha") or "")); r+=1
    _drow(ws,r,"HORA",str(solicitud.get("Hora") or "")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DETALLE DE PAGO","FFC900","1E293B"); r+=1
    _drow(ws,r,"DETALLE",solicitud.get("DetallePago")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    docs=[(f"DOCUMENTO {i}",solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    if docs:
        _hrow(ws,r,"DOCUMENTOS ADJUNTOS — SHAREFILE","FFC900","1E293B"); r+=1
        for lbl,url in docs: _urow(ws,r,lbl,url); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="autorizacion_ahorro_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ── COMPROBANTES DE PAGO ─────────────────────────────────────────
@login_required
def comprobantes_pago_buscar(request, campo):
    if campo not in CAMPOS_COMP_PAGO: return JsonResponse({"results": []})
    col_db=CAMPOS_COMP_PAGO[campo]; term=request.GET.get("term","").strip()
    filas=ReporteComprobantesPago.buscar_opciones(col_db,term)
    return JsonResponse({"results":[{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()],"pagination":{"more":False}})

@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_lista(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","banco","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}
    datos=ReporteComprobantesPago.obtener_datos(filtros); kpis=ReporteComprobantesPago.obtener_kpis(filtros)
    table=ComprobantesPagoTable(datos); tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_comprobantes_pago.html",
                  {"table":table,"filtros":filtros,"kpis":kpis})

@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_detalle(request, respuesta_id):
    solicitud=ReporteComprobantesPago.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    documentos=[(f"Documento {i}",solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    return render(request,"dashboard/formularios/reportes/soli_comprobantes_pago_detalle.html",
                  {"solicitud":solicitud,"documentos":documentos})

@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","banco","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteComprobantesPago.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Comprobantes Pago"
    headers=["ID","Fecha","Hora","Cédula","Nombre","Teléfono","Banco","N° Depósito","Fecha Depósito","Monto (₡)","Detalle Pago","Doc1","Doc2","Doc3","Doc4","Doc5","Doc6","Doc7","Doc8","Doc9","Doc10"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        monto=row.get("Monto")
        ws.cell(row=ri,column=1,value=row.get("respuesta_id")); ws.cell(row=ri,column=2,value=str(row.get("Fecha") or ""))
        ws.cell(row=ri,column=3,value=str(row.get("Hora") or "")); ws.cell(row=ri,column=4,value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri,column=5,value=(row.get("NombreCompleto") or "").strip()); ws.cell(row=ri,column=6,value=(row.get("NumeroTelefonico") or "").strip())
        ws.cell(row=ri,column=7,value=row.get("Banco")); ws.cell(row=ri,column=8,value=row.get("NumeroDeposito"))
        ws.cell(row=ri,column=9,value=str(row.get("FechaDeposito") or ""))
        cm=ws.cell(row=ri,column=10,value=int(float(monto)) if monto else None)
        if monto: cm.number_format="₡#,##0"
        ws.cell(row=ri,column=11,value=row.get("DetallePago"))
        for i,doc in enumerate(DOCS_COLS,12): ws.cell(row=ri,column=i,value=row.get(doc))
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,22): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,12,10,14,30,14,20,16,14,14,35]+[50]*10,1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="comprobantes_pago_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_export_detalle(request, respuesta_id):
    solicitud=ReporteComprobantesPago.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"Comprobante #{respuesta_id}"
    for col,w in enumerate([5,25,50,10],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Comprobante de Pago / Depósito Bancario",solicitud["respuesta_id"])
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"TELÉFONO",(solicitud.get("NumeroTelefonico") or "").strip()); r+=1
    _drow(ws,r,"FECHA",str(solicitud.get("Fecha") or "")); r+=1
    _drow(ws,r,"HORA",str(solicitud.get("Hora") or "")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL DEPÓSITO","FFC900","1E293B"); r+=1
    _drow(ws,r,"BANCO",solicitud.get("Banco")); r+=1
    _drow(ws,r,"N° DEPÓSITO",solicitud.get("NumeroDeposito")); r+=1
    _drow(ws,r,"FECHA DEPÓSITO",str(solicitud.get("FechaDeposito") or "")); r+=1
    monto=solicitud.get("Monto")
    ws.merge_cells(f"A{r}:B{r}"); lc=ws.cell(row=r,column=1,value="MONTO"); lc.fill=_fill("F1F5F9"); lc.font=Font(bold=True,color="64748B",size=9,name="Arial"); lc.alignment=Alignment(vertical="center",indent=2); lc.border=_borde()
    ws.merge_cells(f"C{r}:D{r}"); vc=ws.cell(row=r,column=3,value=int(float(monto)) if monto else "—")
    if monto: vc.number_format="₡#,##0"
    vc.font=Font(color="1E293B",size=10,name="Arial",bold=True); vc.alignment=Alignment(vertical="center",indent=2); vc.border=_borde(); ws.row_dimensions[r].height=22; r+=1
    _drow(ws,r,"DETALLE PAGO",solicitud.get("DetallePago")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    docs=[(f"DOCUMENTO {i}",solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    if docs:
        _hrow(ws,r,"DOCUMENTOS ADJUNTOS — SHAREFILE","FFC900","1E293B"); r+=1
        for lbl,url in docs: _urow(ws,r,lbl,url); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="comprobante_pago_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response