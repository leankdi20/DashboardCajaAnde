# apps/dashboard/views/formularios/seguros.py
# Seguro Viajero · Marchamo
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import django_tables2 as tables
from apps.core.decorators import permiso_requerido
from apps.dashboard.reports.soli_seguro_viajero import ReporteSolicitudSeguroViajero
from apps.dashboard.reports.soli_marchamo import ReporteSolicitudMarchamo
from apps.dashboard.tables import SolicitudSeguroViajeroTable, SolicitudMarchamoTable
from apps.dashboard.views._base import MESES_ES

CAMPOS_SEGURO_VIAJERO = {"cedula": "Cedula", "nombre": "NombreCompleto", "destino": "Destino"}
CAMPOS_MARCHAMO = {"cedula": "Cedula", "nombre": "NombreCompleto", "placa": "NumeroPlaca"}


def _es_si(valor):
    if valor is None: return False
    return str(valor).strip().lower() in ('sí', 'si', '1', 'yes', 'true', 's')

def _fill(h): return PatternFill("solid", fgColor=h)
def _borde(): return Border(bottom=Side(style="thin", color="E2E8F0"))
def _hrow(ws, row, texto, cf, ct="FFFFFF"):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.fill = _fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
def _drow(ws, row, label, valor):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
    vc.font = Font(color="1E293B", size=10, name="Arial")
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _srow(ws, row, label, activo):
    ws.merge_cells(f"A{row}:B{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill("F1F5F9"); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = _borde()
    ws.merge_cells(f"C{row}:D{row}")
    vc = ws.cell(row=row, column=3, value="✓ Sí" if activo else "✗ No")
    vc.font = Font(color="166534" if activo else "94A3B8", size=10, name="Arial", bold=activo)
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = _borde()
    ws.row_dimensions[row].height = 22
def _header_inst(ws, titulo, rid):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = _fill("003FB7"); t.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1, value=f"ID #{rid}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = _fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8; return 4
def _pie(ws, r):
    ws.row_dimensions[r].height = 8; r += 1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = _fill("F1F5F9"); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18


@login_required
@permiso_requerido("dashboard.view_formulario_seguros")
def formulario_seguros(request):
    return render(request, "dashboard/formularios/seguros.html", {"active_tab": "seguros"})


# ── SEGURO VIAJERO ────────────────────────────────────────────────
@login_required
def soli_seguro_viajero_buscar(request, campo):
    if campo not in CAMPOS_SEGURO_VIAJERO: return JsonResponse({"results": []})
    col_db=CAMPOS_SEGURO_VIAJERO[campo]; term=request.GET.get("term","").strip()
    filas=ReporteSolicitudSeguroViajero.buscar_opciones(col_db,term)
    return JsonResponse({"results":[{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()],"pagination":{"more":False}})

@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_lista(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","destino","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}
    datos=ReporteSolicitudSeguroViajero.obtener_datos(filtros); kpis=ReporteSolicitudSeguroViajero.obtener_kpis(filtros)
    top_destinos=ReporteSolicitudSeguroViajero.obtener_top_destinos()
    raw_timeline=ReporteSolicitudSeguroViajero.obtener_timeline()
    timeline_data=[{"label":f"{MESES_ES[r['mes']]} {r['anio']}","total":r["total"]} for r in raw_timeline]
    heatmap_dict={}; heatmap_anios=[]
    for r in raw_timeline:
        a,m,t=r["anio"],r["mes"],r["total"]
        if a not in heatmap_dict: heatmap_dict[a]={}; heatmap_anios.append(a)
        heatmap_dict[a][m]=t
    table=SolicitudSeguroViajeroTable(datos); tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_seguro_viajero.html",{
        "table":table,"filtros":filtros,"kpis":kpis,"top_destinos":top_destinos,
        "timeline_data":timeline_data,"heatmap_anios":heatmap_anios,"heatmap_json":json.dumps(heatmap_dict)})

@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudSeguroViajero.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    return render(request,"dashboard/formularios/reportes/soli_seguro_viajero_detalle.html",{"solicitud":solicitud})

@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","destino","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteSolicitudSeguroViajero.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Seguro Viajero"
    headers=["ID","Fecha/Hora","Cédula","Nombre","Correo","Teléfono","Destino","Inicio Viaje","Fin Viaje","Nombre Beneficiario","Cédula Beneficiario","Parentesco"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        fecha=row.get("FechaHora"); fstr=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
        for col,v in enumerate([row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),row.get("Correo"),row.get("Telefono"),row.get("Destino"),str(row.get("FechaInicioViaje") or ""),str(row.get("FechaFinalViaje") or ""),row.get("NombreBeneficiario"),row.get("CedulaBeneficiario"),row.get("Parentesco")],1): ws.cell(row=ri,column=col,value=v)
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,13): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,18,14,30,28,12,25,12,12,30,14,15],1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="seguro_viajero_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_export_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudSeguroViajero.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"Viajero #{respuesta_id}"
    for col,w in enumerate([5,25,35,15],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Solicitud Seguro Viajero",solicitud["respuesta_id"])
    fecha=solicitud.get("FechaHora"); fecha_str=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"CORREO",solicitud.get("Correo")); r+=1
    _drow(ws,r,"TELÉFONO",solicitud.get("Telefono")); r+=1
    _drow(ws,r,"FECHA / HORA",fecha_str); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL VIAJE","FFC900","1E293B"); r+=1
    _drow(ws,r,"DESTINO",solicitud.get("Destino")); r+=1
    _drow(ws,r,"FECHA INICIO",str(solicitud.get("FechaInicioViaje") or "—")); r+=1
    _drow(ws,r,"FECHA FINAL",str(solicitud.get("FechaFinalViaje") or "—")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL BENEFICIARIO","003FB7"); r+=1
    _drow(ws,r,"NOMBRE",solicitud.get("NombreBeneficiario")); r+=1
    _drow(ws,r,"CÉDULA",solicitud.get("CedulaBeneficiario")); r+=1
    _drow(ws,r,"PARENTESCO",solicitud.get("Parentesco")); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="viajero_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


# ── MARCHAMO ──────────────────────────────────────────────────────
@login_required
def soli_marchamo_buscar(request, campo):
    if campo not in CAMPOS_MARCHAMO: return JsonResponse({"results": []})
    col_db=CAMPOS_MARCHAMO[campo]; term=request.GET.get("term","").strip()
    filas=ReporteSolicitudMarchamo.buscar_opciones(col_db,term)
    return JsonResponse({"results":[{"id":(r[col_db] or "").strip(),"text":(r[col_db] or "").strip()} for r in filas if r.get(col_db) and r[col_db].strip()],"pagination":{"more":False}})

@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_lista(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","placa","tipo_vehiculo","sucursal","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}
    datos=ReporteSolicitudMarchamo.obtener_datos(filtros); kpis=ReporteSolicitudMarchamo.obtener_kpis(filtros)
    tipos_vehiculo=ReporteSolicitudMarchamo.obtener_tipos_vehiculo(); sucursales=ReporteSolicitudMarchamo.obtener_sucursales()
    table=SolicitudMarchamoTable(datos); tables.RequestConfig(request,paginate={"per_page":50}).configure(table)
    return render(request,"dashboard/formularios/reportes/soli_marchamo.html",
                  {"table":table,"filtros":filtros,"kpis":kpis,"tipos_vehiculo":tipos_vehiculo,"sucursales":sucursales})

@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudMarchamo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    seguros=[]
    mapa=[("SeguroRC","Responsabilidad Civil (RC)"),("SeguroVida","Seguro de Vida"),("SeguroSalud","Seguro de Salud"),
          ("SeguroAsistencia","Seguro de Asistencia"),("SeguroVidaPlus","Seguro Vida Plus"),("SeguroComprensivo","Seguro Comprensivo")]
    for campo,label in mapa: seguros.append({"label":label,"activo":_es_si(solicitud.get(campo))})
    return render(request,"dashboard/formularios/reportes/soli_marchamo_detalle.html",{"solicitud":solicitud,"seguros":seguros})

@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_export(request):
    filtros={k:request.GET.get(k,"").strip() for k in ("cedula","nombre","placa","tipo_vehiculo","sucursal","fecha_inicio","fecha_fin")}
    filtros={k:v for k,v in filtros.items() if v}; datos=ReporteSolicitudMarchamo.obtener_datos(filtros)
    wb=Workbook(); ws=wb.active; ws.title="Marchamo"
    headers=["ID","Fecha/Hora","Cédula","Nombre","Teléfono","Tipo Vehículo","Placa","Dueño Registral","Pago Ahorro","Pago Tarjeta","Sucursal Retiro","Seg. RC","Seg. Vida","Seg. Salud","Seg. Asistencia","Seg. Vida Plus","Seg. Comprensivo"]
    hf=PatternFill("solid",fgColor="003FB7"); hfont=Font(bold=True,color="FFFFFF",name="Arial")
    for col,h in enumerate(headers,1): c=ws.cell(row=1,column=col,value=h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,row in enumerate(datos,2):
        fecha=row.get("FechaHora"); fstr=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
        for col,v in enumerate([row.get("respuesta_id"),fstr,(row.get("Cedula") or "").strip(),(row.get("NombreCompleto") or "").strip(),row.get("telefono"),row.get("TipoVehiculo"),(row.get("NumeroPlaca") or "").strip(),row.get("NombreDuenoRegistral"),row.get("AutorizoPagoAhorro"),row.get("AutorizoPagoTarjeta"),row.get("SucursalRetiro"),row.get("SeguroRC"),row.get("SeguroVida"),row.get("SeguroSalud"),row.get("SeguroAsistencia"),row.get("SeguroVidaPlus"),row.get("SeguroComprensivo")],1): ws.cell(row=ri,column=col,value=v)
        if ri%2==0:
            ff=PatternFill("solid",fgColor="E8EFFE")
            for c in range(1,18): ws.cell(row=ri,column=c).fill=ff
    for i,w in enumerate([8,18,14,30,12,14,10,25,12,12,22,10,10,10,14,13,15],1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    fstr=datetime.now().strftime("%Y%m%d_%H%M")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="marchamo_{fstr}.xlsx"'
    wb.save(response); return response

@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_export_detalle(request, respuesta_id):
    solicitud=ReporteSolicitudMarchamo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb=Workbook(); ws=wb.active; ws.title=f"Marchamo #{respuesta_id}"
    for col,w in enumerate([5,25,35,15],1): ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    r=_header_inst(ws,"CAJA DE ANDE — Solicitud Pago Marchamo",solicitud["respuesta_id"])
    fecha=solicitud.get("FechaHora"); fecha_str=fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha,"strftime") else str(fecha or "")
    _hrow(ws,r,"DATOS DEL ACCIONISTA","003FB7"); r+=1
    _drow(ws,r,"CÉDULA",(solicitud.get("Cedula") or "").strip()); r+=1
    _drow(ws,r,"NOMBRE",(solicitud.get("NombreCompleto") or "").strip()); r+=1
    _drow(ws,r,"TELÉFONO",solicitud.get("telefono")); r+=1
    _drow(ws,r,"FECHA / HORA",fecha_str); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"DATOS DEL VEHÍCULO","FFC900","1E293B"); r+=1
    _drow(ws,r,"TIPO DE VEHÍCULO",solicitud.get("TipoVehiculo")); r+=1
    _drow(ws,r,"NÚMERO DE PLACA",(solicitud.get("NumeroPlaca") or "").strip()); r+=1
    _drow(ws,r,"DUEÑO REGISTRAL",solicitud.get("NombreDuenoRegistral")); r+=1
    _drow(ws,r,"PAGO CON AHORRO",solicitud.get("AutorizoPagoAhorro")); r+=1
    _drow(ws,r,"PAGO CON TARJETA",solicitud.get("AutorizoPagoTarjeta")); r+=1
    _drow(ws,r,"SUCURSAL RETIRO",solicitud.get("SucursalRetiro")); r+=1
    ws.row_dimensions[r].height=8; r+=1
    _hrow(ws,r,"SEGUROS INCLUIDOS","003FB7"); r+=1
    mapa=[("SeguroRC","Responsabilidad Civil (RC)"),("SeguroVida","Seguro de Vida"),("SeguroSalud","Seguro de Salud"),
          ("SeguroAsistencia","Seguro de Asistencia"),("SeguroVidaPlus","Seguro Vida Plus"),("SeguroComprensivo","Seguro Comprensivo")]
    for campo,label in mapa: _srow(ws,r,label,_es_si(solicitud.get(campo))); r+=1
    _pie(ws,r); ws.freeze_panes="A3"; ws.sheet_view.zoomScale=110
    nombre=(solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]=f'attachment; filename="marchamo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response