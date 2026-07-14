# apps/dashboard/views/logs.py
import io
import json
from datetime import datetime as dt

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages

from apps.core.decorators import permiso_requerido
from apps.dashboard.services.api_client import APIClient

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _get_filtros(request) -> dict:
    return {
        "usuario":   request.GET.get("usuario", ""),
        "accion":    request.GET.get("accion", ""),
        "modulo":    request.GET.get("modulo", ""),
        "severidad": request.GET.get("severidad", ""),
        "ip":        request.GET.get("ip", ""),
        "desde":     request.GET.get("desde", ""),
        "hasta":     request.GET.get("hasta", ""),
    }


@login_required
@permiso_requerido("dashboard.view_logs")
def logs_home(request):
    filtros = _get_filtros(request)
    page    = request.GET.get("page", 1)
    query_params = request.GET.copy()
    query_params.pop("page", None)

    try:
        kpis       = APIClient.get("logs/kpis/",       params=filtros, request=request)
        actividad  = APIClient.get("logs/actividad/",  params=filtros, request=request)
        dist       = APIClient.get("logs/distribucion/",params=filtros, request=request)
        logs_data  = APIClient.get("logs/", params={**filtros, "page": page, "per_page": 10}, request=request)
    except Exception as e:
        messages.error(request, f"Error al obtener los logs: {e}")
        kpis = {}; actividad = []; dist = []; logs_data = {"results": [], "count": 0, "pages": 1}

    return render(request, "dashboard/Vista_Logs/logs_home.html", {
        "logs":               logs_data.get("results", []),
        "page":               int(page),
        "total_pages":        logs_data.get("pages", 1),
        "total_count":        logs_data.get("count", 0),
        "querystring_without_page": query_params.urlencode(),
        "filtro_usuario":     filtros["usuario"],
        "filtro_accion":      filtros["accion"],
        "filtro_modulo":      filtros["modulo"],
        "filtro_severidad":   filtros["severidad"],
        "filtro_ip":          filtros["ip"],
        "filtro_desde":       filtros["desde"],
        "filtro_hasta":       filtros["hasta"],
        "total_eventos":      kpis.get("total_eventos", 0),
        "logins_fallidos":    kpis.get("logins_fallidos", 0),
        "acciones_criticas":  kpis.get("acciones_criticas", 0),
        "usuarios_unicos":    kpis.get("usuarios_unicos", 0),
        "grafico_labels":     json.dumps([r["dia"]   for r in actividad]),
        "grafico_data":       json.dumps([r["total"] for r in actividad]),
        "dist_labels":        json.dumps([r["accion_display"] for r in dist]),
        "dist_data":          json.dumps([r["total"]          for r in dist]),
        "acciones":           [
            ("LOGIN_OK","Inicio de sesión"), ("LOGIN_FAIL","Intento fallido"),
            ("LOGOUT","Cierre de sesión"), ("CREATE","Creación"),
            ("UPDATE","Modificación"), ("DELETE","Eliminación"),
            ("EXPORT_EXCEL","Exportación Excel"), ("EXPORT_PDF","Exportación PDF"),
            ("VIEW","Vista de reporte"), ("DENEGADO","Acceso denegado"),
        ],
        "modulos": [
            ("SISTEMA","Sistema"), ("ENCUESTAS","Encuestas"),
            ("FORMULARIOS","Formularios"), ("AGENTES","Agentes"),
            ("USUARIOS","Usuarios"), ("LOGS","Logs"),
        ],
        "severidades": [
            ("INFO","Informativo"), ("WARNING","Advertencia"), ("CRITICAL","Crítico"),
        ],
    })


@login_required
@permiso_requerido("dashboard.view_logs")
def log_detalle(request, log_id):
    try:
        log = APIClient.get(f"logs/{log_id}/", request=request)
    except Exception:
        messages.error(request, "Registro no encontrado.")
        return redirect("dashboard:logs_home")

    def _parse(raw):
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return raw

    return render(request, "dashboard/Vista_Logs/logs_detalle.html", {
        "log":              log,
        "datos_anteriores": _parse(log.get("datos_anteriores")),
        "datos_nuevos":     _parse(log.get("datos_nuevos")),
    })


@login_required
@permiso_requerido("dashboard.view_logs")
def logs_export_excel(request):
    filtros = _get_filtros(request)

    try:
        logs = APIClient.get("logs/", params={**filtros, "per_page": 5000}, request=request)
        qs   = logs.get("results", [])
    except Exception as e:
        messages.error(request, f"Error al exportar: {e}")
        return redirect("dashboard:logs_home")

    wb = Workbook(); ws = wb.active; ws.title = "Logs de Auditoría"
    AZUL = "003FB7"; GRIS = "F1F5F9"

    headers = ["ID", "Fecha / Hora", "Usuario", "Acción", "Módulo",
               "Severidad", "Descripción", "IP Address", "URL", "Método HTTP"]

    hf    = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ri, log in enumerate(qs, 2):
        sev = log.get("severidad", "")
        if sev == "CRITICAL":
            rf = PatternFill("solid", fgColor="FEE2E2")
        elif sev == "WARNING":
            rf = PatternFill("solid", fgColor="FEF9C3")
        else:
            rf = PatternFill("solid", fgColor="FFFFFF" if ri % 2 == 0 else GRIS)

        for col, val in enumerate([
            str(log.get("id", "")),
            log.get("fecha", ""),
            log.get("username", ""),
            log.get("accion_display", ""),
            log.get("modulo_display", ""),
            log.get("severidad_display", ""),
            log.get("descripcion", ""),
            str(log.get("ip_address", "") or ""),
            log.get("url", "") or "",
            log.get("metodo_http", "") or "",
        ], 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = rf; c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[ri].height = 18

    for i, w in enumerate([8, 18, 20, 22, 16, 12, 60, 16, 50, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    fstr     = dt.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="audit_log_{fstr}.xlsx"'
    wb.save(response)
    return response

@login_required
@permiso_requerido("dashboard.view_logs")
def logs_export_pdf(request):
    filtros = _get_filtros(request)

    try:
        logs = APIClient.get("logs/", params={**filtros, "per_page": 500}, request=request)
        qs   = logs.get("results", [])
    except Exception as e:
        messages.error(request, f"Error al exportar: {e}")
        return redirect("dashboard:logs_home")

    try:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                   rightMargin=1*cm, leftMargin=1*cm,
                                   topMargin=1.5*cm, bottomMargin=1*cm)

        styles     = getSampleStyleSheet()
        azul       = colors.HexColor("#003FB7")
        gris       = colors.HexColor("#F1F5F9")
        titulo_style = ParagraphStyle("titulo", parent=styles["Heading1"],
                                      textColor=azul, fontSize=14, spaceAfter=4)
        sub_style    = ParagraphStyle("sub", parent=styles["Normal"],
                                      textColor=colors.HexColor("#64748B"),
                                      fontSize=8, spaceAfter=12)
        celda_style  = ParagraphStyle("celda", parent=styles["Normal"], fontSize=7)

        elementos = [
            Paragraph("Reporte de Auditoría — HorizonZero · Caja de ANDE", titulo_style),
            Paragraph(
                f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} "
                f"por {request.user.get_full_name() or request.user.username} · "
                "Clasificación: USO INTERNO RESTRINGIDO",
                sub_style,
            ),
        ]

        data = [["ID", "Fecha / Hora", "Usuario", "Acción", "Módulo",
                 "Severidad", "Descripción", "IP"]]

        SEV_COLOR = {
            "CRITICAL": colors.HexColor("#FEE2E2"),
            "WARNING":  colors.HexColor("#FEF9C3"),
            "INFO":     colors.white,
        }
        row_colors = [colors.HexColor("#003FB7")]

        for log in qs:
            data.append([
                str(log.get("id", "")),
                log.get("fecha", ""),
                log.get("username", ""),
                Paragraph(log.get("accion_display", ""), celda_style),
                log.get("modulo_display", ""),
                log.get("severidad_display", ""),
                Paragraph((log.get("descripcion", "") or "")[:80], celda_style),
                str(log.get("ip_address", "") or "—"),
            ])
            row_colors.append(SEV_COLOR.get(log.get("severidad", ""), colors.white))

        col_widths = [1.2*cm, 2.5*cm, 3*cm, 3.5*cm, 2.5*cm, 2*cm, 8*cm, 2.5*cm]
        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), azul),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 7),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 7),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, gris]),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",   (0,0), (-1,-1), 4),
            ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ]))
        for i, color in enumerate(row_colors[1:], 1):
            if color != colors.white:
                tabla.setStyle(TableStyle([("BACKGROUND", (0,i), (-1,i), color)]))

        elementos.append(tabla)
        doc.build(elementos)
        buffer.seek(0)

        fstr     = dt.now().strftime("%Y%m%d_%H%M")
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="audit_log_{fstr}.pdf"'
        return response

    except ImportError:
        messages.error(request, "ReportLab no está instalado.")
        return redirect("dashboard:logs_home")
