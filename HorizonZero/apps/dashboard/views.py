import io
import base64
import zipfile
import qrcode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from apps.core.decorators import permiso_requerido
from django_tables2 import SingleTableView, tables
from django.http import HttpResponse, Http404, JsonResponse
from django.core.paginator import Paginator

from django.views.decorators.http import require_POST


from .services.db_service import ReportesDBService

from .reports.encuesta_satisfaccion import ReporteEncuestaSatisfaccion

from .reports import encuesta_oficina_digital,encuesta_pagina_web, encuesta_whatsapp_agente, encuesta_whatsapp, encuesta_feria_salud
from .reports.soli_tarj_credito import ReporteSolicitudTarjetaCredito
from .reports.soli_tarj_debito_report import ReporteSolicitudTarjetaDebito
from .reports.soli_tarj_debito import ReporteSolicitudTarjetaDebitoGestion
from .reports.soli_redencion_puntos_report import ReporteSolicitudRedencionPuntos
from .reports.soli_caja_ande_asistencia import ReporteCajaAndeAsistencia

from .reports.soli_deposito_salario import ReporteSolicitudDepositoSalario
from .reports.soli_ahorro_mod_cuota import ReporteSolicitudAhorroModCuota
from .reports.soli_reinversion_ahorro import ReporteSolicitudReinversionAhorro
from .reports.soli_autorizacion_ahorro_nuevo import ReporteSolicitudAutorizacionAhorroNuevo
from .reports.soli_compra_vehiculo import ReporteSolicitudCompraVehiculo
from .reports.soli_prestamo_vivienda import ReporteSolicitudPrestamoVivienda
from .reports.soli_prestamo_desarrollo import ReporteSolicitudPrestamoDesarrollo
from .reports.soli_presolicitud_credito_personal import ReporteSolicitudPresolicitudCreditoPersonal
from .reports.soli_comprobante_autorizacion_ahorro import ReporteComprobanteAutorizacionAhorro
from .reports.soli_comprobantes_pago import ReporteComprobantesPago
from .reports.soli_clave_temporal_cajatel import ReporteSolicitudClaveTemporalCajaTel
from .reports.soli_seguro_viajero import ReporteSolicitudSeguroViajero
from .reports.soli_marchamo import ReporteSolicitudMarchamo

from apps.dashboard.models import AuditLog
from django.db.models.functions import TruncDate
from django.db.models import Count, Q
from apps.core.audit import audit, audit_login_fail


from .tables import EncuestaSatisfaccionTable,EncuestaOficinaDigitalTable, EncuestaPaginaWebTable, EncuestaWhatsappAgenteTable, EncuestaWhatsappTable, EncuestaFeriaSaludTable, SolicitudTarjetaCreditoTable, SolicitudTarjetaDebito_CiuOroTable, SolicitudTarjetaDebitoGestionTable, SolicitudRedencionPuntosTable, CajaAndeAsistenciaTable
from .tables import SolicitudDepositoSalarioTable, SolicitudAhorroModCuotaTable, SolicitudReinversionAhorroTable, SolicitudAutorizacionAhorroNuevoTable ,SolicitudCompraVehiculoTable, SolicitudPrestamoViviendaTable, SolicitudPrestamoDesarrolloTable, SolicitudPresolicitudCreditoPersonalTable, ComprobanteAutorizacionAhorroTable, ComprobantesPagoTable, SolicitudClaveTemporalCajaTelTable, SolicitudSeguroViajeroTable, SolicitudMarchamoTable


from .services.agentes_service import AgentesService,UNIDAD_ENCUESTA,ENCUESTA_NOMBRE, ENCUESTAS_QR, build_encuesta_url, generar_qr_base64

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,Border, Side


from django.http import JsonResponse


 
@login_required
def dashboard_home(request):
    for attr in ('_perm_cache', '_user_perm_cache', '_group_perm_cache'):
        try:
            delattr(request.user, attr)
        except AttributeError:
            pass
    return render(request, "dashboard/home.html", {   # ← home.html, no dashboard_base.html
        "show_welcome": True,
    })






# ── ENCUESTAS ───────────── ENCUESTAS ─────────────────── ENCUESTAS ──────────────────

# ── ENCUESTAS ──────────────SATISFACCION────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion(request):
    # ── Filtros de la TABLA ──────────────────────────────
    filtros = {
        "agente":       request.GET.get("agente"),
        "sucursal":     request.GET.get("sucursales"),
        "unidad":       request.GET.get("unidad"),
        "gestion":      request.GET.get("gestion"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
        "clasificacion": request.GET.get("clasificacion"), 
    }

    # ── Filtros de los KPIs ──────────────────────────────
    kpi_sucursales   = request.GET.getlist("kpi_sucursal")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    # ── Inicializar variables ────────────────────────────
    datos = []
    opciones_agente = opciones_unidad = opciones_sucursal = opciones_gestion = opciones_nombre = []
    kpis_globales = {}

    # ── Queries ──────────────────────────────────────────
    try:
        datos         = ReporteEncuestaSatisfaccion.obtener_datos_agrupados(filtros)
        kpis_globales = ReporteEncuestaSatisfaccion.obtener_kpis_globales(
            sucursales=kpi_sucursales if kpi_sucursales else None,
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
        opciones_agente   = ReportesDBService.ejecutar_query("SELECT DISTINCT Agente FROM dbo.vw_reporte_encuestas_satisfaccion WHERE Agente IS NOT NULL ORDER BY Agente")
        opciones_unidad   = ReportesDBService.ejecutar_query("SELECT DISTINCT Unidad FROM dbo.vw_reporte_encuestas_satisfaccion WHERE Unidad IS NOT NULL ORDER BY Unidad")
        opciones_sucursal = ReportesDBService.ejecutar_query("SELECT DISTINCT Sucursal FROM dbo.vw_reporte_encuestas_satisfaccion WHERE Sucursal IS NOT NULL ORDER BY Sucursal")
        opciones_gestion  = ReportesDBService.ejecutar_query("SELECT DISTINCT Gestion FROM dbo.vw_reporte_encuestas_satisfaccion WHERE Gestion IS NOT NULL ORDER BY Gestion")
        opciones_nombre   = ReportesDBService.ejecutar_query("SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion WHERE Nombre IS NOT NULL ORDER BY Nombre")
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    # ── Tabla ────────────────────────────────────────────
    table = EncuestaSatisfaccionTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion.html", {
        "table":             table,
        "filtros":           filtros,
        "kpis_globales":     kpis_globales,
        "kpi_sucursales":    kpi_sucursales,
        "kpi_fecha_inicio":  kpi_fecha_inicio,
        "kpi_fecha_fin":     kpi_fecha_fin,
        "opciones_agente":   opciones_agente,
        "opciones_unidad":   opciones_unidad,
        "opciones_sucursal": opciones_sucursal,
        "opciones_gestion":  opciones_gestion,
        "opciones_nombre":   opciones_nombre,
    })


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_detalle(request, respuesta_id):
    try:
        filas = ReporteEncuestaSatisfaccion.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

     # Transformar respuestas numéricas
    for fila in filas:
        pregunta = (fila.get("Pregunta") or "").lower()
        respuesta = fila.get("Respuesta", "")
        
        if "recomienda" in pregunta:
            if str(respuesta).strip() == "1":
                fila["Respuesta"] = "Sí"
            elif str(respuesta).strip() == "0" or str(respuesta).strip() == "2":
                fila["Respuesta"] = "No"

    encabezado = filas[0] if filas else {}

     # ── Calcular KPIs ──────────────────────────────────────
    #Solo filas numericas con respuestas numericas para promedios
    respuestas_numericas = []
    for fila in filas:
        try:
            respuestas_numericas.append(float(fila.get("Respuesta", 0)))
        except (ValueError, TypeError):
            pass
  

    #Satisfacción = % de respuestas con valor >= 4
    satisfechos = sum(1 for r in respuestas_numericas if r >= 4)
    satisfaccion_pct = round((satisfechos / len(respuestas_numericas) * 100)) if respuestas_numericas else 0
    
   

    # Después de obtener encabezado
    stats_agente    = ReporteEncuestaSatisfaccion.obtener_promedio_agente(
        encabezado.get("Agente", "")
    )
    promedio_encuesta = ReporteEncuestaSatisfaccion.obtener_promedio_encuesta(respuesta_id)
     # Convertir escala 1-5 a porcentaje
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 5) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0
    
    # Buscar respuesta de la pregunta de satisfacción
    respuesta_satisfaccion = ""
    for fila in filas:
        if "satisfecho" in (fila.get("Pregunta") or "").lower():
            respuesta_satisfaccion = fila.get("Respuesta", "")
            break

    kpis = {
        "promedio_general":   promedio_agente_pct, 
        "total_encuestas":    stats_agente["total_encuestas"],
        "promedio_encuesta":   promedio_encuesta_pct, 
        "respuesta_satisfaccion": respuesta_satisfaccion, 
        "satisfaccion_pct":   satisfaccion_pct,
        "agente":             encabezado.get("Agente", ""),
        "sucursal":           encabezado.get("Sucursal", ""),
        "unidad":             encabezado.get("Unidad", ""),
    }



    return render(request, "dashboard/encuestas/satisfaccion_detalle.html", {
        "encabezado":   encabezado,
        "preguntas":    filas,
        "respuesta_id": respuesta_id,
        "kpis":         kpis,
    })


    # Es necesario que, al boton exportar excel por detalle de agente se haga tambien la exportacion de kís


# Exportar filtro de encuesta satisfacción a hoja de excel:
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_exportar(request):
    filtros = {
        "agente":       request.GET.get("agente"),
        "sucursal":     request.GET.get("sucursal"),
        "unidad":       request.GET.get("unidad"),
        "gestion":      request.GET.get("gestion"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
        "clasificacion":   request.GET.get("clasificacion"),
    }

    datos = ReporteEncuestaSatisfaccion.obtener_datos(filtros)

    # Agrupar por respuesta_id y pivotar preguntas en columnas
    encuestas = {}
    preguntas_orden = []  # mantener orden de preguntas

    for fila in datos:
        rid = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        respuesta = fila.get("Respuesta", "")

        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha":    fila.get("Fecha", ""),
                "Hora":     fila.get("Hora", ""),
                "Agente":   fila.get("Agente", ""),
                "Unidad":   fila.get("Unidad", ""),
                "Sucursal": fila.get("Sucursal", ""),
                "Gestion":  fila.get("Gestion", ""),
                "Nombre":   fila.get("Nombre", ""),
                "Cedula":   fila.get("Cedula", ""),
            }

        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)

        encuestas[rid][pregunta] = respuesta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Encuesta Satisfacción"

    # Columnas fijas + una columna por pregunta
    cols_fijas = ["ID", "Fecha",  "Agente", "Unidad",
                  "Sucursal", "Gestión", "Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha", "Agente", "Unidad",
                  "Sucursal", "Gestion", "Nombre", "Cedula"]

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados fijos
    for col_idx, col_name in enumerate(cols_fijas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Encabezados de preguntas
    for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = PatternFill("solid", fgColor="FFC900")
        cell.font = Font(bold=True, color="1A1000")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Datos — una fila por encuesta
    for row_idx, (rid, enc) in enumerate(encuestas.items(), start=2):
        # Columnas fijas
        for col_idx, key in enumerate(keys_fijas, start=1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")

        # Columnas de preguntas — DENTRO del loop de encuestas
        for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")
            # Transformar recomendación
            if "recomienda" in pregunta.lower():
                if str(respuesta).strip() == "1":
                    respuesta = "Sí"
                elif str(respuesta).strip() in ["0", "2"]:
                    respuesta = "No"
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    # Autoajustar columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Altura de fila del encabezado para preguntas largas
    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="encuesta_satisfaccion.xlsx"'
    wb.save(response)
    return response

@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion")
def encuesta_satisfaccion_detalle_exportar(request, respuesta_id):
    try:
        filas = ReporteEncuestaSatisfaccion.obtener_detalle(respuesta_id)
    except Exception:
        filas = []

    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado = filas[0]

    for fila in filas:
        pregunta = (fila.get("Pregunta") or "").lower()
        respuesta = fila.get("Respuesta", "")
        if "recomienda" in pregunta:
            if str(respuesta).strip() == "1":
                fila["Respuesta"] = "Sí"
            elif str(respuesta).strip() in ["0", "2"]:
                fila["Respuesta"] = "No"
    stats_agente       = ReporteEncuestaSatisfaccion.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta  = ReporteEncuestaSatisfaccion.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 5) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    from openpyxl.styles import Border, Side
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Encuesta"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    # ── Anchos de columna ───────────────────────────────
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 2

    # ── HEADER (filas 1-4) ───────────────────────────────
    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    # Logo
    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = centro

    # Título
    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte de Encuesta de Satisfacción")
    c.font = font(bold=True, color="003FB7", size=15)
    c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9)
    c.alignment = izq

    # Fecha
    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(bold=True, color="FFC900", size=8)
    c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11)
    c.alignment = der

    # Línea separadora
    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    # ── KPI CARDS (fila 5-6) ─────────────────────────────
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    kpi_cols   = [4, 5, 6]
    kpi_labels = ["PROMEDIO GENERAL AGENTE", "ENCUESTAS DEL AGENTE", "PROMEDIO ESTA ENCUESTA"]
    kpi_vals   = [f"{promedio_agente_pct}%", str(stats_agente["total_encuestas"]), f"{promedio_encuesta_pct}%"]
    kpi_fills  = ["003FB7", "E8EEFF", "8B1A0A"]
    kpi_fonts  = ["FFFFFF", "003FB7", "FFFFFF"]

    for i, col in enumerate(kpi_cols):
        c = ws.cell(row=5, column=col, value=kpi_labels[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=8)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

        c = ws.cell(row=6, column=col, value=kpi_vals[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=22)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

    # ── PANEL IZQUIERDO + PREGUNTAS (desde fila 8) ───────
    panel_items = [
        ("AGENTE RESPONSABLE",    encabezado.get("Agente",   ""), True),
        ("UNIDAD INSTITUCIONAL",  encabezado.get("Unidad",   ""), True),
        ("SUCURSAL REGIONAL",     encabezado.get("Sucursal", ""), True),
        ("TIPO DE GESTIÓN",       encabezado.get("Gestion",  ""), True),
        ("NOMBRE ACCIONISTA",     encabezado.get("Nombre",   "—"), True),
        ("CÉDULA",                encabezado.get("Cedula",   "—"), True),
        ("FECHA",                 str(encabezado.get("Fecha", "")), True),
    ]

    # Encabezado tabla preguntas
    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18

    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = izq

    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = centro

    # Filas de datos — panel izquierdo alineado con preguntas
    row_data = 9
    for i in range(max(len(panel_items) * 2, len(filas))):
        ws.row_dimensions[row_data + i].height = 22

    for i, (label, valor, _) in enumerate(panel_items):
        r_label = row_data + (i * 2)
        r_valor = row_data + (i * 2) + 1

        # Label
        c = ws.cell(row=r_label, column=2, value=label)
        c.fill = fill("FFF0C0")
        c.font = font(bold=True, color="003FB7", size=8)
        c.alignment = izq

        # Valor
        c = ws.cell(row=r_valor, column=2, value=valor)
        c.fill = fill("FFF8E0")
        c.font = font(color="1A1A1A", size=10)
        c.alignment = izq

    # Preguntas alineadas desde row_data
    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9)
        c.alignment = izq
        c.border = border_full("E0E0E0")

        respuesta = fila.get("Respuesta", "")
        color_resp = "4CAF50" if respuesta in ["Muy satisfecho", "Muy fácil", "Sí"] else \
             "F44336" if respuesta in ["No", "Nada satisfecho", "Muy difícil"] else \
             "003FB7"
        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro
        c.border = border_full(color_resp)

    # ── PIE DE PÁGINA ────────────────────────────────────
    row_pie = row_data + max(len(panel_items) * 2, len(filas)) + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")

    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie+1, start_column=2, end_row=row_pie+1, end_column=6)
    c = ws.cell(row=row_pie+1, column=2, value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A", size=8)
    c.alignment = centro

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="encuesta_detalle_{respuesta_id}.xlsx"'
    wb.save(response)
    return response







# ── ENCUESTAS ───── OFICINA DIGITAL─────────────────────────────────────────
# ── ENCUESTAS ───── OFICINA DIGITAL─────────────────────────────────────────

# VISTA GENERAL DE LA TABLA DE OFICINA DIGITAL
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina")
def encuesta_satisfaccion_oficina(request):
        # ── Filtros de la TABLA ──────────────────────────────
    filtros = {
        "agente":       request.GET.get("agente"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }

    # ── Filtros de los KPIs ──────────────────────────────
    kpi_sucursales   = request.GET.getlist("kpi_sucursal")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    # ── Inicializar variables ────────────────────────────
    datos = []
    opciones_agente = opciones_unidad = opciones_sucursal = opciones_gestion = opciones_nombre = []
    kpis_globales = {}

    # ── Queries ──────────────────────────────────────────
    try:
        datos         = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_datos_agrupados(filtros)
        kpis_globales = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_kpis_globales(
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
        opciones_agente   = ReportesDBService.ejecutar_query("SELECT DISTINCT Agente FROM dbo.vw_reporte_encuestas_satisfaccion_oficina_digital WHERE Agente IS NOT NULL ORDER BY Agente")
        
        try:
            opciones_nombre = ReportesDBService.ejecutar_query(
                "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_oficina_digital WHERE Nombre IS NOT NULL ORDER BY Nombre"
            )
            # print(">>> NOMBRE OK:", len(opciones_nombre))
        except Exception as e:
            print(">>> NOMBRE ERROR:", e)
            opciones_nombre = []
        
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    # ── Tabla ────────────────────────────────────────────
    table = EncuestaOficinaDigitalTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    

    return render(request, "dashboard/encuestas/satisfaccion_of_dig.html", {
        "table":             table,
        "filtros":           filtros,
        "kpis_globales":     kpis_globales,
        "kpi_sucursales":    kpi_sucursales,
        "kpi_fecha_inicio":  kpi_fecha_inicio,
        "kpi_fecha_fin":     kpi_fecha_fin,
        "opciones_agente":   opciones_agente,
        "opciones_nombre":   opciones_nombre,
    })

    
#VISTA AL DETALLE DE ATENCION DE CADA AGENTE DE OFICINA DIGITAL
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_detalle_of_dig(request, respuesta_id):
    try:
        filas = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    encabezado = filas[0] if filas else {}

     # ── Calcular KPIs ──────────────────────────────────────
    #Solo filas numericas con respuestas numericas para promedios
    respuestas_numericas = []
    for fila in filas:
        try:
            respuestas_numericas.append(float(fila.get("Respuesta", 0)))
        except (ValueError, TypeError):
            pass
  

    #Satisfacción = % de respuestas con valor >= 4
    satisfechos = sum(1 for r in respuestas_numericas if r >= 4)
    satisfaccion_pct = round((satisfechos / len(respuestas_numericas) * 100)) if respuestas_numericas else 0
    
   

    # Después de obtener encabezado
    stats_agente    = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_agente(
        encabezado.get("Agente", "")
    )
    promedio_encuesta = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_encuesta(respuesta_id)
     # Convertir escala 1-5 a porcentaje
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 3) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 3) * 100) if promedio_encuesta else 0
    
    # Buscar respuesta de la pregunta de satisfacción
    respuesta_satisfaccion = ""
    for fila in filas:
        if "satisfecho" in (fila.get("Pregunta") or "").lower():
            respuesta_satisfaccion = fila.get("Respuesta", "")
            break

    kpis = {
        "promedio_general":   promedio_agente_pct, 
        "total_encuestas":    stats_agente["total_encuestas"],
        "promedio_encuesta":   promedio_encuesta_pct, 
        "respuesta_satisfaccion": respuesta_satisfaccion, 
        "satisfaccion_pct":   satisfaccion_pct,

    }



    return render(request, "dashboard/encuestas/satisfaccion_detalle_of_dig.html", {
        "encabezado":   encabezado,
        "preguntas":    filas,
        "respuesta_id": respuesta_id,
        "kpis":         kpis,
    })


# Exportar filtro de encuesta satisfacción a hoja de excel general:
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_of_dig_exportar(request):
    filtros = {
        "agente":       request.GET.get("agente"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }

    datos = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_datos(filtros)

    # Agrupar por respuesta_id y pivotar preguntas en columnas
    encuestas = {}
    preguntas_orden = []  # mantener orden de preguntas

    for fila in datos:
        rid = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        respuesta = fila.get("Respuesta", "")

        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha":    fila.get("Fecha", ""),
                "Hora":     fila.get("Hora", ""),
                "Agente":   fila.get("Agente", ""),
                "Nombre":   fila.get("Nombre", ""),
                "Cedula":   fila.get("Cedula", ""),
            }

        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)

        encuestas[rid][pregunta] = respuesta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Encuesta Satisfacción Oficina Digital"

    # Columnas fijas + una columna por pregunta
    cols_fijas = ["ID", "Fecha",  "Agente", "Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha", "Agente", "Nombre", "Cedula"]

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados fijos
    for col_idx, col_name in enumerate(cols_fijas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Encabezados de preguntas
    for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = PatternFill("solid", fgColor="FFC900")
        cell.font = Font(bold=True, color="1A1000")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Datos — una fila por encuesta
    for row_idx, (rid, enc) in enumerate(encuestas.items(), start=2):
        for col_idx, key in enumerate(keys_fijas, start=1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")

        for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    # Autoajustar columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Altura de fila del encabezado para preguntas largas
    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="encuesta_satisfaccion_oficina_digital.xlsx"'
    wb.save(response)
    return response


# Exporta en excel el detalle de la encuesta por id del agente a encuesta única
@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_oficina_digital")
def encuesta_satisfaccion_detalle_of_dig_exportar(request, respuesta_id):
    try:
        filas = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_detalle(respuesta_id)
    except Exception:
        filas = []

    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado = filas[0]

    stats_agente       = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta  = encuesta_oficina_digital.ReporteEncuestaOficinaDigital.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 3) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 3) * 100) if promedio_encuesta else 0

    from openpyxl.styles import Border, Side
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Encuesta Satisfacción Oficina Digital"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    # ── Anchos de columna ───────────────────────────────
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 2

    # ── HEADER (filas 1-4) ───────────────────────────────
    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    # Logo
    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = centro

    # Título
    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte de Encuesta de Satisfacción")
    c.font = font(bold=True, color="003FB7", size=15)
    c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9)
    c.alignment = izq

    # Fecha
    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(bold=True, color="FFC900", size=8)
    c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11)
    c.alignment = der

    # Línea separadora
    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    # ── KPI CARDS (fila 5-6) ─────────────────────────────
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    kpi_cols   = [4, 5, 6]
    kpi_labels = ["PROMEDIO GENERAL AGENTE", "ENCUESTAS DEL AGENTE", "PROMEDIO ESTA ENCUESTA"]
    kpi_vals   = [f"{promedio_agente_pct}%", str(stats_agente["total_encuestas"]), f"{promedio_encuesta_pct}%"]
    kpi_fills  = ["003FB7", "E8EEFF", "8B1A0A"]
    kpi_fonts  = ["FFFFFF", "003FB7", "FFFFFF"]

    for i, col in enumerate(kpi_cols):
        c = ws.cell(row=5, column=col, value=kpi_labels[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=8)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

        c = ws.cell(row=6, column=col, value=kpi_vals[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=22)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

    # ── PANEL IZQUIERDO + PREGUNTAS (desde fila 8) ───────
    panel_items = [
        ("AGENTE RESPONSABLE",    encabezado.get("Agente",   ""), True),
        ("NOMBRE ACCIONISTA",     encabezado.get("Nombre",   "—"), True),
        ("CÉDULA",                encabezado.get("Cedula",   "—"), True),
        ("FECHA",                 str(encabezado.get("Fecha", "")), True),
    ]

    # Encabezado tabla preguntas
    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18

    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = izq

    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = centro

    # Filas de datos — panel izquierdo alineado con preguntas
    row_data = 9
    for i in range(max(len(panel_items) * 2, len(filas))):
        ws.row_dimensions[row_data + i].height = 22

    for i, (label, valor, _) in enumerate(panel_items):
        r_label = row_data + (i * 2)
        r_valor = row_data + (i * 2) + 1

        # Label
        c = ws.cell(row=r_label, column=2, value=label)
        c.fill = fill("FFF0C0")
        c.font = font(bold=True, color="003FB7", size=8)
        c.alignment = izq

        # Valor
        c = ws.cell(row=r_valor, column=2, value=valor)
        c.fill = fill("FFF8E0")
        c.font = font(color="1A1A1A", size=10)
        c.alignment = izq

    # Preguntas alineadas desde row_data
    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9)
        c.alignment = izq
        c.border = border_full("E0E0E0")

        respuesta = fila.get("Respuesta", "")
        color_resp = "4CAF50" if respuesta in ["Muy satisfecho", "Muy fácil", "Sí"] else "003FB7"
        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro
        c.border = border_full(color_resp)

    # ── PIE DE PÁGINA ────────────────────────────────────
    row_pie = row_data + max(len(panel_items) * 2, len(filas)) + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")

    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie+1, start_column=2, end_row=row_pie+1, end_column=6)
    c = ws.cell(row=row_pie+1, column=2, value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A", size=8)
    c.alignment = centro

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="encuesta_detalle_of_dig{respuesta_id}.xlsx"'
    wb.save(response)
    return response









# ── ENCUESTAS ───── EXPERIENCIA EN PÁGINA WEB─────────────────────────────────────────
# ── ENCUESTAS ───── EXPERIENCIA EN PÁGINA WEB─────────────────────────────────────────

# VISTA GENERAL DE LOS DATOS EN LA TABLA
@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web(request):
    filtros = {
        "nombre": request.GET.get("nombre"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin": request.GET.get("fecha_fin"),
        "sitio_evaluado": request.GET.get("sitio_evaluado"),
        "segmento_nps": request.GET.get("segmento_nps"),
    }

    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin = request.GET.get("kpi_fecha_fin")
    kpi_sitio_evaluado = request.GET.get("kpi_sitio_evaluado")

    datos = []
    opciones_nombre = []
    opciones_sitio = []
    kpis_globales = {}

    try:
        datos = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_datos_agrupados(filtros)

        kpis_globales = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_kpis_globales(
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
            sitio_evaluado=kpi_sitio_evaluado,
        )

        opciones_nombre = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_pagina_web WHERE Nombre IS NOT NULL ORDER BY Nombre"
        )

        opciones_sitio = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_opciones_sitio()

    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaPaginaWebTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_experiencia_web.html", {
        "table": table,
        "filtros": filtros,
        "kpis_globales": kpis_globales,
        "kpi_fecha_inicio": kpi_fecha_inicio,
        "kpi_fecha_fin": kpi_fecha_fin,
        "kpi_sitio_evaluado": kpi_sitio_evaluado,
        "opciones_nombre": opciones_nombre,
        "opciones_sitio": opciones_sitio,
    })

# VISTA PARA INGRESAR Y REFLEJAR LOS DATOS DE LOS DETALLES DE LA ENCUESTA
@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web_detalle(request, respuesta_id):
    # print("respuesta_id recibido:", respuesta_id)

    try:
        filas = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_detalle(respuesta_id)
        # print("filas detalle:", filas[:5] if filas else [])
    except Exception as e:
        # print("ERROR DETALLE:", e)
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    for fila in filas:
        pregunta = (fila.get("Pregunta") or "").lower()
        respuesta = fila.get("Respuesta", "")
        if "encontró" in pregunta or "encontro" in pregunta:
            if respuesta == "1":
                fila["Respuesta"] = "Sí"
            elif respuesta == "0":
                fila["Respuesta"] = "No"

    encabezado = filas[0] if filas else {}
    encabezado["Nombre"] = encabezado.get("Nombre") or "—"

    fecha = encabezado.get("Fecha")
    if fecha:
        encabezado["Fecha"] = fecha.strftime("%d/%m/%Y %H:%M")
    else:
        encabezado["Fecha"] = "—"

    promedio_encuesta = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_promedio_encuesta(respuesta_id)
    promedio_encuesta_nps = round((promedio_encuesta / 10) * 100) if promedio_encuesta else 0

    kpis = {
        "promedio_encuesta": promedio_encuesta,
        "promedio_pct": promedio_encuesta_nps,
        "nombre": encabezado.get("Nombre", ""),
        "fecha": encabezado.get("Fecha", ""),
    }

    return render(request, "dashboard/encuestas/satisfaccion_experiencia_web_detalle.html", {
        "encabezado": encabezado,
        "preguntas": filas,
        "respuesta_id": respuesta_id,
        "kpis": kpis,
    })


# FUNCION PARA EXPORTAR LOS DATOS A EXCEL
@login_required
@permiso_requerido("dashboard.view_encuesta_experiencia_web")
def encuesta_experiencia_web_exportar(request):
    filtros = {
        "nombre": request.GET.get("nombre"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin": request.GET.get("fecha_fin"),
        "sitio_evaluado": request.GET.get("sitio_evaluado"),
        "segmento_nps": request.GET.get("segmento_nps"),
    }

    datos = encuesta_pagina_web.ReporteEncuestaPaginaWeb.obtener_datos(filtros)

    encuestas = {}
    preguntas_orden = []

    for fila in datos:
        rid = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        respuesta = fila.get("Respuesta", "")

        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha": fila.get("Fecha", ""),
                "Nombre": fila.get("Nombre", ""),
            }

        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)

        encuestas[rid][pregunta] = respuesta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Experiencia Página Web"

    cols_fijas = ["ID", "Fecha", "Nombre"]
    keys_fijas = ["respuesta_id", "Fecha", "Nombre"]

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados fijos
    for col_idx, col_name in enumerate(cols_fijas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados dinámicos de preguntas
    for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = PatternFill("solid", fgColor="FFC900")
        cell.font = Font(bold=True, color="1A1000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos
    for row_idx, (rid, enc) in enumerate(encuestas.items(), start=2):
        for col_idx, key in enumerate(keys_fijas, start=1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")

        for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")

            # Transformar 1/0 a Sí/No para la pregunta de información
            if "encontró" in pregunta.lower() or "encontro" in pregunta.lower():
                if str(respuesta) == "1":
                    respuesta = "Sí"
                elif str(respuesta) == "0":
                    respuesta = "No"

            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="experiencia_pagina_web.xlsx"'

    wb.save(response)
    return response





# ── ENCUESTAS ───── ENCUESTA SATISFACCION WHATSAPP AGENTE ────────────────────────
# ── ENCUESTAS ───── ENCUESTA SATISFACCION WHATSAPP AGENTE ────────────────────────


@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsApp_agente(request):
    filtros = {
        "agente":          request.GET.get("agente"),
        "nombre":          request.GET.get("nombre"),
        "cedula":          request.GET.get("cedula"),
        "fecha_inicio":    request.GET.get("fecha_inicio"),
        "fecha_fin":       request.GET.get("fecha_fin"),
        "clasificacion":   request.GET.get("clasificacion"),
    }

    kpi_agentes      = request.GET.getlist("kpi_agente")
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    datos = []
    opciones_agente = opciones_nombre = []
    kpis_globales = {}

    try:
        datos         = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_datos_agrupados(filtros)
        
        kpis_globales = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_kpis_globales(
            agentes=kpi_agentes if kpi_agentes else None,
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
        opciones_agente = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Agente FROM dbo.vw_reporte_encuestas_satisfaccion_whatsapp_agente WHERE Agente IS NOT NULL ORDER BY Agente"
        )
        opciones_nombre = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_whatsapp_agente WHERE Nombre IS NOT NULL ORDER BY Nombre"
        )
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaWhatsappAgenteTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp_agente.html", {
        "table":            table,
        "filtros":          filtros,
        "kpis_globales":    kpis_globales,
        "kpi_agentes":      kpi_agentes,
        "kpi_fecha_inicio": kpi_fecha_inicio,
        "kpi_fecha_fin":    kpi_fecha_fin,
        "opciones_agente":  opciones_agente,
        "opciones_nombre":  opciones_nombre,
    })




@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_detalle(request, respuesta_id):
    try:
        filas = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_detalle(respuesta_id)
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    encabezado = filas[0] if filas else {}

    stats_agente      = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_promedio_agente(
        encabezado.get("Agente", "")
    )
    promedio_encuesta = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 5) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    kpis = {
        "promedio_general":  promedio_agente_pct,
        "total_encuestas":   stats_agente["total_encuestas"],
        "promedio_encuesta": promedio_encuesta_pct,
        "agente":            encabezado.get("Agente", ""),
    }

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp_agente_detalle.html", {
        "encabezado":   encabezado,
        "preguntas":    filas,
        "respuesta_id": respuesta_id,
        "kpis":         kpis,
    })



@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_exportar(request):
    filtros = {
        "agente":       request.GET.get("agente"),
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }

    datos = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_datos(filtros)

    encuestas = {}
    preguntas_orden = []

    for fila in datos:
        rid      = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        respuesta = fila.get("Respuesta", "")

        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha":   fila.get("Fecha", ""),
                "Agente":  fila.get("Agente", ""),
                "Nombre":  fila.get("Nombre", ""),
                "Cedula":  fila.get("Cedula", ""),
            }

        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)

        encuestas[rid][pregunta] = respuesta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WhatsApp Agente"

    cols_fijas = ["ID", "Fecha", "Agente", "Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha", "Agente", "Nombre", "Cedula"]

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(cols_fijas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = PatternFill("solid", fgColor="FFC900")
        cell.font = Font(bold=True, color="1A1000")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, (rid, enc) in enumerate(encuestas.items(), start=2):
        for col_idx, key in enumerate(keys_fijas, start=1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")

        for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="whatsapp_agente.xlsx"'
    wb.save(response)
    return response



@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp_agente")
def encuesta_whatsapp_agente_detalle_exportar(request, respuesta_id):
    try:
        filas = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_detalle(respuesta_id)
    except Exception:
        filas = []

    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado = filas[0]

    stats_agente      = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_promedio_agente(encabezado.get("Agente", ""))
    promedio_encuesta = encuesta_whatsapp_agente.ReporteEncuestaWhatsappAgente.obtener_promedio_encuesta(respuesta_id)
    promedio_agente_pct   = round((stats_agente["promedio_agente"] / 5) * 100) if stats_agente["promedio_agente"] else 0
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    from openpyxl.styles import Border, Side
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte WhatsApp Agente"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 2

    # ── HEADER ───────────────────────────────────────────
    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = centro

    ws.merge_cells("D2:E2")
    c = ws.cell(row=2, column=4, value="Reporte Encuesta WhatsApp Agente")
    c.font = font(bold=True, color="003FB7", size=15)
    c.alignment = izq

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9)
    c.alignment = izq

    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(bold=True, color="FFC900", size=8)
    c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11)
    c.alignment = der

    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    # ── KPI CARDS ────────────────────────────────────────
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    kpi_cols   = [4, 5, 6]
    kpi_labels = ["PROMEDIO GENERAL AGENTE", "ENCUESTAS DEL AGENTE", "PROMEDIO ESTA ENCUESTA"]
    kpi_vals   = [f"{promedio_agente_pct}%", str(stats_agente["total_encuestas"]), f"{promedio_encuesta_pct}%"]
    kpi_fills  = ["003FB7", "E8EEFF", "8B1A0A"]
    kpi_fonts  = ["FFFFFF", "003FB7", "FFFFFF"]

    for i, col in enumerate(kpi_cols):
        c = ws.cell(row=5, column=col, value=kpi_labels[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=8)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

        c = ws.cell(row=6, column=col, value=kpi_vals[i])
        c.fill = fill(kpi_fills[i])
        c.font = font(bold=True, color=kpi_fonts[i], size=22)
        c.alignment = centro
        c.border = border_full(kpi_fills[i])

    # ── PANEL IZQUIERDO ───────────────────────────────────
    panel_items = [
        ("AGENTE RESPONSABLE", encabezado.get("Agente", ""), True),
        ("NOMBRE ACCIONISTA",  encabezado.get("Nombre", "—"), True),
        ("CÉDULA",             encabezado.get("Cedula", "—"), True),
        ("FECHA",              str(encabezado.get("Fecha", "")), True),
        ("HORA",               str(encabezado.get("Hora", "")), True),
    ]

    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18

    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = izq

    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = centro

    row_data = 9
    for i in range(max(len(panel_items) * 2, len(filas))):
        ws.row_dimensions[row_data + i].height = 22

    for i, (label, valor, _) in enumerate(panel_items):
        r_label = row_data + (i * 2)
        r_valor = row_data + (i * 2) + 1

        c = ws.cell(row=r_label, column=2, value=label)
        c.fill = fill("FFF0C0")
        c.font = font(bold=True, color="003FB7", size=8)
        c.alignment = izq

        c = ws.cell(row=r_valor, column=2, value=valor)
        c.fill = fill("FFF8E0")
        c.font = font(color="1A1A1A", size=10)
        c.alignment = izq

    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9)
        c.alignment = izq
        c.border = border_full("E0E0E0")

        respuesta = fila.get("Respuesta", "")
        try:
            val = float(respuesta)
            color_resp = "4CAF50" if val >= 4 else "FFC107" if val == 3 else "F44336"
        except (ValueError, TypeError):
            color_resp = "003FB7"

        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro
        c.border = border_full(color_resp)

    # ── PIE DE PÁGINA ────────────────────────────────────
    row_pie = row_data + max(len(panel_items) * 2, len(filas)) + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")

    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie+1, start_column=2, end_row=row_pie+1, end_column=6)
    c = ws.cell(row=row_pie+1, column=2, value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A", size=8)
    c.alignment = centro

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="whatsapp_agente_detalle_{respuesta_id}.xlsx"'
    wb.save(response)
    return response






# ── ENCUESTAS ───── ENCUESTA SATISFACCION WHATSAPP  ────────────────────────
# ── ENCUESTAS ───── ENCUESTA SATISFACCION WHATSAPP  ────────────────────────

@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp")
def encuesta_whatsApp_(request):
    filtros = {
        "nombre":          request.GET.get("nombre"),
        "cedula":          request.GET.get("cedula"),
        "fecha_inicio":    request.GET.get("fecha_inicio"),
        "fecha_fin":       request.GET.get("fecha_fin"),
        "clasificacion":   request.GET.get("clasificacion"),
    }

    
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin")

    datos = []
    opciones_nombre = []
    kpis_globales = {}

    try:
        datos         = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_datos_agrupados(filtros)
        
        kpis_globales = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_kpis_globales(
            
            fecha_inicio=kpi_fecha_inicio,
            fecha_fin=kpi_fecha_fin,
        )
        
        opciones_nombre = ReportesDBService.ejecutar_query(
            "SELECT DISTINCT Nombre FROM dbo.vw_reporte_encuestas_satisfaccion_WhatsApp_sin_agente WHERE Nombre IS NOT NULL ORDER BY Nombre"
        )
    except Exception as e:
        print(">>> ERROR:", e)
        messages.error(request, "Error al obtener los datos.")

    table = EncuestaWhatsappTable(datos)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=10)
    except Exception:
        table.paginate(page=1, per_page=10)

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp.html", {
        "table":            table,
        "filtros":          filtros,
        "kpis_globales":    kpis_globales,
        "kpi_fecha_inicio": kpi_fecha_inicio,
        "kpi_fecha_fin":    kpi_fecha_fin,
        "opciones_nombre":  opciones_nombre,
    })




@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp")
def encuesta_whatsapp_detalle(request, respuesta_id):
    try:
        filas = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_detalle(respuesta_id)
        # print(f"Este es el id: {respuesta_id}")
    except Exception:
        filas = []
        messages.error(request, "Error al obtener el detalle.")

    encabezado = filas[0] if filas else {}
    print("Encabezado:", encabezado)
    
    promedio_encuesta = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_promedio_encuesta(respuesta_id)
    
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    kpis = {
        "promedio_encuesta": promedio_encuesta_pct,
    }

    return render(request, "dashboard/encuestas/satisfaccion_whatsapp_detalle.html", {
        "encabezado":   encabezado,
        "preguntas":    filas,
        "respuesta_id": respuesta_id,
        "kpis":         kpis,
    })



@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp")
def encuesta_whatsapp_exportar(request):
    filtros = {
        "nombre":       request.GET.get("nombre"),
        "cedula":       request.GET.get("cedula"),
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }

    datos = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_datos(filtros)

    encuestas = {}
    preguntas_orden = []

    for fila in datos:
        rid      = fila["respuesta_id"]
        pregunta = fila.get("Pregunta", "")
        respuesta = fila.get("Respuesta", "")

        if rid not in encuestas:
            encuestas[rid] = {
                "respuesta_id": rid,
                "Fecha":   fila.get("Fecha", ""),
                "Nombre":  fila.get("Nombre", ""),
                "Cedula":  fila.get("Cedula", ""),
            }

        if pregunta and pregunta not in preguntas_orden:
            preguntas_orden.append(pregunta)

        encuestas[rid][pregunta] = respuesta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WhatsApp"

    cols_fijas = ["ID", "Fecha","Accionista", "Cédula"]
    keys_fijas = ["respuesta_id", "Fecha","Nombre", "Cedula"]

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(cols_fijas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = PatternFill("solid", fgColor="FFC900")
        cell.font = Font(bold=True, color="1A1000")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, (rid, enc) in enumerate(encuestas.items(), start=2):
        for col_idx, key in enumerate(keys_fijas, start=1):
            valor = enc.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else "")

        for i, pregunta in enumerate(preguntas_orden, start=len(cols_fijas) + 1):
            respuesta = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(respuesta) if respuesta else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="whatsapp.xlsx"'
    wb.save(response)
    return response



@login_required
@permiso_requerido("dashboard.view_encuesta_satisfaccion_whatsapp")
def encuesta_whatsapp_detalle_exportar(request, respuesta_id):
    try:
        filas = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_detalle(respuesta_id)
    except Exception:
        filas = []

    if not filas:
        return HttpResponse("Sin datos", status=404)

    encabezado = filas[0]

    promedio_encuesta = encuesta_whatsapp.ReporteEncuestaWhatsapp.obtener_promedio_encuesta(respuesta_id)
    promedio_encuesta_pct = round((promedio_encuesta / 5) * 100) if promedio_encuesta else 0

    from openpyxl.styles import Border, Side
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte WhatsApp"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border_full(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="1A1A1A", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    der    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 2

    # ── HEADER ───────────────────────────────────────────
    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5

    ws.merge_cells("B2:B3")
    c = ws.cell(row=2, column=2, value="CAJA DE\nANDE")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = centro

    ws.merge_cells("D3:E3")
    c = ws.cell(row=3, column=4, value=f"ID Reporte: #{respuesta_id}")
    c.font = font(color="74788A", size=9)
    c.alignment = izq

    c = ws.cell(row=2, column=6, value="FECHA DE GENERACIÓN")
    c.font = font(bold=True, color="FFC900", size=8)
    c.alignment = der

    c = ws.cell(row=3, column=6, value=datetime.now().strftime("%d de %B, %Y"))
    c.font = font(bold=True, color="003FB7", size=11)
    c.alignment = der

    ws.row_dimensions[4].height = 4
    for col in range(2, 7):
        ws.cell(row=4, column=col).fill = fill("003FB7")

    # ── KPI CARD (solo promedio encuesta) ────────────────
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 5

    # Label
    c = ws.cell(row=5, column=4, value="PROMEDIO ESTA ENCUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=8)
    c.alignment = centro
    c.border = border_full("E8EEFF")

    # Valor
    c = ws.cell(row=6, column=4, value=f"{promedio_encuesta_pct}%")
    c.fill = fill("003FB7")
    c.font = font(bold=True, color="FFFFFF", size=22)
    c.alignment = centro
    c.border = border_full("003FB7")

    # ── ENCABEZADO TABLA ─────────────────────────────────
    panel_items = [
        ("NOMBRE ACCIONISTA", encabezado.get("Nombre", "—")),
        ("CÉDULA",            encabezado.get("Cedula", "—")),
        ("FECHA",             str(encabezado.get("Fecha", ""))),
        ("HORA",              str(encabezado.get("Hora", ""))),
    ]

    row_enc_tabla = 8
    ws.row_dimensions[row_enc_tabla].height = 18

    ws.merge_cells(start_row=row_enc_tabla, start_column=4, end_row=row_enc_tabla, end_column=5)
    c = ws.cell(row=row_enc_tabla, column=4, value="PREGUNTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = izq

    c = ws.cell(row=row_enc_tabla, column=6, value="RESPUESTA")
    c.fill = fill("E8EEFF")
    c.font = font(bold=True, color="003FB7", size=9)
    c.alignment = centro

    # ── DATOS ────────────────────────────────────────────
    row_data = 9
    total_filas = max(len(panel_items) * 2, len(filas))
    for i in range(total_filas):
        ws.row_dimensions[row_data + i].height = 22

    # Panel izquierdo (info accionista)
    for i, (label, valor) in enumerate(panel_items):
        r_label = row_data + (i * 2)
        r_valor = row_data + (i * 2) + 1

        c = ws.cell(row=r_label, column=2, value=label)
        c.fill = fill("FFF0C0")
        c.font = font(bold=True, color="003FB7", size=8)
        c.alignment = izq

        c = ws.cell(row=r_valor, column=2, value=valor)
        c.fill = fill("FFF8E0")
        c.font = font(color="1A1A1A", size=10)
        c.alignment = izq

    # Panel derecho (preguntas y respuestas)
    for i, fila in enumerate(filas):
        r = row_data + i
        ws.row_dimensions[r].height = 28

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c = ws.cell(row=r, column=4, value=fila.get("Pregunta", ""))
        c.fill = fill("FFFFFF") if i % 2 == 0 else fill("F4F7FF")
        c.font = font(color="1A1A1A", size=9)
        c.alignment = izq
        c.border = border_full("E0E0E0")

        respuesta = fila.get("Respuesta", "")
        try:
            val = float(respuesta)
            color_resp = "4CAF50" if val >= 4 else "FFC107" if val == 3 else "F44336"
        except (ValueError, TypeError):
            color_resp = "003FB7"

        c = ws.cell(row=r, column=6, value=respuesta)
        c.fill = fill(color_resp)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = centro
        c.border = border_full(color_resp)

    # ── PIE DE PÁGINA ────────────────────────────────────
    row_pie = row_data + total_filas + 1
    ws.row_dimensions[row_pie].height = 4
    for col in range(2, 7):
        ws.cell(row=row_pie, column=col).fill = fill("FFC900")

    ws.row_dimensions[row_pie + 1].height = 14
    ws.merge_cells(start_row=row_pie+1, start_column=2, end_row=row_pie+1, end_column=6)
    c = ws.cell(row=row_pie+1, column=2,
                value="© 2026 Caja de Ande · Documento generado automáticamente · Confidencial")
    c.font = font(color="74788A", size=8)
    c.alignment = centro

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="whatsapp_detalle_{respuesta_id}.xlsx"'
    wb.save(response)
    return response









# ── ENCUESTAS ───── ENCUESTA SATISFACCION SALUD FINANCIERA ────────────────────────
# ── ENCUESTAS ───── ENCUESTA SATISFACCION SALUD FINANCIERA ────────────────────────

def build_encuesta_feria_salud_table(datos, preguntas_orden):
    attrs = {
        "Fecha": tables.Column(verbose_name="Fecha"),
    }
    for pregunta in preguntas_orden:
        attrs[pregunta] = tables.Column(
            verbose_name=pregunta,
            default="—",
            orderable=False,
        )

    Meta = type("Meta", (), {
        "template_name": "dashboard/components/table.html",
        "attrs":         {"class": "w-full text-left border-collapse"},
        "row_attrs":     {"class": "hover:bg-surface-container-low transition-colors cursor-pointer"},
        "sequence":      ("Fecha", *preguntas_orden),
    })
    attrs["Meta"] = Meta

    TableClass = type("EncuestaFeriaSaludDynamicTable", (tables.Table,), attrs)
    return TableClass(datos)
@login_required
@permiso_requerido("dashboard.view_encuesta_feria_salud")
def encuesta_feria_salud_(request):
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio", ""),
        "fecha_fin":    request.GET.get("fecha_fin", ""),
    }
    kpi_fecha_inicio = request.GET.get("kpi_fecha_inicio", "")
    kpi_fecha_fin    = request.GET.get("kpi_fecha_fin", "")

    try:
        datos, preguntas_orden = encuesta_feria_salud.ReporteEncuestaFeriaSalud.obtener_datos_agrupados(filtros)
        # print(f"DEBUG: {len(datos)} encuestas agrupadas")
    except Exception:
        datos, preguntas_orden = [], []

    try:
        total = encuesta_feria_salud.ReporteEncuestaFeriaSalud.obtener_total(
            filtros.get("fecha_inicio"), filtros.get("fecha_fin")
        )
    except Exception:
        total = 0

    table = build_encuesta_feria_salud_table(datos, preguntas_orden)
    try:
        table.paginate(page=request.GET.get("page", 1), per_page=25)
    except Exception:
        table.paginate(page=1, per_page=25)

    return render(request, "dashboard/encuestas/satisfaccion_feria_salud.html", {
        "table":            table,
        "filtros":          filtros,
        "total_encuestas":  total,
        "kpi_fecha_inicio": kpi_fecha_inicio,
        "kpi_fecha_fin":    kpi_fecha_fin,
    })





@login_required
@permiso_requerido("dashboard.view_encuesta_feria_salud")
def encuesta_feria_salud_exportar(request):
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio"),
        "fecha_fin":    request.GET.get("fecha_fin"),
    }

    try:
        datos, preguntas_orden = encuesta_feria_salud.ReporteEncuestaFeriaSalud.obtener_datos_agrupados(filtros)
    except Exception:
        datos, preguntas_orden = [], []

    if not datos:
        return HttpResponse("Sin datos", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feria de la Salud"

    header_fill = PatternFill("solid", fgColor="003FB7")
    header_font = Font(bold=True, color="FFFFFF")
    pregunta_fill = PatternFill("solid", fgColor="FFC900")
    pregunta_font = Font(bold=True, color="1A1000")

    # Columnas fijas
    ws.cell(row=1, column=1, value="ID").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=1, column=2, value="Fecha").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    # Columnas dinámicas (una por pregunta)
    for i, pregunta in enumerate(preguntas_orden, start=3):
        cell = ws.cell(row=1, column=i, value=pregunta)
        cell.fill = pregunta_fill
        cell.font = pregunta_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Filas de datos
    for row_idx, enc in enumerate(datos, start=2):
        ws.cell(row=row_idx, column=1, value=enc.get("respuesta_id", ""))
        ws.cell(row=row_idx, column=2, value=str(enc.get("Fecha", "")))
        for i, pregunta in enumerate(preguntas_orden, start=3):
            valor = enc.get(pregunta, "")
            ws.cell(row=row_idx, column=i, value=str(valor) if valor else "")

    # Ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="feria_salud.xlsx"'
    wb.save(response)
    return response






# ── FORMULARIOS ────────────────────────────────────────────
# ── FORMULARIOS ─────────TARJETAS───────────────────────────────────



@login_required
@permiso_requerido("dashboard.view_formulario_tarjetas")
def formulario_tarjetas(request):
    return render(request, "dashboard/formularios/tarjetas.html", {
        "active_tab": "tarjetas"
    })

# ─────────────────────────────────────────────
#  Campos permitidos para la búsqueda AJAX
#  clave URL → columna real en la BD
# ─────────────────────────────────────────────
CAMPOS_TARJETA = {
    "cedula":   "Cedula",
    "nombre":   "Nombre",
    "telefono": "Telefono",
    "correo":   "Correo",
}


# ----- SOLICITUD TARJETA CREDITO -----

@login_required
def soli_tarj_credito_buscar(request, campo):
    """
    Endpoint AJAX para Select2.
    URL: /dashboard/formularios/tarjetas/solicitudes/buscar/<campo>/
    Devuelve: { results: [{id, text}, ...] }
    """
    if campo not in CAMPOS_TARJETA:
        return JsonResponse({"results": []})
 
    col_db = CAMPOS_TARJETA[campo]
    term   = request.GET.get("term", "").strip()
 
    filas = ReporteSolicitudTarjetaCredito.buscar_opciones(col_db, term)
    data  = [
        {"id": (row[col_db] or "").strip(), "text": (row[col_db] or "").strip()}
        for row in filas
        if row.get(col_db) and row[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
# ─────────────────────────────────────────────
#  LISTA — liviana, sin pre-cargar opciones
# ─────────────────────────────────────────────
@permiso_requerido("dashboard.view_soli_tarj_credito")
def soli_tarj_credito_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "telefono":     request.GET.get("telefono",     "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "tipo_tramite": request.GET.get("tipo_tramite", "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos    = ReporteSolicitudTarjetaCredito.obtener_datos(filtros)
    kpis     = ReporteSolicitudTarjetaCredito.obtener_kpis(filtros)
    tramites = ReporteSolicitudTarjetaCredito.obtener_por_tramite()
 
    table = SolicitudTarjetaCreditoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_tarj_credito_lista.html", {
        "table":    table,
        "filtros":  filtros,
        "kpis":     kpis,
        "tramites": tramites,
    })
 
 
# ─────────────────────────────────────────────
#  DETALLE
# ─────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_soli_tarj_credito")
def soli_tarj_credito_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudTarjetaCredito.obtener_detalle(respuesta_id)
    if not solicitud:
        from django.http import Http404
        raise Http404
 
    return render(request, "dashboard/formularios/reportes/soli_tarj_credito_detalle.html", {
        "solicitud": solicitud,
    })
 
 
# ─────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_soli_tarj_credito")
def soli_tarj_credito_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula", "").strip(),
        "nombre":       request.GET.get("nombre", "").strip(),
        "tipo_tramite": request.GET.get("tipo_tramite", "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin", "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudTarjetaCredito.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes Tarjeta"
 
    AZUL     = "003FB7"
    AMARILLO = "FFC900"
 
    headers = [
        "ID", "Fecha / Hora", "Cédula", "Nombre",
        "Teléfono", "Correo", "Tipo de Trámite",
        "Dirección de Envío", "URL Cédula Frente", "URL Cédula Reverso",
    ]
 
    header_fill = PatternFill("solid", fgColor=AZUL)
    header_font = Font(bold=True, color="FFFFFF")
 
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
 
    for row_idx, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
 
        ws.cell(row=row_idx, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=row_idx, column=2,  value=fecha_str)
        ws.cell(row=row_idx, column=3,  value=row.get("Cedula"))
        ws.cell(row=row_idx, column=4,  value=row.get("Nombre"))
        ws.cell(row=row_idx, column=5,  value=row.get("Telefono"))
        ws.cell(row=row_idx, column=6,  value=row.get("Correo"))
        ws.cell(row=row_idx, column=7,  value=row.get("TipoTramite"))
        ws.cell(row=row_idx, column=8,  value=row.get("DireccionEnvio"))
        ws.cell(row=row_idx, column=9,  value=row.get("URL_Cedula_Frente"))
        ws.cell(row=row_idx, column=10, value=row.get("URL_Cedula_Reverso"))
 
        if row_idx % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 11):
                ws.cell(row=row_idx, column=c).fill = fila_fill
 
    col_widths = [8, 18, 14, 30, 14, 30, 22, 40, 50, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="solicitudes_tarjeta_{fecha_str}.xlsx"'
    wb.save(response)
    return response


 
# ─────────────────────────────────────────────
#  EXPORT EXCEL Detalle 1 solicitud
# ─────────────────────────────────────────────

@login_required
@permiso_requerido("dashboard.view_soli_tarj_credito")
def soli_tarj_credito_export_detalle(request, respuesta_id):
    """
    Exporta una sola solicitud con el mismo layout del detalle:
    - Encabezado institucional
    - Sección Datos del Solicitante (azul)
    - Sección Datos del Trámite (amarillo)
    - Sección Documentos ShareFile
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, PatternFill
    )
    from openpyxl.utils import get_column_letter
    from datetime import datetime
 
    solicitud = ReporteSolicitudTarjetaCredito.obtener_detalle(respuesta_id)
    if not solicitud:
        from django.http import Http404
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Solicitud #{respuesta_id}"
 
    # ── Constantes de color ───────────────────────
    AZUL        = "003FB7"
    AMARILLO    = "FFC900"
    BLANCO      = "FFFFFF"
    GRIS_LABEL  = "F1F5F9"   # fondo de la etiqueta
    GRIS_BORDE  = "E2E8F0"
    NEGRO       = "1E293B"
 
    # ── Helpers ───────────────────────────────────
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
 
    def border_bottom():
        return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(ws, row, texto, color_fondo, color_texto=BLANCO, icono=""):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1,
                    value=f"  {icono}  {texto}" if icono else f"  {texto}")
        c.fill    = fill(color_fondo)
        c.font    = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(ws, row, label, valor):
        # Columna A-B: label
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill      = fill(GRIS_LABEL)
        lc.font      = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border    = border_bottom()
 
        # Columna C-D: valor
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor or "—")
        vc.font      = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border    = border_bottom()
 
        ws.row_dimensions[row].height = 22
 
    def spacer(ws, row):
        ws.row_dimensions[row].height = 8
 
    # ── Anchos de columna ─────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25
 
    r = 1  # cursor de fila
 
    # ── ENCABEZADO INSTITUCIONAL ──────────────────
    ws.merge_cells(f"A{r}:D{r}")
    titulo = ws.cell(row=r, column=1,
                     value="CAJA DE ANDE — Solicitud de Tarjeta de Crédito")
    titulo.fill      = fill(AZUL)
    titulo.font      = Font(bold=True, color=BLANCO, size=13, name="Arial")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36
    r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    sub = ws.cell(row=r, column=1,
                  value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    sub.fill      = fill("002A80")
    sub.font      = Font(color="93C5FD", size=9, name="Arial")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1
 
    spacer(ws, r); r += 1
 
    # ── SECCIÓN 1: DATOS DEL SOLICITANTE ─────────
    header_row(ws, r, "DATOS DEL SOLICITANTE", AZUL, BLANCO)
    r += 1
 
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "—")
 
    data_row(ws, r, "CÉDULA",    (solicitud.get("Cedula") or "").strip()); r += 1
    data_row(ws, r, "NOMBRE",    (solicitud.get("Nombre") or "").strip()); r += 1
    data_row(ws, r, "TELÉFONO",  (solicitud.get("Telefono") or "").strip()); r += 1
    data_row(ws, r, "CORREO",    (solicitud.get("Correo") or "").strip()); r += 1
 
    spacer(ws, r); r += 1
 
    # ── SECCIÓN 2: DATOS DEL TRÁMITE ─────────────
    header_row(ws, r, "DATOS DEL TRÁMITE", AMARILLO, NEGRO)
    r += 1
 
    data_row(ws, r, "TIPO DE TRÁMITE",     solicitud.get("TipoTramite")); r += 1
    data_row(ws, r, "DIRECCIÓN DE ENVÍO",  solicitud.get("DireccionEnvio")); r += 1
    data_row(ws, r, "FECHA / HORA",        fecha_str); r += 1
 
    spacer(ws, r); r += 1
 
    # ── SECCIÓN 3: DOCUMENTOS SHAREFILE ──────────
    header_row(ws, r, "DOCUMENTOS ADJUNTOS — SHAREFILE", AZUL, BLANCO)
    r += 1
 
    # Cédula Frente
    ws.merge_cells(f"A{r}:B{r}")
    lf = ws.cell(row=r, column=1, value="CÉDULA — FRENTE")
    lf.fill      = fill(GRIS_LABEL)
    lf.font      = Font(bold=True, color="64748B", size=9, name="Arial")
    lf.alignment = Alignment(vertical="center", indent=2)
    lf.border    = border_bottom()
 
    ws.merge_cells(f"C{r}:D{r}")
    url_frente = (solicitud.get("URL_Cedula_Frente") or "").strip()
    vf = ws.cell(row=r, column=3, value=url_frente or "No disponible")
    if url_frente:
        vf.hyperlink = url_frente
        vf.font = Font(color="003FB7", underline="single", size=10, name="Arial")
    else:
        vf.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
    vf.alignment = Alignment(vertical="center", indent=2)
    vf.border    = border_bottom()
    ws.row_dimensions[r].height = 22
    r += 1
 
    # Cédula Reverso
    ws.merge_cells(f"A{r}:B{r}")
    lr = ws.cell(row=r, column=1, value="CÉDULA — REVERSO")
    lr.fill      = fill(GRIS_LABEL)
    lr.font      = Font(bold=True, color="64748B", size=9, name="Arial")
    lr.alignment = Alignment(vertical="center", indent=2)
    lr.border    = border_bottom()
 
    ws.merge_cells(f"C{r}:D{r}")
    url_reverso = (solicitud.get("URL_Cedula_Reverso") or "").strip()
    vr = ws.cell(row=r, column=3, value=url_reverso or "No disponible")
    if url_reverso:
        vr.hyperlink = url_reverso
        vr.font = Font(color="003FB7", underline="single", size=10, name="Arial")
    else:
        vr.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
    vr.alignment = Alignment(vertical="center", indent=2)
    vr.border    = border_bottom()
    ws.row_dimensions[r].height = 22
    r += 1
 
    spacer(ws, r); r += 1
 
    # ── PIE ───────────────────────────────────────
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill      = fill(GRIS_LABEL)
    pie.font      = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    # ── Freeze panes & zoom ───────────────────────
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    # ── Respuesta HTTP ────────────────────────────
    nombre = (solicitud.get("Nombre") or "solicitud").strip().replace(" ", "_")
    filename = f"solicitud_tarjeta_{respuesta_id}_{nombre}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






#---------------------------------Solicitud de tarjeta Ciudadano de Oro-------------
 
# ══════════════════════════════════════════════
 
CAMPOS_TARJETA_DEBITO_CIUDADANO_ORO = {
    "cedula":   "Cedula",
    "nombre":   "Nombre",
    "telefono": "Teléfono",
    "correo":   "Correo",
}
 
@login_required

def soli_tarj_debito_buscar(request, campo):
    if campo not in CAMPOS_TARJETA_DEBITO_CIUDADANO_ORO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_TARJETA_DEBITO_CIUDADANO_ORO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudTarjetaDebito.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito")
def soli_tarj_debito_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "telefono":     request.GET.get("telefono",     "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros  = {k: v for k, v in filtros.items() if v}
    datos    = ReporteSolicitudTarjetaDebito.obtener_datos(filtros)
    kpis     = ReporteSolicitudTarjetaDebito.obtener_kpis(filtros)
    destinos = ReporteSolicitudTarjetaDebito.obtener_por_destino()
 
    table = SolicitudTarjetaDebito_CiuOroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_tarj_debito_ciu_oro.html", {
        "table":    table,
        "filtros":  filtros,
        "kpis":     kpis,
        "destinos": destinos,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito")
def soli_tarj_debito_detalle(request, respuesta_id):
    filas = ReporteSolicitudTarjetaDebito.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404
    # La primera fila tiene los datos del accionista; todas tienen preguntas
    accionista = filas[0]
    return render(request, "dashboard/formularios/reportes/soli_tarj_debito_ciu_oro_detalle.html", {
        "accionista": accionista,
        "filas":      filas,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito")
def soli_tarj_debito_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "telefono":     request.GET.get("telefono",     "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudTarjetaDebito.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes Débito C.Oro"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Teléfono", "Correo", "Dirección de Envío"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
 
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre") or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Telefono") or "").strip())
        ws.cell(row=ri, column=6, value=(row.get("Correo") or "").strip())
        ws.cell(row=ri, column=7, value=row.get("DireccionEnvio"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 8):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 30, 14, 30, 25], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="solicitudes_debito_{fstr}.xlsx"'
    wb.save(response)
    return response





#--------------------------------Solicitud de tarjeta de debito común--------------
CAMPOS_DEBITO_GESTION = {
    "cedula":   "Cedula",
    "nombre":   "Nombre",
    "telefono": "Telefono",
    "correo":   "Correo",
}
 
@login_required

def soli_tarj_debito_gestion_buscar(request, campo):
    if campo not in CAMPOS_DEBITO_GESTION:
        return JsonResponse({"results": []})
    col_db = CAMPOS_DEBITO_GESTION[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudTarjetaDebitoGestion.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito_gestion")
def soli_tarj_debito_gestion_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "telefono":     request.GET.get("telefono",     "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "tipo_tramite": request.GET.get("tipo_tramite", "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos         = ReporteSolicitudTarjetaDebitoGestion.obtener_datos(filtros)
    kpis          = ReporteSolicitudTarjetaDebitoGestion.obtener_kpis(filtros)
    tipos_tramite = ReporteSolicitudTarjetaDebitoGestion.obtener_tipos_tramite()
    destinos      = ReporteSolicitudTarjetaDebitoGestion.obtener_destinos()
 
    table = SolicitudTarjetaDebitoGestionTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_tarj_debito_gestion.html", {
        "table":         table,
        "filtros":       filtros,
        "kpis":          kpis,
        "tipos_tramite": tipos_tramite,
        "destinos":      destinos,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito_gestion")
def soli_tarj_debito_gestion_detalle(request, respuesta_id):
    filas = ReporteSolicitudTarjetaDebitoGestion.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_tarj_debito_gestion_detalle.html", {
        "accionista": filas[0],
        "filas":      filas,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_tarj_debito_gestion")
def soli_tarj_debito_gestion_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "telefono":     request.GET.get("telefono",     "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "tipo_tramite": request.GET.get("tipo_tramite", "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudTarjetaDebitoGestion.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Gestión Tarjeta Débito"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Teléfono", "Correo", "Tipo de Trámite", "Dirección de Envío"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")    or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre")    or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Telefono")  or "").strip())
        ws.cell(row=ri, column=6, value=(row.get("Correo")    or "").strip())
        ws.cell(row=ri, column=7, value=row.get("TipoTramite"))
        ws.cell(row=ri, column=8, value=row.get("DireccionEnvio"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 9):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 30, 14, 30, 35, 30], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="gestion_tarjeta_debito_{fstr}.xlsx"'
    wb.save(response)
    return response





#--------------------------------Solicitud dpara redención de puntos--------------


CAMPOS_REDENCION = {
    "cedula":   "Cedula",
    "nombre":   "Nombre",
    "telefono": "Teléfono",
    "correo":   "Correo",
}
 
@login_required
def soli_redencion_puntos_buscar(request, campo):
    if campo not in CAMPOS_REDENCION:
        return JsonResponse({"results": []})
    col_db = CAMPOS_REDENCION[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudRedencionPuntos.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_redencion_puntos")
def soli_redencion_puntos_lista(request):
    filtros = {
        "cedula":         request.GET.get("cedula",         "").strip(),
        "nombre":         request.GET.get("nombre",         "").strip(),
        "telefono":       request.GET.get("telefono",       "").strip(),
        "correo":         request.GET.get("correo",         "").strip(),
        "tipo_redencion": request.GET.get("tipo_redencion", "").strip(),
        "fecha_inicio":   request.GET.get("fecha_inicio",   "").strip(),
        "fecha_fin":      request.GET.get("fecha_fin",      "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos   = ReporteSolicitudRedencionPuntos.obtener_datos(filtros)
    kpis    = ReporteSolicitudRedencionPuntos.obtener_kpis(filtros)
    tipos   = ReporteSolicitudRedencionPuntos.obtener_por_tipo()
 
    table = SolicitudRedencionPuntosTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_redencion_puntos.html", {
        "table":   table,
        "filtros": filtros,
        "kpis":    kpis,
        "tipos":   tipos,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_redencion_puntos")
def soli_redencion_puntos_detalle(request, respuesta_id):
    filas = ReporteSolicitudRedencionPuntos.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_redencion_puntos_detalle.html", {
        "accionista": filas[0],
        "filas":      filas,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_redencion_puntos")
def soli_redencion_puntos_export(request):
    filtros = {
        "cedula":         request.GET.get("cedula",         "").strip(),
        "nombre":         request.GET.get("nombre",         "").strip(),
        "telefono":       request.GET.get("telefono",       "").strip(),
        "correo":         request.GET.get("correo",         "").strip(),
        "tipo_redencion": request.GET.get("tipo_redencion", "").strip(),
        "fecha_inicio":   request.GET.get("fecha_inicio",   "").strip(),
        "fecha_fin":      request.GET.get("fecha_fin",      "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudRedencionPuntos.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Redención de Puntos"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Teléfono", "Correo", "Tipo de Redención"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")       or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre")       or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Telefono")     or "").strip())
        ws.cell(row=ri, column=6, value=(row.get("Correo")       or "").strip())
        ws.cell(row=ri, column=7, value=row.get("TipoRedencion"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 8):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 30, 14, 30, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="redencion_puntos_{fstr}.xlsx"'
    wb.save(response)
    return response

@login_required
@permiso_requerido("dashboard.view_soli_redencion_puntos")
def soli_redencion_puntos_export_detalle(request, respuesta_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filas = ReporteSolicitudRedencionPuntos.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404

    accionista = filas[0]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Redención #{respuesta_id}"

    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"

    def fill(hex): return PatternFill("solid", fgColor=hex)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))

    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28

    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor or "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25

    r = 1

    # Encabezado institucional
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Redención de Puntos")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1

    ws.merge_cells(f"A{r}:D{r}")
    s = ws.cell(row=r, column=1,
                value=f"ID #{accionista['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # Sección accionista
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    fecha = accionista.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    data_row(r, "CÉDULA",   (accionista.get("Cedula")   or "").strip()); r += 1
    data_row(r, "NOMBRE",   (accionista.get("Nombre")   or "").strip()); r += 1
    data_row(r, "TELÉFONO", (accionista.get("Telefono") or "").strip()); r += 1
    data_row(r, "CORREO",   (accionista.get("Correo")   or "").strip()); r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # Sección redención
    header_row(r, "DATOS DE LA REDENCIÓN", AMARILLO, NEGRO); r += 1
    for fila in filas:
        data_row(r, (fila.get("Pregunta") or "").upper(), fila.get("Respuesta")); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110

    nombre = (accionista.get("Nombre") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="redencion_puntos_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response





# ------------------------------Solicitud Caja de ANDE Asistencia ------------------


CAMPOS_ASISTENCIA = {
    "cedula":  "Cedula",
    "nombre":  "Nombre",
    "correo":  "Correo",
}
 
@login_required
def caja_ande_asistencia_buscar(request, campo):
    if campo not in CAMPOS_ASISTENCIA:
        return JsonResponse({"results": []})
    col_db = CAMPOS_ASISTENCIA[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteCajaAndeAsistencia.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_caja_ande_asistencia")
def caja_ande_asistencia_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "tipo_plan":    request.GET.get("tipo_plan",    "").strip(),
        "tipo_tarjeta": request.GET.get("tipo_tarjeta", "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros  = {k: v for k, v in filtros.items() if v}
    datos    = ReporteCajaAndeAsistencia.obtener_datos(filtros)
    kpis     = ReporteCajaAndeAsistencia.obtener_kpis(filtros)
    planes   = ReporteCajaAndeAsistencia.obtener_planes()
    tarjetas = ReporteCajaAndeAsistencia.obtener_tarjetas()
 
    table = CajaAndeAsistenciaTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_caja_ande_asistencia.html", {
        "table":    table,
        "filtros":  filtros,
        "kpis":     kpis,
        "planes":   planes,
        "tarjetas": tarjetas,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_caja_ande_asistencia")
def caja_ande_asistencia_detalle(request, respuesta_id):
    filas = ReporteCajaAndeAsistencia.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_caja_ande_asistencia_detalle.html", {
        "accionista": filas[0],
        "filas":      filas,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_caja_ande_asistencia")
def caja_ande_asistencia_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "correo":       request.GET.get("correo",       "").strip(),
        "tipo_plan":    request.GET.get("tipo_plan",    "").strip(),
        "tipo_tarjeta": request.GET.get("tipo_tarjeta", "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteCajaAndeAsistencia.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Caja ANDE Asistencia"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Correo", "Plan", "Tipo Tarjeta"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")      or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre")      or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Correo")      or "").strip())
        ws.cell(row=ri, column=6, value=row.get("TipoPlan"))
        ws.cell(row=ri, column=7, value=row.get("TipoTarjeta"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 8):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 30, 30, 55, 12], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="caja_ande_asistencia_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_caja_ande_asistencia")
def caja_ande_asistencia_export_detalle(request, respuesta_id):
    filas = ReporteCajaAndeAsistencia.obtener_detalle(respuesta_id)
    if not filas:
        raise Http404
 
    accionista = filas[0]
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencia #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex): return PatternFill("solid", fgColor=hex)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor or "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 15
 
    r = 1
 
    # Encabezado
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Caja de ANDE Asistencia")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = accionista.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{accionista['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Sección accionista
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",  (accionista.get("Cedula") or "").strip()); r += 1
    data_row(r, "NOMBRE",  (accionista.get("Nombre") or "").strip()); r += 1
    data_row(r, "CORREO",  (accionista.get("Correo") or "").strip()); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Sección formulario
    header_row(r, "DATOS DEL FORMULARIO", AMARILLO, NEGRO); r += 1
    for fila in filas:
        data_row(r, (fila.get("Pregunta") or "").upper(), fila.get("Respuesta")); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (accionista.get("Nombre") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="asistencia_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response









# ── FORMULARIOS ────────────AHORROS────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_ahorros")
def formulario_ahorros(request):
    return render(request, "dashboard/formularios/ahorros.html", {
        "active_tab": "ahorros"
    })


#-------------Solicitud Deposito de salario----------------------------

@login_required
def soli_deposito_salario_buscar_cedula(request):
    term  = request.GET.get("term", "").strip()
    filas = ReporteSolicitudDepositoSalario.buscar_cedulas(term)
    data  = [
        {"id": (r["Cedula"] or "").strip(), "text": (r["Cedula"] or "").strip()}
        for r in filas if r.get("Cedula") and r["Cedula"].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos = ReporteSolicitudDepositoSalario.obtener_datos(filtros)
    kpis  = ReporteSolicitudDepositoSalario.obtener_kpis(filtros)
 
    table = SolicitudDepositoSalarioTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_deposito_salario.html", {
        "table":   table,
        "filtros": filtros,
        "kpis":    kpis,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudDepositoSalario.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_deposito_salario_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudDepositoSalario.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Depósito de Salario"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula",
               "URL Boleta Solicitud", "URL Cédula Frente", "URL Cédula Reverso"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula") or "").strip())
        ws.cell(row=ri, column=4, value=row.get("BoletaSolicitud_URL"))
        ws.cell(row=ri, column=5, value=row.get("FrenteCedula_URL"))
        ws.cell(row=ri, column=6, value=row.get("ReversoCedula_URL"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 7):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 55, 55, 55], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="deposito_salario_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_deposito_salario")
def soli_deposito_salario_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudDepositoSalario.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Depósito Salario #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor or "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    def url_row(row, label, url):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=url or "No disponible")
        if url:
            vc.hyperlink = url
            vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
        else:
            vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 10
 
    r = 1
 
    # Encabezado institucional
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Depósito de Salario")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos del solicitante
    header_row(r, "DATOS DEL SOLICITANTE", AZUL); r += 1
    data_row(r, "CÉDULA",      (solicitud.get("Cedula") or "").strip()); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Documentos ShareFile
    header_row(r, "DOCUMENTOS ADJUNTOS — SHAREFILE", AMARILLO, NEGRO); r += 1
    url_row(r, "BOLETA DE SOLICITUD", solicitud.get("BoletaSolicitud_URL")); r += 1
    url_row(r, "CÉDULA — FRENTE",     solicitud.get("FrenteCedula_URL"));    r += 1
    url_row(r, "CÉDULA — REVERSO",    solicitud.get("ReversoCedula_URL"));   r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    cedula = (solicitud.get("Cedula") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="deposito_salario_{respuesta_id}_{cedula}.xlsx"'
    wb.save(response)
    return response





#------------------------Solicitud de ahorr: Modificacion de cuota------------------------------

CAMPOS_AHORRO_MOD = {
    "cedula":  "Cedula",
    "nombre":  "Nombre_Completo",
}
 
@login_required
def soli_ahorro_mod_cuota_buscar(request, campo):
    if campo not in CAMPOS_AHORRO_MOD:
        return JsonResponse({"results": []})
    col_db = CAMPOS_AHORRO_MOD[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudAhorroModCuota.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_lista(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "tipo_modificacion":request.GET.get("tipo_modificacion","").strip(),
        "numero_contrato":  request.GET.get("numero_contrato",  "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos             = ReporteSolicitudAhorroModCuota.obtener_datos(filtros)
    kpis              = ReporteSolicitudAhorroModCuota.obtener_kpis(filtros)
    tipos_ahorro      = ReporteSolicitudAhorroModCuota.obtener_tipos_ahorro()
    tipos_modificacion = ReporteSolicitudAhorroModCuota.obtener_tipos_modificacion()
 
    table = SolicitudAhorroModCuotaTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_ahorro_mod_cuota.html", {
        "table":              table,
        "filtros":            filtros,
        "kpis":               kpis,
        "tipos_ahorro":       tipos_ahorro,
        "tipos_modificacion": tipos_modificacion,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAhorroModCuota.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_ahorro_mod_cuota_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_export(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "tipo_modificacion":request.GET.get("tipo_modificacion","").strip(),
        "numero_contrato":  request.GET.get("numero_contrato",  "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudAhorroModCuota.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Modificación Cuota Ahorro"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre",
               "Tipo de Ahorro", "N° Contrato", "Modificación", "Monto (₡)"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        monto = row.get("MontoCuotaDeducir")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=5, value=row.get("TipoAhorro"))
        ws.cell(row=ri, column=6, value=row.get("NumeroContrato"))
        ws.cell(row=ri, column=7, value=row.get("TipoModificacion"))
        c_monto = ws.cell(row=ri, column=8, value=int(monto) if monto else 0)
        c_monto.number_format = '₡#,##0'
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 9):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 30, 25, 15, 14, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ahorro_mod_cuota_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_ahorro_mod_cuota")
def soli_ahorro_mod_cuota_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAhorroModCuota.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Mod. Cuota #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
 
    r = 1
 
    # Encabezado
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Modificación de Cuota de Ahorro")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos del accionista
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",  (solicitud.get("Cedula")         or "").strip()); r += 1
    data_row(r, "NOMBRE",  (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos de la modificación
    header_row(r, "DATOS DE LA MODIFICACIÓN", AMARILLO, NEGRO); r += 1
    data_row(r, "TIPO DE AHORRO",     solicitud.get("TipoAhorro")); r += 1
    data_row(r, "N° DE CONTRATO",     solicitud.get("NumeroContrato")); r += 1
    data_row(r, "TIPO MODIFICACIÓN",  solicitud.get("TipoModificacion")); r += 1
 
    # Monto con formato
    ws.merge_cells(f"A{r}:B{r}")
    lc = ws.cell(row=r, column=1, value="MONTO CUOTA A DEDUCIR")
    lc.fill = fill(GRIS_LABEL)
    lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2)
    lc.border = borde()
    ws.merge_cells(f"C{r}:D{r}")
    monto = solicitud.get("MontoCuotaDeducir")
    vc = ws.cell(row=r, column=3, value=int(monto) if monto else 0)
    vc.font = Font(color=NEGRO, size=10, name="Arial", bold=True)
    vc.number_format = '₡#,##0'
    vc.alignment = Alignment(vertical="center", indent=2)
    vc.border = borde()
    ws.row_dimensions[r].height = 22; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="mod_cuota_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response





#---------------------------- Solicitud reinversion ahorro ---------------------------------
CAMPOS_REINVERSION = {
    "cedula":  "Cedula",
    "nombre":  "Nombre_Completo",
}
 
@login_required
def soli_reinversion_ahorro_buscar(request, campo):
    if campo not in CAMPOS_REINVERSION:
        return JsonResponse({"results": []})
    col_db = CAMPOS_REINVERSION[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudReinversionAhorro.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_lista(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "tipo_reinversion": request.GET.get("tipo_reinversion", "").strip(),
        "numero_contrato":  request.GET.get("numero_contrato",  "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos              = ReporteSolicitudReinversionAhorro.obtener_datos(filtros)
    kpis               = ReporteSolicitudReinversionAhorro.obtener_kpis(filtros)
    tipos_ahorro       = ReporteSolicitudReinversionAhorro.obtener_tipos_ahorro()
    tipos_reinversion  = ReporteSolicitudReinversionAhorro.obtener_tipos_reinversion()
 
    table = SolicitudReinversionAhorroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_reinversion_ahorro.html", {
        "table":             table,
        "filtros":           filtros,
        "kpis":              kpis,
        "tipos_ahorro":      tipos_ahorro,
        "tipos_reinversion": tipos_reinversion,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudReinversionAhorro.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_reinversion_ahorro_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_export(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "tipo_reinversion": request.GET.get("tipo_reinversion", "").strip(),
        "numero_contrato":  request.GET.get("numero_contrato",  "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudReinversionAhorro.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Reinversión Ahorro"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha", "Hora", "Cédula", "Nombre",
               "Tipo de Ahorro", "Tipo Reinversión", "N° Contrato", "Reinversión 2"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("Fecha")
        hora  = row.get("Hora")
        fstr  = str(fecha) if fecha else ""
        hstr  = str(hora)  if hora  else ""
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=hstr)
        ws.cell(row=ri, column=4, value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=6, value=row.get("SolicitoReinversion"))
        ws.cell(row=ri, column=7, value=row.get("TipoReinversion"))
        ws.cell(row=ri, column=8, value=row.get("NumeroContrato"))
        ws.cell(row=ri, column=9, value=row.get("TipoReinversion2"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 10):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 12, 10, 14, 30, 20, 30, 15, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="reinversion_ahorro_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_reinversion_ahorro")
def soli_reinversion_ahorro_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudReinversionAhorro.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reinversión #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
 
    r = 1
 
    # Encabezado
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Reinversión Ahorro Existente")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("Fecha")
    hora  = solicitud.get("Hora")
    fecha_hora_str = f"{fecha} {hora}" if fecha and hora else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos del accionista
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",      (solicitud.get("Cedula")          or "").strip()); r += 1
    data_row(r, "NOMBRE",      (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    data_row(r, "FECHA",       str(fecha) if fecha else "—"); r += 1
    data_row(r, "HORA",        str(hora)  if hora  else "—"); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos de la reinversión
    header_row(r, "DATOS DE LA REINVERSIÓN", AMARILLO, NEGRO); r += 1
    data_row(r, "TIPO DE AHORRO",      solicitud.get("SolicitoReinversion")); r += 1
    data_row(r, "N° DE CONTRATO",      solicitud.get("NumeroContrato")); r += 1
    data_row(r, "TIPO REINVERSIÓN",    solicitud.get("TipoReinversion")); r += 1
 
    # TipoReinversion2 solo si tiene valor
    if solicitud.get("TipoReinversion2"):
        data_row(r, "REINVERSIÓN CUOTA",  solicitud.get("TipoReinversion2")); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="reinversion_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response





# ------------------Solicitud Autorizacion ahorro nuevo-------------------------
CAMPOS_AUTORIZACION_AHORRO = {
    "cedula":  "Cedula",
    "nombre":  "Nombre_Completo",
}
 
@login_required
def soli_autorizacion_ahorro_nuevo_buscar(request, campo):
    if campo not in CAMPOS_AUTORIZACION_AHORRO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_AUTORIZACION_AHORRO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudAutorizacionAhorroNuevo.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_lista(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "forma_pago":       request.GET.get("forma_pago",       "").strip(),
        "tipo_reinversion": request.GET.get("tipo_reinversion", "").strip(),
        "retiro_intereses": request.GET.get("retiro_intereses", "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos             = ReporteSolicitudAutorizacionAhorroNuevo.obtener_datos(filtros)
    kpis              = ReporteSolicitudAutorizacionAhorroNuevo.obtener_kpis(filtros)
    tipos_ahorro      = ReporteSolicitudAutorizacionAhorroNuevo.obtener_tipos_ahorro()
    formas_pago       = ReporteSolicitudAutorizacionAhorroNuevo.obtener_formas_pago()
    tipos_reinversion = ReporteSolicitudAutorizacionAhorroNuevo.obtener_tipos_reinversion()
 
    table = SolicitudAutorizacionAhorroNuevoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_autorizacion_ahorro_nuevo.html", {
        "table":             table,
        "filtros":           filtros,
        "kpis":              kpis,
        "tipos_ahorro":      tipos_ahorro,
        "formas_pago":       formas_pago,
        "tipos_reinversion": tipos_reinversion,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAutorizacionAhorroNuevo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_autorizacion_ahorro_nuevo_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_export(request):
    filtros = {
        "cedula":           request.GET.get("cedula",           "").strip(),
        "nombre":           request.GET.get("nombre",           "").strip(),
        "tipo_ahorro":      request.GET.get("tipo_ahorro",      "").strip(),
        "forma_pago":       request.GET.get("forma_pago",       "").strip(),
        "tipo_reinversion": request.GET.get("tipo_reinversion", "").strip(),
        "retiro_intereses": request.GET.get("retiro_intereses", "").strip(),
        "fecha_inicio":     request.GET.get("fecha_inicio",     "").strip(),
        "fecha_fin":        request.GET.get("fecha_fin",        "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudAutorizacionAhorroNuevo.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Autorización Ahorro Nuevo"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha", "Cédula", "Nombre", "Tipo de Ahorro",
               "Forma de Pago", "Tipo Reinversión", "Retiro Intereses",
               "Monto Cuota (₡)", "Monto Débito Ahorro (₡)"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("Fecha")
        fstr  = str(fecha) if fecha else ""
        mc    = row.get("MontoCuota")
        mda   = row.get("MontoDebitarAhorro")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("Nombre_Completo") or "").strip())
        ws.cell(row=ri, column=5,  value=row.get("TipoAhorro"))
        ws.cell(row=ri, column=6,  value=row.get("FormaPago"))
        ws.cell(row=ri, column=7,  value=row.get("TipoReinversion"))
        ws.cell(row=ri, column=8,  value=row.get("RetiroIntereses"))
        c_mc  = ws.cell(row=ri, column=9,  value=int(mc)  if mc  else None)
        c_mda = ws.cell(row=ri, column=10, value=int(mda) if mda else None)
        if mc:  c_mc.number_format  = '₡#,##0'
        if mda: c_mda.number_format = '₡#,##0'
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 11):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 12, 14, 30, 22, 15, 30, 15, 16, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="autorizacion_ahorro_nuevo_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_autorizacion_ahorro_nuevo")
def soli_autorizacion_ahorro_nuevo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudAutorizacionAhorroNuevo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ahorro Nuevo #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor, formato_moneda=False):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3,
                     value=int(valor) if formato_moneda and valor else (valor if valor is not None else "—"))
        vc.font = Font(color=NEGRO, size=10, name="Arial",
                       bold=True if formato_moneda else False)
        if formato_moneda and valor:
            vc.number_format = '₡#,##0'
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
 
    r = 1
 
    # Encabezado
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Autorización Ahorro Nuevo")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos del accionista
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",  (solicitud.get("Cedula")          or "").strip()); r += 1
    data_row(r, "NOMBRE",  (solicitud.get("Nombre_Completo") or "").strip()); r += 1
    data_row(r, "FECHA",   str(solicitud.get("Fecha")) if solicitud.get("Fecha") else "—"); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos del ahorro
    header_row(r, "DATOS DEL AHORRO", AMARILLO, NEGRO); r += 1
    data_row(r, "TIPO DE AHORRO",     solicitud.get("TipoAhorro")); r += 1
    data_row(r, "FORMA DE PAGO",      solicitud.get("FormaPago")); r += 1
    data_row(r, "TIPO REINVERSIÓN",   solicitud.get("TipoReinversion")); r += 1
    data_row(r, "RETIRO DE INTERESES", solicitud.get("RetiroIntereses")); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Montos
    header_row(r, "MONTOS", AZUL); r += 1
    data_row(r, "MONTO CUOTA",          solicitud.get("MontoCuota"),         formato_moneda=True); r += 1
    data_row(r, "MONTO DÉBITO AHORRO",  solicitud.get("MontoDebitarAhorro"), formato_moneda=True); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("Nombre_Completo") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ahorro_nuevo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response




# ── FORMULARIOS ───────────VIVIENDA─────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_vivienda")
def formulario_vivienda(request):
    return render(request, "dashboard/formularios/vivienda.html", {
        "active_tab": "vivienda"
    })



# ------FORMULARIO SOLICITUD PRESTAMO COMPRA VEHICULO


CAMPOS_VEHICULO = {
    "cedula":  "Cedula",
    "nombre":  "NombreCompleto",
}
 
@login_required
def soli_compra_vehiculo_buscar(request, campo):
    if campo not in CAMPOS_VEHICULO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_VEHICULO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudCompraVehiculo.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_lista(request):
    filtros = {
        "cedula":               request.GET.get("cedula",               "").strip(),
        "nombre":               request.GET.get("nombre",               "").strip(),
        "tipo_vehiculo":        request.GET.get("tipo_vehiculo",        "").strip(),
        "gastos_formalizacion": request.GET.get("gastos_formalizacion", "").strip(),
        "provincia_domicilio":  request.GET.get("provincia_domicilio",  "").strip(),
        "fecha_inicio":         request.GET.get("fecha_inicio",         "").strip(),
        "fecha_fin":            request.GET.get("fecha_fin",            "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos          = ReporteSolicitudCompraVehiculo.obtener_datos(filtros)
    kpis           = ReporteSolicitudCompraVehiculo.obtener_kpis(filtros)
    tipos_vehiculo = ReporteSolicitudCompraVehiculo.obtener_tipos_vehiculo()
 
    table = SolicitudCompraVehiculoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_compra_vehiculo.html", {
        "table":          table,
        "filtros":        filtros,
        "kpis":           kpis,
        "tipos_vehiculo": tipos_vehiculo,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudCompraVehiculo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_compra_vehiculo_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_export(request):
    filtros = {
        "cedula":               request.GET.get("cedula",               "").strip(),
        "nombre":               request.GET.get("nombre",               "").strip(),
        "tipo_vehiculo":        request.GET.get("tipo_vehiculo",        "").strip(),
        "gastos_formalizacion": request.GET.get("gastos_formalizacion", "").strip(),
        "provincia_domicilio":  request.GET.get("provincia_domicilio",  "").strip(),
        "fecha_inicio":         request.GET.get("fecha_inicio",         "").strip(),
        "fecha_fin":            request.GET.get("fecha_fin",            "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudCompraVehiculo.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes Vehículo"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha/Hora", "Cédula", "Nombre", "Correo",
               "Fecha Nac.", "Estado Civil", "Tel. Celular", "Tel. Casa", "Tel. Trabajo",
               "Nacionalidad", "Dir. Domicilio", "Prov. Domicilio", "Cantón Dom.", "Distrito Dom.",
               "Dir. Trabajo", "Prov. Trabajo", "Cantón Trab.", "Distrito Trab.",
               "Monto Solicitado", "Plazo", "Gastos Formaliz.", "Fecha Entrega",
               "Tipo Vehículo", "Garantía"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        fecha_nac = row.get("FechaNacimiento")
        fn_str = str(fecha_nac) if fecha_nac else ""
        monto = row.get("MontoCreditoSolicitado")
 
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula")         or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("NombreCompleto") or "").strip())
        ws.cell(row=ri, column=5,  value=row.get("Correo"))
        ws.cell(row=ri, column=6,  value=fn_str)
        ws.cell(row=ri, column=7,  value=row.get("EstadoCivil"))
        ws.cell(row=ri, column=8,  value=row.get("TelefonoCelular"))
        ws.cell(row=ri, column=9,  value=row.get("TelefonoCasa"))
        ws.cell(row=ri, column=10, value=row.get("TelefonoTrabajo"))
        ws.cell(row=ri, column=11, value=row.get("Nacionalidad"))
        ws.cell(row=ri, column=12, value=row.get("DireccionDomicilio"))
        ws.cell(row=ri, column=13, value=row.get("ProvinciaDomicilio"))
        ws.cell(row=ri, column=14, value=row.get("CantonDomicilio"))
        ws.cell(row=ri, column=15, value=row.get("DistritoDomicilio"))
        ws.cell(row=ri, column=16, value=row.get("DireccionTrabajo"))
        ws.cell(row=ri, column=17, value=row.get("ProvinciaTrabajo"))
        ws.cell(row=ri, column=18, value=row.get("CantonTrabajo"))
        ws.cell(row=ri, column=19, value=row.get("DistritoTrabajo"))
        c_m = ws.cell(row=ri, column=20, value=int(float(monto)) if monto else None)
        if monto: c_m.number_format = '₡#,##0'
        ws.cell(row=ri, column=21, value=row.get("Plazo"))
        ws.cell(row=ri, column=22, value=row.get("GastosFormalizacion"))
        ws.cell(row=ri, column=23, value=str(row.get("FechaEntrega")) if row.get("FechaEntrega") else "")
        ws.cell(row=ri, column=24, value=row.get("TipoVehiculo"))
        ws.cell(row=ri, column=25, value=row.get("Garantia"))
 
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 26):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    col_widths = [8,18,14,30,28,12,14,14,14,14,14,30,16,16,16,30,16,16,16,16,8,12,12,12,30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="solicitudes_vehiculo_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_compra_vehiculo")
def soli_compra_vehiculo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudCompraVehiculo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Vehículo #{respuesta_id}"
 
    AZUL, AMARILLO, VERDE, BLANCO = "003FB7", "FFC900", "166534", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor, col_label_end="C", col_val_start="D", col_val_end="F"):
        ws.merge_cells(f"A{row}:{col_label_end}{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"{col_val_start}{row}:{col_val_end}{row}")
        vc = ws.cell(row=row, column=4, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    # Anchos
    for col, w in enumerate([5, 20, 15, 15, 20, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
 
    r = 1
 
    # Encabezado
    ws.merge_cells(f"A{r}:F{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Préstamo Compra de Vehículo")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:F{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    # Datos personales
    header_row(r, "DATOS PERSONALES", AZUL); r += 1
    data_row(r, "CÉDULA",         (solicitud.get("Cedula")         or "").strip()); r += 1
    data_row(r, "NOMBRE",         (solicitud.get("NombreCompleto") or "").strip()); r += 1
    data_row(r, "CORREO",         solicitud.get("Correo")); r += 1
    data_row(r, "FECHA NAC.",     str(solicitud.get("FechaNacimiento")) if solicitud.get("FechaNacimiento") else "—"); r += 1
    data_row(r, "ESTADO CIVIL",   solicitud.get("EstadoCivil")); r += 1
    data_row(r, "NACIONALIDAD",   solicitud.get("Nacionalidad")); r += 1
    data_row(r, "TEL. CELULAR",   solicitud.get("TelefonoCelular")); r += 1
    data_row(r, "TEL. CASA",      solicitud.get("TelefonoCasa")); r += 1
    data_row(r, "TEL. TRABAJO",   solicitud.get("TelefonoTrabajo")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    # Domicilio
    header_row(r, "DOMICILIO", AMARILLO, NEGRO); r += 1
    data_row(r, "DIRECCIÓN",  solicitud.get("DireccionDomicilio")); r += 1
    data_row(r, "PROVINCIA",  solicitud.get("ProvinciaDomicilio")); r += 1
    data_row(r, "CANTÓN",     solicitud.get("CantonDomicilio")); r += 1
    data_row(r, "DISTRITO",   solicitud.get("DistritoDomicilio")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    # Trabajo
    header_row(r, "LUGAR DE TRABAJO", AZUL); r += 1
    data_row(r, "DIRECCIÓN",  solicitud.get("DireccionTrabajo")); r += 1
    data_row(r, "PROVINCIA",  solicitud.get("ProvinciaTrabajo")); r += 1
    data_row(r, "CANTÓN",     solicitud.get("CantonTrabajo")); r += 1
    data_row(r, "DISTRITO",   solicitud.get("DistritoTrabajo")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    # Préstamo
    header_row(r, "DATOS DEL PRÉSTAMO", AMARILLO, NEGRO); r += 1
    monto = solicitud.get("MontoCreditoSolicitado")
    ws.merge_cells(f"A{r}:C{r}")
    lc = ws.cell(row=r, column=1, value="MONTO SOLICITADO")
    lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
    ws.merge_cells(f"D{r}:F{r}")
    vc = ws.cell(row=r, column=4, value=int(float(monto)) if monto else "—")
    if monto: vc.number_format = '₡#,##0'
    vc.font = Font(color=NEGRO, size=10, name="Arial", bold=True)
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
    ws.row_dimensions[r].height = 22; r += 1
    data_row(r, "PLAZO",              solicitud.get("Plazo")); r += 1
    data_row(r, "GASTOS FORMALIZ.",   solicitud.get("GastosFormalizacion")); r += 1
    data_row(r, "FECHA ENTREGA",      str(solicitud.get("FechaEntrega")) if solicitud.get("FechaEntrega") else "—"); r += 1
    data_row(r, "TIPO DE VEHÍCULO",   solicitud.get("TipoVehiculo")); r += 1
    data_row(r, "GARANTÍA",           solicitud.get("Garantia")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    # Pie
    ws.merge_cells(f"A{r}:F{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="vehiculo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response




#--------------FORMULARIO SOLICITUD PRESTAMO VIVIENDA----------------

CAMPOS_PRESTAMO_VIVIENDA = {
    "cedula":   "Cedula",
    "nombre":   "NombreCompleto",
    "telefono": "Telefono",
}
 
@login_required
def soli_prestamo_vivienda_buscar(request, campo):
    if campo not in CAMPOS_PRESTAMO_VIVIENDA:
        return JsonResponse({"results": []})
    col_db = CAMPOS_PRESTAMO_VIVIENDA[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudPrestamoVivienda.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_lista(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "telefono":      request.GET.get("telefono",      "").strip(),
        "tipo_prestamo": request.GET.get("tipo_prestamo", "").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos           = ReporteSolicitudPrestamoVivienda.obtener_datos(filtros)
    kpis            = ReporteSolicitudPrestamoVivienda.obtener_kpis(filtros)
    tipos_prestamo  = ReporteSolicitudPrestamoVivienda.obtener_tipos_prestamo()
 
    table = SolicitudPrestamoViviendaTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_prestamo_vivienda.html", {
        "table":          table,
        "filtros":        filtros,
        "kpis":           kpis,
        "tipos_prestamo": tipos_prestamo,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPrestamoVivienda.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_prestamo_vivienda_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_export(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "telefono":      request.GET.get("telefono",      "").strip(),
        "tipo_prestamo": request.GET.get("tipo_prestamo", "").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudPrestamoVivienda.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos Vivienda"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre", "Teléfono", "Tipo de Préstamo"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")         or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("NombreCompleto") or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Telefono")       or "").strip())
        ws.cell(row=ri, column=6, value=row.get("TipoPrestamo"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 7):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 35, 14, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="prestamo_vivienda_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_vivienda")
def soli_prestamo_vivienda_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPrestamoVivienda.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Vivienda #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
 
    r = 1
 
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Préstamo para Vivienda")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula")         or "").strip()); r += 1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r += 1
    data_row(r, "TELÉFONO",  (solicitud.get("Telefono")       or "").strip()); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL PRÉSTAMO", AMARILLO, NEGRO); r += 1
    data_row(r, "TIPO DE PRÉSTAMO", solicitud.get("TipoPrestamo")); r += 1
 
    ws.row_dimensions[r].height = 8; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="vivienda_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response






#-------------Formulario Desarrollo Economico----------------

CAMPOS_DESARROLLO = {
    "cedula":   "Cedula",
    "nombre":   "NombreCompleto",
    "telefono": "Telefono",
}
 
@login_required
def soli_prestamo_desarrollo_buscar(request, campo):
    if campo not in CAMPOS_DESARROLLO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_DESARROLLO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudPrestamoDesarrollo.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_lista(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "telefono":      request.GET.get("telefono",      "").strip(),
        "plan_inversion":request.GET.get("plan_inversion","").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos  = ReporteSolicitudPrestamoDesarrollo.obtener_datos(filtros)
    kpis   = ReporteSolicitudPrestamoDesarrollo.obtener_kpis(filtros)
    planes = ReporteSolicitudPrestamoDesarrollo.obtener_planes()
 
    table = SolicitudPrestamoDesarrolloTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_prestamo_desarrollo.html", {
        "table":   table,
        "filtros": filtros,
        "kpis":    kpis,
        "planes":  planes,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPrestamoDesarrollo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_prestamo_desarrollo_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_export(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "telefono":      request.GET.get("telefono",      "").strip(),
        "plan_inversion":request.GET.get("plan_inversion","").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudPrestamoDesarrollo.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Desarrollo Económico"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre", "Teléfono", "Plan de Inversión"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")         or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("NombreCompleto") or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("Telefono")       or "").strip())
        ws.cell(row=ri, column=6, value=row.get("PlanInversion"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 7):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 35, 14, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="prestamo_desarrollo_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_prestamo_desarrollo")
def soli_prestamo_desarrollo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPrestamoDesarrollo.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Desarrollo #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo)
        c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL)
        lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2)
        lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2)
        vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
 
    r = 1
 
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Préstamo Desarrollo Económico")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula")         or "").strip()); r += 1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r += 1
    data_row(r, "TELÉFONO",  (solicitud.get("Telefono")       or "").strip()); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL PRÉSTAMO", AMARILLO, NEGRO); r += 1
    data_row(r, "PLAN DE INVERSIÓN", solicitud.get("PlanInversion")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL)
    pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="desarrollo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response








# ── FORMULARIOS ────────────PRESTAMOS────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_prestamos")
def formulario_prestamos(request):
    return render(request, "dashboard/formularios/prestamos.html", {
        "active_tab": "prestamos"
    })

#--------------------Pre solicitud de crédito personal -------------------------------


CAMPOS_CREDITO_PERSONAL = {
    "cedula":  "Cedula",
    "nombre":  "NombreCompleto",
}
 
@login_required
def soli_presolicitud_credito_personal_buscar(request, campo):
    if campo not in CAMPOS_CREDITO_PERSONAL:
        return JsonResponse({"results": []})
    col_db = CAMPOS_CREDITO_PERSONAL[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudPresolicitudCreditoPersonal.buscar_opciones(col_db, term)
    data   = [
        {"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
        for r in filas if r.get(col_db) and r[col_db].strip()
    ]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "tipo_credito": request.GET.get("tipo_credito", "").strip(),
        "sucursal":     request.GET.get("sucursal",     "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos          = ReporteSolicitudPresolicitudCreditoPersonal.obtener_datos(filtros)
    kpis           = ReporteSolicitudPresolicitudCreditoPersonal.obtener_kpis(filtros)
    tipos_credito  = ReporteSolicitudPresolicitudCreditoPersonal.obtener_tipos_credito()
    sucursales     = ReporteSolicitudPresolicitudCreditoPersonal.obtener_sucursales()
 
    table = SolicitudPresolicitudCreditoPersonalTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_presolicitud_credito_personal.html", {
        "table":         table,
        "filtros":       filtros,
        "kpis":          kpis,
        "tipos_credito": tipos_credito,
        "sucursales":    sucursales,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPresolicitudCreditoPersonal.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_presolicitud_credito_personal_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "tipo_credito": request.GET.get("tipo_credito", "").strip(),
        "sucursal":     request.GET.get("sucursal",     "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudPresolicitudCreditoPersonal.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Solicitud Crédito"
 
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre", "Teléfono",
               "Tipo de Crédito", "Sucursal", "Monto (₡)",
               "URL Cédula Frente", "URL Cédula Reverso", "URL Desglose Pensión"]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        monto = row.get("Monto")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula")         or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("NombreCompleto") or "").strip())
        ws.cell(row=ri, column=5,  value=(row.get("Telefono")       or "").strip())
        ws.cell(row=ri, column=6,  value=row.get("TipoCredito"))
        ws.cell(row=ri, column=7,  value=row.get("SucursalFormalizacion"))
        c_m = ws.cell(row=ri, column=8, value=int(float(monto)) if monto else None)
        if monto: c_m.number_format = '₡#,##0'
        ws.cell(row=ri, column=9,  value=row.get("FrenteCedula_URL"))
        ws.cell(row=ri, column=10, value=row.get("ReversoCedula_URL"))
        ws.cell(row=ri, column=11, value=row.get("DesglosePension_URL"))
        if ri % 2 == 0:
            fila_fill = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 12):
                ws.cell(row=ri, column=c).fill = fila_fill
 
    for i, w in enumerate([8, 18, 14, 35, 14, 42, 20, 14, 55, 55, 55], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="presolicitud_credito_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_presolicitud_credito")
def soli_presolicitud_credito_personal_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudPresolicitudCreditoPersonal.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
 
    wb = Workbook()
    ws = wb.active
    ws.title = f"Crédito #{respuesta_id}"
 
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
 
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
 
    def header_row(row, texto, color_fondo, color_texto=BLANCO):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.fill = fill(color_fondo); c.font = Font(bold=True, color=color_texto, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
 
    def data_row(row, label, valor):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    def url_row(row, label, url):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{row}:D{row}")
        vc = ws.cell(row=row, column=3, value=url or "No disponible")
        if url:
            vc.hyperlink = url
            vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
        else:
            vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[row].height = 22
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10
 
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Pre-Solicitud Crédito Personal")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1,
                value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r += 1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula")         or "").strip()); r += 1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r += 1
    data_row(r, "TELÉFONO",  (solicitud.get("Telefono")       or "").strip()); r += 1
    data_row(r, "FECHA / HORA", fecha_str); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DATOS DEL CRÉDITO", AMARILLO, NEGRO); r += 1
    data_row(r, "TIPO DE CRÉDITO", solicitud.get("TipoCredito")); r += 1
    data_row(r, "SUCURSAL",        solicitud.get("SucursalFormalizacion")); r += 1
 
    monto = solicitud.get("Monto")
    ws.merge_cells(f"A{r}:B{r}")
    lc = ws.cell(row=r, column=1, value="MONTO")
    lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
    lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
    ws.merge_cells(f"C{r}:D{r}")
    vc = ws.cell(row=r, column=3, value=int(float(monto)) if monto else "—")
    if monto: vc.number_format = '₡#,##0'
    vc.font = Font(color=NEGRO, size=10, name="Arial", bold=True)
    vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
    ws.row_dimensions[r].height = 22; r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    header_row(r, "DOCUMENTOS ADJUNTOS — SHAREFILE", AMARILLO, NEGRO); r += 1
    url_row(r, "CÉDULA — FRENTE",      solicitud.get("FrenteCedula_URL")); r += 1
    url_row(r, "CÉDULA — REVERSO",     solicitud.get("ReversoCedula_URL")); r += 1
    url_row(r, "DESGLOSE DE PENSIÓN",  solicitud.get("DesglosePension_URL")); r += 1
    ws.row_dimensions[r].height = 8; r += 1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1,
                  value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
 
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="credito_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response)
    return response









# ── FORMULARIOS ────────────CONTROL CREDITO────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_control_credito")
def formulario_control_credito(request):
    return render(request, "dashboard/formularios/control_credito.html", {
        "active_tab": "control_credito"
    })




# ----------------------Solicitud Formulario autorizacion de pagos -------------------
DOCS_COLS = ["Documento1","Documento2","Documento3","Documento4","Documento5",
             "Documento6","Documento7","Documento8","Documento9","Documento10"]
 
# ─── Comprobante Autorización Ahorro ──────────────────
CAMPOS_COMP_AHORRO = {"cedula": "Cedula", "nombre": "NombreCompleto"}
 
@login_required
def comprobante_autorizacion_ahorro_buscar(request, campo):
    if campo not in CAMPOS_COMP_AHORRO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_COMP_AHORRO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteComprobanteAutorizacionAhorro.buscar_opciones(col_db, term)
    data   = [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
              for r in filas if r.get(col_db) and r[col_db].strip()]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteComprobanteAutorizacionAhorro.obtener_datos(filtros)
    kpis  = ReporteComprobanteAutorizacionAhorro.obtener_kpis(filtros)
    table = ComprobanteAutorizacionAhorroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_comprobante_autorizacion_ahorro.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_detalle(request, respuesta_id):
    solicitud = ReporteComprobanteAutorizacionAhorro.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    documentos = [(f"Documento {i}", solicitud.get(f"Documento{i}"))
                  for i in range(1, 11) if solicitud.get(f"Documento{i}")]
    return render(request, "dashboard/formularios/reportes/soli_comprobante_autorizacion_ahorro_detalle.html", {
        "solicitud": solicitud, "documentos": documentos,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteComprobanteAutorizacionAhorro.obtener_datos(filtros)
    wb = Workbook(); ws = wb.active; ws.title = "Autorizacion Ahorro"
    AZUL = "003FB7"
    headers = ["ID","Fecha","Hora","Cédula","Nombre","Teléfono","Detalle Pago",
               "Doc1","Doc2","Doc3","Doc4","Doc5","Doc6","Doc7","Doc8","Doc9","Doc10"]
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=str(row.get("Fecha") or ""))
        ws.cell(row=ri, column=3,  value=str(row.get("Hora")  or ""))
        ws.cell(row=ri, column=4,  value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=5,  value=(row.get("NombreCompleto")  or "").strip())
        ws.cell(row=ri, column=6,  value=(row.get("NumeroTelefonico") or "").strip())
        ws.cell(row=ri, column=7,  value=row.get("DetallePago"))
        for i, doc in enumerate(DOCS_COLS, 8):
            ws.cell(row=ri, column=i, value=row.get(doc))
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 18): ws.cell(row=ri, column=c).fill = ff
    widths = [8,12,10,14,30,14,35]+[50]*10
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="autorizacion_ahorro_{fstr}.xlsx"'
    wb.save(response); return response
 
 
@login_required
@permiso_requerido("dashboard.view_comprobante_autorizacion_ahorro")
def comprobante_autorizacion_ahorro_export_detalle(request, respuesta_id):
    solicitud = ReporteComprobanteAutorizacionAhorro.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb = Workbook(); ws = wb.active; ws.title = f"Autorizacion #{respuesta_id}"
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
    def header_row(r, texto, cf, ct=BLANCO):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {texto}")
        c.fill = fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 28
    def data_row(r, label, valor):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
    def url_row(r, label, url):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=url or "No disponible")
        if url: vc.hyperlink = url; vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
        else: vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
    for col, w in enumerate([5,25,50,10], 1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = w
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Comprobante Autorización Ahorro Voluntario")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 36; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    s = ws.cell(row=r, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18; r+=1
    ws.row_dimensions[r].height = 8; r+=1
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r+=1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula") or "").strip()); r+=1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r+=1
    data_row(r, "TELÉFONO",  (solicitud.get("NumeroTelefonico") or "").strip()); r+=1
    data_row(r, "FECHA",     str(solicitud.get("Fecha") or "")); r+=1
    data_row(r, "HORA",      str(solicitud.get("Hora") or "")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    header_row(r, "DETALLE DE PAGO", AMARILLO, NEGRO); r+=1
    data_row(r, "DETALLE", solicitud.get("DetallePago")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    docs = [(f"DOCUMENTO {i}", solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    if docs:
        header_row(r, "DOCUMENTOS ADJUNTOS — SHAREFILE", AMARILLO, NEGRO); r+=1
        for label, url in docs:
            url_row(r, label, url); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="autorizacion_ahorro_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response
 




#-----------Solicitud Formulario Comprobante de pago ----------------------------------------------------
# ─── Comprobantes Pago ──────────────────────────────
CAMPOS_COMP_PAGO = {"cedula": "Cedula", "nombre": "NombreCompleto", "banco": "Banco"}
 
@login_required
def comprobantes_pago_buscar(request, campo):
    if campo not in CAMPOS_COMP_PAGO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_COMP_PAGO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteComprobantesPago.buscar_opciones(col_db, term)
    data   = [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
              for r in filas if r.get(col_db) and r[col_db].strip()]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "banco":        request.GET.get("banco",        "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteComprobantesPago.obtener_datos(filtros)
    kpis  = ReporteComprobantesPago.obtener_kpis(filtros)
    table = ComprobantesPagoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_comprobantes_pago.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_detalle(request, respuesta_id):
    solicitud = ReporteComprobantesPago.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    documentos = [(f"Documento {i}", solicitud.get(f"Documento{i}"))
                  for i in range(1, 11) if solicitud.get(f"Documento{i}")]
    return render(request, "dashboard/formularios/reportes/soli_comprobantes_pago_detalle.html", {
        "solicitud": solicitud, "documentos": documentos,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "banco":        request.GET.get("banco",        "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteComprobantesPago.obtener_datos(filtros)
    wb = Workbook(); ws = wb.active; ws.title = "Comprobantes Pago"
    AZUL = "003FB7"
    headers = ["ID","Fecha","Hora","Cédula","Nombre","Teléfono",
               "Banco","N° Depósito","Fecha Depósito","Monto (₡)","Detalle Pago",
               "Doc1","Doc2","Doc3","Doc4","Doc5","Doc6","Doc7","Doc8","Doc9","Doc10"]
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        monto = row.get("Monto")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=str(row.get("Fecha") or ""))
        ws.cell(row=ri, column=3,  value=str(row.get("Hora")  or ""))
        ws.cell(row=ri, column=4,  value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=5,  value=(row.get("NombreCompleto")  or "").strip())
        ws.cell(row=ri, column=6,  value=(row.get("NumeroTelefonico") or "").strip())
        ws.cell(row=ri, column=7,  value=row.get("Banco"))
        ws.cell(row=ri, column=8,  value=row.get("NumeroDeposito"))
        ws.cell(row=ri, column=9,  value=str(row.get("FechaDeposito") or ""))
        c_m = ws.cell(row=ri, column=10, value=int(float(monto)) if monto else None)
        if monto: c_m.number_format = '₡#,##0'
        ws.cell(row=ri, column=11, value=row.get("DetallePago"))
        for i, doc in enumerate(DOCS_COLS, 12):
            ws.cell(row=ri, column=i, value=row.get(doc))
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 22): ws.cell(row=ri, column=c).fill = ff
    widths = [8,12,10,14,30,14,20,16,14,14,35]+[50]*10
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="comprobantes_pago_{fstr}.xlsx"'
    wb.save(response); return response
 
 
@login_required
@permiso_requerido("dashboard.view_comprobantes_pago")
def comprobantes_pago_export_detalle(request, respuesta_id):
    solicitud = ReporteComprobantesPago.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
    wb = Workbook(); ws = wb.active; ws.title = f"Comprobante #{respuesta_id}"
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
    def header_row(r, texto, cf, ct=BLANCO):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {texto}")
        c.fill = fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 28
    def data_row(r, label, valor, moneda=False):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=int(float(valor)) if moneda and valor else (valor if valor is not None else "—"))
        if moneda and valor: vc.number_format = '₡#,##0'
        vc.font = Font(color=NEGRO, size=10, name="Arial", bold=moneda)
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
    def url_row(r, label, url):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=url or "No disponible")
        if url: vc.hyperlink = url; vc.font = Font(color="003FB7", underline="single", size=10, name="Arial")
        else: vc.font = Font(color="94A3B8", size=10, name="Arial", italic=True)
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
    for col, w in enumerate([5,25,50,10], 1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = w
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Comprobante de Pago / Depósito Bancario")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 36; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    s = ws.cell(row=r, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18; r+=1
    ws.row_dimensions[r].height = 8; r+=1
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r+=1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula") or "").strip()); r+=1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r+=1
    data_row(r, "TELÉFONO",  (solicitud.get("NumeroTelefonico") or "").strip()); r+=1
    data_row(r, "FECHA",     str(solicitud.get("Fecha") or "")); r+=1
    data_row(r, "HORA",      str(solicitud.get("Hora") or "")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    header_row(r, "DATOS DEL DEPÓSITO", AMARILLO, NEGRO); r+=1
    data_row(r, "BANCO",           solicitud.get("Banco")); r+=1
    data_row(r, "N° DEPÓSITO",     solicitud.get("NumeroDeposito")); r+=1
    data_row(r, "FECHA DEPÓSITO",  str(solicitud.get("FechaDeposito") or "")); r+=1
    data_row(r, "MONTO",           solicitud.get("Monto"), moneda=True); r+=1
    data_row(r, "DETALLE PAGO",    solicitud.get("DetallePago")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    docs = [(f"DOCUMENTO {i}", solicitud.get(f"Documento{i}")) for i in range(1,11) if solicitud.get(f"Documento{i}")]
    if docs:
        header_row(r, "DOCUMENTOS ADJUNTOS — SHAREFILE", AMARILLO, NEGRO); r+=1
        for label, url in docs:
            url_row(r, label, url); r+=1
    ws.row_dimensions[r].height = 8; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ","_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="comprobante_pago_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response




# ── FORMULARIOS ──────────────SERVICIO ACCIONISTA──────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_servicio_accionista")
def formulario_servicio_accionista(request):
    return render(request, "dashboard/formularios/servicio_accionista.html", {
        "active_tab": "servicio_accionista"
    })


#-------------Solicitud Formulario Caja Tel ----------------------------------
CAMPOS_CAJATEL = {"cedula": "Cedula", "nombre": "NombreCompleto"}
 
@login_required
def soli_clave_temporal_cajatel_buscar(request, campo):
    if campo not in CAMPOS_CAJATEL:
        return JsonResponse({"results": []})
    col_db = CAMPOS_CAJATEL[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudClaveTemporalCajaTel.buscar_opciones(col_db, term)
    data   = [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
              for r in filas if r.get(col_db) and r[col_db].strip()]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos = ReporteSolicitudClaveTemporalCajaTel.obtener_datos(filtros)
    kpis  = ReporteSolicitudClaveTemporalCajaTel.obtener_kpis(filtros)
    table = SolicitudClaveTemporalCajaTelTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_clave_temporal_cajatel.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudClaveTemporalCajaTel.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_clave_temporal_cajatel_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudClaveTemporalCajaTel.obtener_datos(filtros)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Clave CajaTel"
    AZUL    = "003FB7"
    headers = ["ID", "Fecha / Hora", "Cédula", "Nombre", "Correo Personal"]
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1, value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2, value=fstr)
        ws.cell(row=ri, column=3, value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=4, value=(row.get("NombreCompleto")  or "").strip())
        ws.cell(row=ri, column=5, value=(row.get("CorreoPersonal")  or "").strip())
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 6): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 18, 14, 35, 35], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="clave_cajatel_{fstr}.xlsx"'
    wb.save(response); return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_clave_cajatel")
def soli_clave_temporal_cajatel_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudClaveTemporalCajaTel.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
 
    wb = Workbook(); ws = wb.active; ws.title = f"CajaTel #{respuesta_id}"
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
    def header_row(r, texto, cf, ct=BLANCO):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {texto}")
        c.fill = fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 28
    def data_row(r, label, valor):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
 
    for col, w in enumerate([5, 25, 40, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
 
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Clave Temporal CajaTel")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 36; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    s = ws.cell(row=r, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18; r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r+=1
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    data_row(r, "CÉDULA",          (solicitud.get("Cedula")          or "").strip()); r+=1
    data_row(r, "NOMBRE",          (solicitud.get("NombreCompleto")  or "").strip()); r+=1
    data_row(r, "CORREO PERSONAL", (solicitud.get("CorreoPersonal")  or "").strip()); r+=1
    data_row(r, "FECHA / HORA",    fecha_str); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="cajatel_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response




# ── FORMULARIOS ─────────────SEGUROS───────────────────────────────
@login_required
@permiso_requerido("dashboard.view_formulario_seguros")
def formulario_seguros(request):
    return render(request, "dashboard/formularios/seguros.html", {
        "active_tab": "seguros"
    })


# -------------- Formulario solicitud Seguro Viajero -------------------------
import json

CAMPOS_SEGURO_VIAJERO = {
    "cedula":  "Cedula",
    "nombre":  "NombreCompleto",
    "destino": "Destino",
}

MESES_ES = {
    1:'Ene', 2:'Feb', 3:'Mar', 4:'Abr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Ago', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dic'
}
 

@login_required
def soli_seguro_viajero_buscar(request, campo):
    if campo not in CAMPOS_SEGURO_VIAJERO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_SEGURO_VIAJERO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudSeguroViajero.buscar_opciones(col_db, term)
    data   = [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
              for r in filas if r.get(col_db) and r[col_db].strip()]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_lista(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
 
    datos        = ReporteSolicitudSeguroViajero.obtener_datos(filtros)
    kpis         = ReporteSolicitudSeguroViajero.obtener_kpis(filtros)
    top_destinos = ReporteSolicitudSeguroViajero.obtener_top_destinos()
    raw_timeline = ReporteSolicitudSeguroViajero.obtener_timeline()
 
    # ── Timeline: lista de {label, total} ──
    timeline_data = [
        {
            "label": f"{MESES_ES[r['mes']]} {r['anio']}",
            "total": r["total"]
        }
        for r in raw_timeline
    ]
 
    # ── Heatmap: {anio: {mes: total}} ──
    heatmap_dict  = {}
    heatmap_anios = []
    for r in raw_timeline:
        a, m, t = r["anio"], r["mes"], r["total"]
        if a not in heatmap_dict:
            heatmap_dict[a] = {}
            heatmap_anios.append(a)
        heatmap_dict[a][m] = t
 
    table = SolicitudSeguroViajeroTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
 
    return render(request, "dashboard/formularios/reportes/soli_seguro_viajero.html", {
        "table":          table,
        "filtros":        filtros,
        "kpis":           kpis,
        "top_destinos":   top_destinos,
        "timeline_data":  timeline_data,
        "heatmap_anios":  heatmap_anios,
        "heatmap_json":   json.dumps(heatmap_dict),
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudSeguroViajero.obtener_detalle(respuesta_id)
    if not solicitud:
        raise Http404
    return render(request, "dashboard/formularios/reportes/soli_seguro_viajero_detalle.html", {
        "solicitud": solicitud,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_export(request):
    filtros = {
        "cedula":       request.GET.get("cedula",       "").strip(),
        "nombre":       request.GET.get("nombre",       "").strip(),
        "destino":      request.GET.get("destino",      "").strip(),
        "fecha_inicio": request.GET.get("fecha_inicio", "").strip(),
        "fecha_fin":    request.GET.get("fecha_fin",    "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudSeguroViajero.obtener_datos(filtros)
 
    wb = Workbook(); ws = wb.active; ws.title = "Seguro Viajero"
    AZUL    = "003FB7"
    headers = ["ID", "Fecha/Hora", "Cédula", "Nombre", "Correo", "Teléfono",
               "Destino", "Inicio Viaje", "Fin Viaje",
               "Nombre Beneficiario", "Cédula Beneficiario", "Parentesco"]
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula")          or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("NombreCompleto")  or "").strip())
        ws.cell(row=ri, column=5,  value=row.get("Correo"))
        ws.cell(row=ri, column=6,  value=row.get("Telefono"))
        ws.cell(row=ri, column=7,  value=row.get("Destino"))
        ws.cell(row=ri, column=8,  value=str(row.get("FechaInicioViaje") or ""))
        ws.cell(row=ri, column=9,  value=str(row.get("FechaFinalViaje")  or ""))
        ws.cell(row=ri, column=10, value=row.get("NombreBeneficiario"))
        ws.cell(row=ri, column=11, value=row.get("CedulaBeneficiario"))
        ws.cell(row=ri, column=12, value=row.get("Parentesco"))
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 13): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8, 18, 14, 30, 28, 12, 25, 12, 12, 30, 14, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="seguro_viajero_{fstr}.xlsx"'
    wb.save(response); return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_seguro_viajero")
def soli_seguro_viajero_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudSeguroViajero.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
 
    wb = Workbook(); ws = wb.active; ws.title = f"Viajero #{respuesta_id}"
    AZUL, AMARILLO, BLANCO = "003FB7", "FFC900", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
    def header_row(r, texto, cf, ct=BLANCO):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {texto}")
        c.fill = fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 28
    def data_row(r, label, valor):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
 
    for col, w in enumerate([5, 25, 35, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
 
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Seguro Viajero")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 36; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18; r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r+=1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula")          or "").strip()); r+=1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto")  or "").strip()); r+=1
    data_row(r, "CORREO",    solicitud.get("Correo")); r+=1
    data_row(r, "TELÉFONO",  solicitud.get("Telefono")); r+=1
    data_row(r, "FECHA / HORA", fecha_str); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL VIAJE", AMARILLO, NEGRO); r+=1
    data_row(r, "DESTINO",        solicitud.get("Destino")); r+=1
    data_row(r, "FECHA INICIO",   str(solicitud.get("FechaInicioViaje") or "—")); r+=1
    data_row(r, "FECHA FINAL",    str(solicitud.get("FechaFinalViaje")  or "—")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL BENEFICIARIO", AZUL); r+=1
    data_row(r, "NOMBRE",     solicitud.get("NombreBeneficiario")); r+=1
    data_row(r, "CÉDULA",     solicitud.get("CedulaBeneficiario")); r+=1
    data_row(r, "PARENTESCO", solicitud.get("Parentesco")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="viajero_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response


#-------------------Formulario solicitud Marchamo -----------------
CAMPOS_MARCHAMO = {
    "cedula":  "Cedula",
    "nombre":  "NombreCompleto",
    "placa":   "NumeroPlaca",
}
 
def _es_si(valor):  
    """Normaliza valores Sí/No independiente del formato en BD."""
    if valor is None: return False
    return str(valor).strip().lower() in ('sí', 'si', '1', 'yes', 'true', 's')
 
 
@login_required
def soli_marchamo_buscar(request, campo):
    if campo not in CAMPOS_MARCHAMO:
        return JsonResponse({"results": []})
    col_db = CAMPOS_MARCHAMO[campo]
    term   = request.GET.get("term", "").strip()
    filas  = ReporteSolicitudMarchamo.buscar_opciones(col_db, term)
    data   = [{"id": (r[col_db] or "").strip(), "text": (r[col_db] or "").strip()}
              for r in filas if r.get(col_db) and r[col_db].strip()]
    return JsonResponse({"results": data, "pagination": {"more": False}})
 
 
@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_lista(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "placa":         request.GET.get("placa",         "").strip(),
        "tipo_vehiculo": request.GET.get("tipo_vehiculo", "").strip(),
        "sucursal":      request.GET.get("sucursal",      "").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros         = {k: v for k, v in filtros.items() if v}
    datos           = ReporteSolicitudMarchamo.obtener_datos(filtros)
    kpis            = ReporteSolicitudMarchamo.obtener_kpis(filtros)
    tipos_vehiculo  = ReporteSolicitudMarchamo.obtener_tipos_vehiculo()
    sucursales      = ReporteSolicitudMarchamo.obtener_sucursales()
    table = SolicitudMarchamoTable(datos)
    tables.RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "dashboard/formularios/reportes/soli_marchamo.html", {
        "table": table, "filtros": filtros, "kpis": kpis,
        "tipos_vehiculo": tipos_vehiculo, "sucursales": sucursales,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudMarchamo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
 
    # Construir lista de seguros contratados
    seguros = []
    mapa = [
        ("SeguroRC",          "Responsabilidad Civil (RC)"),
        ("SeguroVida",        "Seguro de Vida"),
        ("SeguroSalud",       "Seguro de Salud"),
        ("SeguroAsistencia",  "Seguro de Asistencia"),
        ("SeguroVidaPlus",    "Seguro Vida Plus"),
        ("SeguroComprensivo", "Seguro Comprensivo"),
    ]
    for campo, label in mapa:
        seguros.append({"label": label, "activo": _es_si(solicitud.get(campo))})
 
    return render(request, "dashboard/formularios/reportes/soli_marchamo_detalle.html", {
        "solicitud": solicitud, "seguros": seguros,
    })
 
 
@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_export(request):
    filtros = {
        "cedula":        request.GET.get("cedula",        "").strip(),
        "nombre":        request.GET.get("nombre",        "").strip(),
        "placa":         request.GET.get("placa",         "").strip(),
        "tipo_vehiculo": request.GET.get("tipo_vehiculo", "").strip(),
        "sucursal":      request.GET.get("sucursal",      "").strip(),
        "fecha_inicio":  request.GET.get("fecha_inicio",  "").strip(),
        "fecha_fin":     request.GET.get("fecha_fin",     "").strip(),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    datos   = ReporteSolicitudMarchamo.obtener_datos(filtros)
 
    wb = Workbook(); ws = wb.active; ws.title = "Marchamo"
    AZUL = "003FB7"
    headers = ["ID","Fecha/Hora","Cédula","Nombre","Teléfono",
               "Tipo Vehículo","Placa","Dueño Registral",
               "Pago Ahorro","Pago Tarjeta","Sucursal Retiro",
               "Seg. RC","Seg. Vida","Seg. Salud",
               "Seg. Asistencia","Seg. Vida Plus","Seg. Comprensivo"]
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(datos, 2):
        fecha = row.get("FechaHora")
        fstr  = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
        ws.cell(row=ri, column=1,  value=row.get("respuesta_id"))
        ws.cell(row=ri, column=2,  value=fstr)
        ws.cell(row=ri, column=3,  value=(row.get("Cedula")             or "").strip())
        ws.cell(row=ri, column=4,  value=(row.get("NombreCompleto")     or "").strip())
        ws.cell(row=ri, column=5,  value=row.get("telefono"))
        ws.cell(row=ri, column=6,  value=row.get("TipoVehiculo"))
        ws.cell(row=ri, column=7,  value=(row.get("NumeroPlaca")        or "").strip())
        ws.cell(row=ri, column=8,  value=row.get("NombreDuenoRegistral"))
        ws.cell(row=ri, column=9,  value=row.get("AutorizoPagoAhorro"))
        ws.cell(row=ri, column=10, value=row.get("AutorizoPagoTarjeta"))
        ws.cell(row=ri, column=11, value=row.get("SucursalRetiro"))
        ws.cell(row=ri, column=12, value=row.get("SeguroRC"))
        ws.cell(row=ri, column=13, value=row.get("SeguroVida"))
        ws.cell(row=ri, column=14, value=row.get("SeguroSalud"))
        ws.cell(row=ri, column=15, value=row.get("SeguroAsistencia"))
        ws.cell(row=ri, column=16, value=row.get("SeguroVidaPlus"))
        ws.cell(row=ri, column=17, value=row.get("SeguroComprensivo"))
        if ri % 2 == 0:
            ff = PatternFill("solid", fgColor="E8EFFE")
            for c in range(1, 18): ws.cell(row=ri, column=c).fill = ff
    for i, w in enumerate([8,18,14,30,12,14,10,25,12,12,22,10,10,10,14,13,15], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    fstr = datetime.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="marchamo_{fstr}.xlsx"'
    wb.save(response); return response
 
 
@login_required
@permiso_requerido("dashboard.view_soli_marchamo")
def soli_marchamo_export_detalle(request, respuesta_id):
    solicitud = ReporteSolicitudMarchamo.obtener_detalle(respuesta_id)
    if not solicitud: raise Http404
 
    wb = Workbook(); ws = wb.active; ws.title = f"Marchamo #{respuesta_id}"
    AZUL, AMARILLO, VERDE, BLANCO = "003FB7", "FFC900", "166534", "FFFFFF"
    GRIS_LABEL, GRIS_BORDE, NEGRO = "F1F5F9", "E2E8F0", "1E293B"
    def fill(h): return PatternFill("solid", fgColor=h)
    def borde(): return Border(bottom=Side(style="thin", color=GRIS_BORDE))
    def header_row(r, texto, cf, ct=BLANCO):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {texto}")
        c.fill = fill(cf); c.font = Font(bold=True, color=ct, size=11, name="Arial")
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 28
    def data_row(r, label, valor):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value=valor if valor is not None else "—")
        vc.font = Font(color=NEGRO, size=10, name="Arial")
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
    def seguro_row(r, label, activo):
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = fill(GRIS_LABEL); lc.font = Font(bold=True, color="64748B", size=9, name="Arial")
        lc.alignment = Alignment(vertical="center", indent=2); lc.border = borde()
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws.cell(row=r, column=3, value="✓ Sí" if activo else "✗ No")
        vc.font = Font(color="166534" if activo else "94A3B8", size=10, name="Arial", bold=activo)
        vc.alignment = Alignment(vertical="center", indent=2); vc.border = borde()
        ws.row_dimensions[r].height = 22
 
    for col, w in enumerate([5,25,35,15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
 
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    t = ws.cell(row=r, column=1, value="CAJA DE ANDE — Solicitud Pago Marchamo")
    t.fill = fill(AZUL); t.font = Font(bold=True, color=BLANCO, size=13, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 36; r+=1
    ws.merge_cells(f"A{r}:D{r}")
    fecha = solicitud.get("FechaHora")
    fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if hasattr(fecha, "strftime") else str(fecha or "")
    s = ws.cell(row=r, column=1, value=f"ID #{solicitud['respuesta_id']}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.fill = fill("002A80"); s.font = Font(color="93C5FD", size=9, name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18; r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL ACCIONISTA", AZUL); r+=1
    data_row(r, "CÉDULA",    (solicitud.get("Cedula")         or "").strip()); r+=1
    data_row(r, "NOMBRE",    (solicitud.get("NombreCompleto") or "").strip()); r+=1
    data_row(r, "TELÉFONO",  solicitud.get("telefono")); r+=1
    data_row(r, "FECHA / HORA", fecha_str); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "DATOS DEL VEHÍCULO", AMARILLO, NEGRO); r+=1
    data_row(r, "TIPO DE VEHÍCULO",   solicitud.get("TipoVehiculo")); r+=1
    data_row(r, "NÚMERO DE PLACA",    (solicitud.get("NumeroPlaca") or "").strip()); r+=1
    data_row(r, "DUEÑO REGISTRAL",    solicitud.get("NombreDuenoRegistral")); r+=1
    data_row(r, "PAGO CON AHORRO",    solicitud.get("AutorizoPagoAhorro")); r+=1
    data_row(r, "PAGO CON TARJETA",   solicitud.get("AutorizoPagoTarjeta")); r+=1
    data_row(r, "SUCURSAL RETIRO",    solicitud.get("SucursalRetiro")); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    header_row(r, "SEGUROS INCLUIDOS", AZUL); r+=1
    mapa = [
        ("SeguroRC",          "Responsabilidad Civil (RC)"),
        ("SeguroVida",        "Seguro de Vida"),
        ("SeguroSalud",       "Seguro de Salud"),
        ("SeguroAsistencia",  "Seguro de Asistencia"),
        ("SeguroVidaPlus",    "Seguro Vida Plus"),
        ("SeguroComprensivo", "Seguro Comprensivo"),
    ]
    for campo, label in mapa:
        seguro_row(r, label, _es_si(solicitud.get(campo))); r+=1
    ws.row_dimensions[r].height = 8; r+=1
 
    ws.merge_cells(f"A{r}:D{r}")
    pie = ws.cell(row=r, column=1, value="Documento generado automáticamente por HorizonZero — Caja de ANDE")
    pie.fill = fill(GRIS_LABEL); pie.font = Font(color="94A3B8", size=8, name="Arial", italic=True)
    pie.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"; ws.sheet_view.zoomScale = 110
 
    nombre = (solicitud.get("NombreCompleto") or "solicitud").strip().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="marchamo_{respuesta_id}_{nombre}.xlsx"'
    wb.save(response); return response
    




#========================================================
# APARTADO USUARIOS - Se encuentran las funciones CRUD
#========================================================

# ── LISTADO PRINCIPAL ───────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_agentes")
def agentes_home(request):
    buscar      = request.GET.get("q", "").strip()
    unidad_id   = request.GET.get("unidad_id") or None
    sucursal_id = request.GET.get("sucursal_id") or None

    agentes_lista = AgentesService.listar(buscar, unidad_id, sucursal_id)
    sucursales    = AgentesService.sucursales()
    unidades      = AgentesService.unidades()
    kpis          = AgentesService.kpis()
    conteo        = AgentesService.conteo_por_unidad()

    # Paginación manual — 50 por página como el resto del proyecto
    paginator = Paginator(agentes_lista, 20)
    page_num  = request.GET.get("page", 1)
    page_obj  = paginator.get_page(page_num)

    return render(request, "dashboard/usuarios/agentes_home.html", {
        "agentes":    page_obj,          # ← page_obj en lugar de lista directa
        "sucursales": sucursales,
        "unidades":   unidades,
        "kpis":       kpis,
        "conteo":     conteo,
        "buscar":     buscar,
        "filtro_unidad":   unidad_id,
        "filtro_sucursal": sucursal_id,
        "paginator":  paginator,
        "page_obj":   page_obj,
    })
 
 
# ── CREAR ───────────────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.add_agente")
def agente_crear(request):
    sucursales = AgentesService.sucursales()
    unidades   = AgentesService.unidades()

    if request.method == "POST":
        nombre      = request.POST.get("nombre", "").strip()
        sucursal_id = request.POST.get("sucursal_id")
        unidad_id   = request.POST.get("unidad_id")
        login_ad    = request.POST.get("login_ad", "").strip().lower()

        if not nombre or not sucursal_id or not unidad_id:
            messages.error(request, "Todos los campos son requeridos.")
        else:
            # Verificar si el login_ad ya existe
            if login_ad and login_ad != "pendiente":
                existente = AgentesService.buscar_por_login_ad(login_ad)
                if existente:
                    return render(request, "dashboard/usuarios/agente_form.html", {
                        "sucursales":  sucursales,
                        "unidades":    unidades,
                        "modo":        "crear",
                        "agente_existente": existente,
                        "login_ad_buscado": login_ad,
                    })

            try:
                nuevo_id = AgentesService.crear(nombre, int(sucursal_id), int(unidad_id), login_ad or "pendiente")
                audit(request, AuditLog.Accion.AGENTE_CREATE, AuditLog.Modulo.AGENTES,
                      f"Creó el agente '{nombre}' (login: {login_ad or 'pendiente'})",
                      objeto_id=nuevo_id, objeto_nombre=nombre,
                      datos_nuevos={"nombre": nombre, "unidad_id": unidad_id,
                                    "sucursal_id": sucursal_id, "login_ad": login_ad})
                messages.success(request, f"Agente '{nombre}' creado con ID #{nuevo_id}.")
                return redirect("dashboard:agentes_home")
            except Exception as e:
                messages.error(request, f"Error al crear: {e}")

    return render(request, "dashboard/usuarios/agente_form.html", {
        "sucursales": sucursales,
        "unidades":   unidades,
        "modo":       "crear",
    })

 
 
# ── EDITAR ──────────────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.change_agente")
def agente_editar(request, agente_id):
    agente     = AgentesService.obtener(agente_id)
    if not agente:
        messages.error(request, "Agente no encontrado.")
        return redirect("dashboard:agentes_home")

    sucursales = AgentesService.sucursales()
    unidades   = AgentesService.unidades()

    if request.method == "POST":
        nombre      = request.POST.get("nombre", "").strip()
        sucursal_id = request.POST.get("sucursal_id")
        unidad_id   = request.POST.get("unidad_id")
        login_ad    = request.POST.get("login_ad", "").strip().lower()

        if not nombre or not sucursal_id or not unidad_id:
            messages.error(request, "Todos los campos son requeridos.")
        else:
            # Verificar duplicado de login_ad (excluyendo el agente actual)
            if login_ad and login_ad != "pendiente":
                existente = AgentesService.buscar_por_login_ad(login_ad)
                if existente and existente["agente_id"] != agente_id:
                    messages.error(request, f"El login '{login_ad}' ya está asignado al agente #{existente['agente_id']} — {existente['nombre']}.")
                    return render(request, "dashboard/usuarios/agente_form.html", {
                        "agente": agente, "sucursales": sucursales,
                        "unidades": unidades, "modo": "editar",
                    })

            try:
                AgentesService.actualizar(agente_id, nombre, int(sucursal_id), int(unidad_id), login_ad)
                audit(request, AuditLog.Accion.AGENTE_UPDATE, AuditLog.Modulo.AGENTES,
                      f"Editó el agente '{agente['nombre']}'",
                      objeto_id=agente_id, objeto_nombre=nombre,
                      datos_anteriores={"nombre": agente["nombre"], "login_ad": agente["login_ad"]},
                      datos_nuevos={"nombre": nombre, "login_ad": login_ad})
                messages.success(request, f"Agente #{agente_id} actualizado.")
                return redirect("dashboard:agentes_home")
            except Exception as e:
                messages.error(request, f"Error al actualizar: {e}")

    return render(request, "dashboard/usuarios/agente_form.html", {
        "agente":     agente,
        "sucursales": sucursales,
        "unidades":   unidades,
        "modo":       "editar",
    })
 
 
# ── ELIMINAR (POST) ─────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.delete_agente")
@require_POST
def agente_eliminar(request, agente_id):
    try:
        agente = AgentesService.obtener(agente_id)
        AgentesService.eliminar(agente_id)
        audit(request, AuditLog.Accion.AGENTE_DELETE, AuditLog.Modulo.AGENTES,
            f"Eliminó el agente '{agente['nombre']}' (soft delete)",
            objeto_id=agente_id, objeto_nombre=agente["nombre"],
            severidad=AuditLog.Severidad.WARNING)
        messages.success(request, f"Agente #{agente_id} eliminado.")
    except Exception as e:
        messages.error(request, f"Error al eliminar: {e}")
    return redirect("dashboard:agentes_home")
 
 
# ── VER QR DE UN AGENTE ─────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_agentes")
def agente_qr(request, agente_id):
    agente = AgentesService.obtener(agente_id)
    if not agente:
        messages.error(request, "Agente no encontrado.")
        return redirect("dashboard:agentes_home")
 
    qr_data = AgentesService.qr_data(agente_id, agente['unidad_nombre'])

    return render(request, "dashboard/usuarios/agente_qr.html", {
        "agente":  agente,
        "qr_data": qr_data,
    })
 
 
# ── DESCARGAR QR INDIVIDUAL (PNG) ───────────────────────────────
@login_required
@permiso_requerido("dashboard.view_agentes")
def agente_qr_download(request, agente_id, encuesta_id):
    agente = AgentesService.obtener(agente_id)        
    if not agente:
        return HttpResponse("Not found", status=404)

    unidad_key  = (agente['unidad_nombre'] or "").strip().lower()
    encuesta_id = UNIDAD_ENCUESTA.get(unidad_key, 1)

    url    = build_encuesta_url(int(encuesta_id), agente_id)
    qr_b64 = generar_qr_base64(url, size=12)
    png    = base64.b64decode(qr_b64)

    enc_nombre     = ENCUESTA_NOMBRE.get(int(encuesta_id), f"encuesta_{encuesta_id}")
    nombre_archivo = f"QR_{agente['nombre'].replace(' ', '_')}_{enc_nombre.replace(' ', '_')}.png"

    audit(request, AuditLog.Accion.EXPORT_QR, AuditLog.Modulo.QR,  
          f"Descargó QR del agente '{agente['nombre']}' — {enc_nombre}",
          objeto_id=agente_id, objeto_nombre=agente["nombre"])

    response = HttpResponse(png, content_type="image/png")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response

# ── DESCARGAR QR INDIVIDUAL (PNG ZIP) ───────────────────────────────
@login_required
@permiso_requerido("dashboard.view_agentes")
def agente_qr_download_zip(request, agente_id):
    agente = AgentesService.obtener(agente_id)          
    if not agente:
        return HttpResponse("Not found", status=404)

    qr_data = AgentesService.qr_data(agente_id, agente['unidad_nombre'])
    buffer  = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in qr_data:
            png = base64.b64decode(item["qr_b64"])
            fname = f"QR_{item['nombre'].replace(' ', '_')}.png"
            zf.writestr(fname, png)

    buffer.seek(0)
    nombre_zip = f"QR_{agente['nombre'].replace(' ', '_')}.zip"

    audit(request, AuditLog.Accion.EXPORT_ZIP, AuditLog.Modulo.QR,  # ← después
          f"Descargó ZIP de QR del agente '{agente['nombre']}'",
          objeto_id=agente_id, objeto_nombre=agente["nombre"])

    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{nombre_zip}"'
    return response


@login_required
@permiso_requerido("dashboard.view_agentes")
def agentes_inactivos(request):
    buscar      = request.GET.get("q", "").strip()
    unidad_id   = request.GET.get("unidad_id") or None
    sucursal_id = request.GET.get("sucursal_id") or None

    agentes_lista = AgentesService.listar_inactivos(buscar, unidad_id, sucursal_id)
    sucursales    = AgentesService.sucursales()
    unidades      = AgentesService.unidades()

    paginator = Paginator(agentes_lista, 50)
    page_obj  = paginator.get_page(request.GET.get("page", 1))

    return render(request, "dashboard/usuarios/agentes_inactivos.html", {
        "agentes":         page_obj,
        "sucursales":      sucursales,
        "unidades":        unidades,
        "buscar":          buscar,
        "filtro_unidad":   unidad_id,
        "filtro_sucursal": sucursal_id,
        "paginator":       paginator,
        "page_obj":        page_obj,
    })


@login_required
@permiso_requerido("dashboard.delete_agente")
@require_POST
def agente_restaurar(request, agente_id):
    try:
        AgentesService.restaurar(agente_id)
        audit(request, AuditLog.Accion.AGENTE_RESTORE, AuditLog.Modulo.AGENTES,
            f"Restauró el agente #{agente_id}",
            objeto_id=agente_id,
            severidad=AuditLog.Severidad.WARNING)
        messages.success(request, f"Agente #{agente_id} restaurado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al restaurar: {e}")
    return redirect("dashboard:agentes_inactivos")




#============================================
#============================================
# MODULO DE LOGS    -   AUDITORIA DE USUARIOS
#============================================
#============================================

# ── LISTADO PRINCIPAL DE LOGS ────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_logs")
def logs_home(request):
    qs = AuditLog.objects.all()
 
    # ── Filtros ──────────────────────────────────────────────────
    filtro_usuario   = request.GET.get("usuario", "").strip()
    filtro_accion    = request.GET.get("accion", "").strip()
    filtro_modulo    = request.GET.get("modulo", "").strip()
    filtro_severidad = request.GET.get("severidad", "").strip()
    filtro_ip        = request.GET.get("ip", "").strip()
    filtro_desde     = request.GET.get("desde", "").strip()
    filtro_hasta     = request.GET.get("hasta", "").strip()
 
    if filtro_usuario:
        qs = qs.filter(username__icontains=filtro_usuario)
    if filtro_accion:
        qs = qs.filter(accion=filtro_accion)
    if filtro_modulo:
        qs = qs.filter(modulo=filtro_modulo)
    if filtro_severidad:
        qs = qs.filter(severidad=filtro_severidad)
    if filtro_ip:
        qs = qs.filter(ip_address__icontains=filtro_ip)
    if filtro_desde:
        qs = qs.filter(fecha__date__gte=filtro_desde)
    if filtro_hasta:
        qs = qs.filter(fecha__date__lte=filtro_hasta)
 
    # ── KPIs (sobre el queryset filtrado) ───────────────────────
    total_eventos     = qs.count()
    logins_fallidos   = qs.filter(accion=AuditLog.Accion.LOGIN_FAIL).count()
    acciones_criticas = qs.filter(severidad=AuditLog.Severidad.CRITICAL).count()
    usuarios_unicos   = qs.values("username").distinct().count()
 
    # ── Gráfico: actividad por día (últimos 30 días del filtro) ──
    actividad = (
        qs.annotate(dia=TruncDate("fecha"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")[:30]
    )
    grafico_labels = [str(r["dia"]) for r in actividad]
    grafico_data   = [r["total"] for r in actividad]
 
    # ── Gráfico: distribución por acción ────────────────────────
    dist_accion = (
        qs.values("accion")
        .annotate(total=Count("id"))
        .order_by("-total")[:8]
    )
    dist_labels = [r["accion"] for r in dist_accion]
    dist_data   = [r["total"] for r in dist_accion]
 
    # ── Paginación ───────────────────────────────────────────────
    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get("page", 1))
 
    return render(request, "dashboard/logs/logs_home.html", {
        "logs":               page_obj,
        "paginator":          paginator,
        "page_obj":           page_obj,
        # Filtros activos
        "filtro_usuario":     filtro_usuario,
        "filtro_accion":      filtro_accion,
        "filtro_modulo":      filtro_modulo,
        "filtro_severidad":   filtro_severidad,
        "filtro_ip":          filtro_ip,
        "filtro_desde":       filtro_desde,
        "filtro_hasta":       filtro_hasta,
        # KPIs
        "total_eventos":      total_eventos,
        "logins_fallidos":    logins_fallidos,
        "acciones_criticas":  acciones_criticas,
        "usuarios_unicos":    usuarios_unicos,
        # Gráficos
        "grafico_labels":     json.dumps(grafico_labels),
        "grafico_data":       json.dumps(grafico_data),
        "dist_labels":        json.dumps(dist_labels),
        "dist_data":          json.dumps(dist_data),
        # Opciones para selects
        "acciones":           AuditLog.Accion.choices,
        "modulos":            AuditLog.Modulo.choices,
        "severidades":        AuditLog.Severidad.choices,
    })
 
 
# ── DETALLE DE UN LOG ────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_logs")
def log_detalle(request, log_id):
    try:
        log = AuditLog.objects.get(pk=log_id)
    except AuditLog.DoesNotExist:
        messages.error(request, "Registro no encontrado.")
        return redirect("dashboard:logs_home")
 
    # Parsear JSON para mostrar diff
    datos_anteriores = None
    datos_nuevos     = None
    if log.datos_anteriores:
        try:
            datos_anteriores = json.loads(log.datos_anteriores)
        except Exception:
            datos_anteriores = log.datos_anteriores
    if log.datos_nuevos:
        try:
            datos_nuevos = json.loads(log.datos_nuevos)
        except Exception:
            datos_nuevos = log.datos_nuevos
 
    return render(request, "dashboard/logs/log_detalle.html", {
        "log":               log,
        "datos_anteriores":  datos_anteriores,
        "datos_nuevos":      datos_nuevos,
    })
 
 
# ── EXPORTAR EXCEL ───────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_logs")
def logs_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime as dt
 
    qs = AuditLog.objects.all()
 
    # Aplicar mismos filtros
    if request.GET.get("usuario"):
        qs = qs.filter(username__icontains=request.GET["usuario"])
    if request.GET.get("accion"):
        qs = qs.filter(accion=request.GET["accion"])
    if request.GET.get("modulo"):
        qs = qs.filter(modulo=request.GET["modulo"])
    if request.GET.get("severidad"):
        qs = qs.filter(severidad=request.GET["severidad"])
    if request.GET.get("desde"):
        qs = qs.filter(fecha__date__gte=request.GET["desde"])
    if request.GET.get("hasta"):
        qs = qs.filter(fecha__date__lte=request.GET["hasta"])
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoría"
 
    AZUL     = "003FB7"
    AMARILLO = "FFC900"
    ROJO     = "DC2626"
    GRIS     = "F1F5F9"
 
    headers = [
        "ID", "Fecha / Hora", "Usuario", "Acción", "Módulo",
        "Severidad", "Descripción", "Objeto ID", "Objeto Nombre",
        "IP Address", "URL", "Método HTTP"
    ]
 
    hfill = PatternFill("solid", fgColor=AZUL)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
 
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    ws.row_dimensions[1].height = 22
 
    for ri, log in enumerate(qs, 2):
        fecha_str = log.fecha.strftime("%d/%m/%Y %H:%M:%S") if log.fecha else ""
 
        # Color de fila según severidad
        if log.severidad == "CRITICAL":
            row_fill = PatternFill("solid", fgColor="FEE2E2")
        elif log.severidad == "WARNING":
            row_fill = PatternFill("solid", fgColor="FEF9C3")
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF") if ri % 2 == 0 else PatternFill("solid", fgColor=GRIS)
 
        valores = [
            str(log.id),
            fecha_str,
            log.username,
            log.get_accion_display(),
            log.get_modulo_display(),
            log.get_severidad_display(),
            log.descripcion,
            log.objeto_id or "",
            log.objeto_nombre or "",
            str(log.ip_address) if log.ip_address else "",
            log.url or "",
            log.metodo_http or "",
        ]
 
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = row_fill
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center")
 
        ws.row_dimensions[ri].height = 18
 
    # Anchos de columna
    anchos = [8, 18, 20, 22, 16, 12, 60, 10, 30, 16, 50, 10]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    # Freeze header
    ws.freeze_panes = "A2"
 
    # Hoja resumen para la USI
    ws2 = wb.create_sheet("Resumen USI")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
 
    ws2.cell(row=1, column=1, value="REPORTE DE AUDITORÍA — HorizonZero").font = Font(bold=True, color=AZUL, size=13)
    ws2.cell(row=2, column=1, value=f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')}").font = Font(color="64748B", size=9)
    ws2.cell(row=2, column=2, value=f"por {request.user.get_full_name() or request.user.username}").font = Font(color="64748B", size=9)
 
    resumen = [
        ("Total de eventos", qs.count()),
        ("Logins exitosos",  qs.filter(accion="LOGIN_OK").count()),
        ("Logins fallidos",  qs.filter(accion="LOGIN_FAIL").count()),
        ("Eventos críticos", qs.filter(severidad="CRITICAL").count()),
        ("Eventos advertencia", qs.filter(severidad="WARNING").count()),
        ("Usuarios únicos",  qs.values("username").distinct().count()),
        ("Acciones sobre agentes", qs.filter(modulo="AGENTES").count()),
        ("Exportaciones",    qs.filter(accion__in=["EXPORT_EXCEL","EXPORT_QR","EXPORT_ZIP"]).count()),
    ]
 
    for i, (label, valor) in enumerate(resumen, 4):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        c = ws2.cell(row=i, column=2, value=valor)
        c.font = Font(bold=True, color=AZUL, size=10)
        c.alignment = Alignment(horizontal="center")
 
    fstr = dt.now().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="audit_log_{fstr}.xlsx"'
    wb.save(response)
    return response
 
 
# ── EXPORTAR PDF ─────────────────────────────────────────────────
@login_required
@permiso_requerido("dashboard.view_logs")
def logs_export_pdf(request):
    """
    Genera un PDF del reporte de auditoría usando ReportLab.
    Instalar: pip install reportlab
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from datetime import datetime as dt
        import io
 
        qs = AuditLog.objects.all()
        if request.GET.get("usuario"):
            qs = qs.filter(username__icontains=request.GET["usuario"])
        if request.GET.get("accion"):
            qs = qs.filter(accion=request.GET["accion"])
        if request.GET.get("modulo"):
            qs = qs.filter(modulo=request.GET["modulo"])
        if request.GET.get("severidad"):
            qs = qs.filter(severidad=request.GET["severidad"])
        if request.GET.get("desde"):
            qs = qs.filter(fecha__date__gte=request.GET["desde"])
        if request.GET.get("hasta"):
            qs = qs.filter(fecha__date__lte=request.GET["hasta"])
 
        # Limitar a 500 registros en PDF para no romper la memoria
        qs = qs[:500]
 
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm, leftMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm,
        )
 
        styles  = getSampleStyleSheet()
        azul    = colors.HexColor("#003FB7")
        rojo    = colors.HexColor("#DC2626")
        amarillo= colors.HexColor("#D97706")
        gris    = colors.HexColor("#F1F5F9")
 
        titulo_style = ParagraphStyle(
            "titulo", parent=styles["Heading1"],
            textColor=azul, fontSize=14, spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            "sub", parent=styles["Normal"],
            textColor=colors.HexColor("#64748B"), fontSize=8, spaceAfter=12,
        )
        celda_style = ParagraphStyle(
            "celda", parent=styles["Normal"], fontSize=7,
        )
 
        elementos = [
            Paragraph("Reporte de Auditoría — HorizonZero · Caja de ANDE", titulo_style),
            Paragraph(
                f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} "
                f"por {request.user.get_full_name() or request.user.username} · "
                f"Clasificación: USO INTERNO RESTRINGIDO",
                sub_style,
            ),
        ]
 
        # Tabla
        data = [["ID", "Fecha / Hora", "Usuario", "Acción", "Módulo", "Severidad", "Descripción", "IP"]]
 
        SEV_COLOR = {
            "CRITICAL": colors.HexColor("#FEE2E2"),
            "WARNING":  colors.HexColor("#FEF9C3"),
            "INFO":     colors.white,
        }
 
        row_colors = [colors.HexColor("#003FB7")]  # header
 
        for log in qs:
            fecha_str = log.fecha.strftime("%d/%m/%Y\n%H:%M:%S") if log.fecha else ""
            data.append([
                str(log.id),
                fecha_str,
                log.username,
                Paragraph(log.get_accion_display(), celda_style),
                log.get_modulo_display(),
                log.get_severidad_display(),
                Paragraph(log.descripcion[:80], celda_style),
                str(log.ip_address) if log.ip_address else "—",
            ])
            row_colors.append(SEV_COLOR.get(log.severidad, colors.white))
 
        col_widths = [1.2*cm, 2.5*cm, 3*cm, 3.5*cm, 2.5*cm, 2*cm, 8*cm, 2.5*cm]
 
        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            # Header
            ("BACKGROUND",  (0,0), (-1,0), azul),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 7),
            ("ALIGN",       (0,0), (-1,0), "CENTER"),
            ("BOTTOMPADDING",(0,0),(-1,0), 6),
            # Cuerpo
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",    (0,1), (-1,-1), 7),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, gris]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("RIGHTPADDING",(0,0), (-1,-1), 4),
        ]))
 
        # Colorear filas según severidad
        for i, color in enumerate(row_colors[1:], 1):
            if color != colors.white:
                tabla.setStyle(TableStyle([
                    ("BACKGROUND", (0,i), (-1,i), color),
                ]))
 
        elementos.append(tabla)
 
        doc.build(elementos)
        buffer.seek(0)
 
        fstr = dt.now().strftime("%Y%m%d_%H%M")
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="audit_log_{fstr}.pdf"'
        return response
 
    except ImportError:
        messages.error(request, "ReportLab no está instalado. Ejecutá: pip install reportlab")
        return redirect("dashboard:logs_home")



 