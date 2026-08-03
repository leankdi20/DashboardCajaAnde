import re
import unicodedata
from io import BytesIO
from urllib.parse import quote

from django.http import Http404
from django.shortcuts import render
from django.utils.timezone import localtime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


EXPORT_TITLE_MAP = {
    "encuesta_satisfaccion_exportar": "Encuesta_Satisfaccion",
    "encuesta_satisfaccion_detalle_exportar": "Encuesta_Satisfaccion_Detalle",
    "encuesta_satisfaccion_oficina_exportar": "Oficina_Digital",
    "encuesta_satisfaccion_oficina_detalle_exportar": "Oficina_Digital_Detalle",
    "encuesta_experiencia_web_exportar": "Experiencia_Web",
    "encuesta_whatsapp_agente_exportar": "WhatsApp_Agente",
    "encuesta_whatsapp_agente_detalle_exportar": "WhatsApp_Agente_Detalle",
    "encuesta_whatsapp_exportar": "WhatsApp",
    "encuesta_whatsapp_detalle_exportar": "WhatsApp_Detalle",
    "encuesta_feria_salud_exportar": "Feria_Salud",
    "logs_export_excel": "Logs_Sistema",
    "soli_tarj_credito_export": "Tarjetas_Credito",
    "soli_tarj_credito_export_detalle": "Tarjetas_Credito_Detalle",
    "soli_tarj_debito_export": "Tarjetas_Debito_Ciudadano_Oro",
    "soli_tarj_debito_gestion_export": "Tarjetas_Debito_Gestion",
    "soli_redencion_puntos_export": "Redencion_Puntos",
    "soli_redencion_puntos_export_detalle": "Redencion_Puntos_Detalle",
    "caja_ande_asistencia_export": "Caja_Ande_Asistencia",
    "caja_ande_asistencia_export_detalle": "Caja_Ande_Asistencia_Detalle",
    "soli_deposito_salario_export": "Deposito_Salario",
    "soli_deposito_salario_export_detalle": "Deposito_Salario_Detalle",
    "soli_ahorro_mod_cuota_export": "Ahorro_Modificacion_Cuota",
    "soli_ahorro_mod_cuota_export_detalle": "Ahorro_Modificacion_Cuota_Detalle",
    "soli_reinversion_ahorro_export": "Reinversion_Ahorro",
    "soli_reinversion_ahorro_export_detalle": "Reinversion_Ahorro_Detalle",
    "soli_autorizacion_ahorro_nuevo_export": "Autorizacion_Ahorro_Nuevo",
    "soli_autorizacion_ahorro_nuevo_export_detalle": "Autorizacion_Ahorro_Nuevo_Detalle",
    "soli_compra_vehiculo_export": "Compra_Vehiculo",
    "soli_compra_vehiculo_export_detalle": "Compra_Vehiculo_Detalle",
    "soli_prestamo_vivienda_export": "Prestamo_Vivienda",
    "soli_prestamo_vivienda_export_detalle": "Prestamo_Vivienda_Detalle",
    "soli_prestamo_desarrollo_export": "Prestamo_Desarrollo",
    "soli_prestamo_desarrollo_export_detalle": "Prestamo_Desarrollo_Detalle",
    "soli_presolicitud_credito_personal_export": "Presolicitud_Credito_Personal",
    "soli_presolicitud_credito_personal_export_detalle": "Presolicitud_Credito_Personal_Detalle",
    "comprobante_autorizacion_ahorro_export": "Comprobante_Autorizacion_Ahorro",
    "comprobante_autorizacion_ahorro_export_detalle": "Comprobante_Autorizacion_Ahorro_Detalle",
    "comprobantes_pago_export": "Comprobantes_Pago",
    "comprobantes_pago_export_detalle": "Comprobantes_Pago_Detalle",
    "soli_clave_temporal_cajatel_export": "Clave_Temporal_CajaTel",
    "soli_clave_temporal_cajatel_export_detalle": "Clave_Temporal_CajaTel_Detalle",
    "soli_seguro_viajero_export": "Seguro_Viajero",
    "soli_seguro_viajero_export_detalle": "Seguro_Viajero_Detalle",
    "soli_marchamo_export": "Marchamo",
    "soli_marchamo_export_detalle": "Marchamo_Detalle",
}

FILTER_PARAM_ALIASES = (
    ("desde", "fecha_inicio"),
    ("hasta", "fecha_fin"),
)
EXTRA_FILTER_KEYS = ("clasificacion", "sucursal", "unidad", "agente", "gestion", "nombre", "cedula")
IGNORED_FILTER_KEYS = {"page", "csrfmiddlewaretoken"}
FILTER_LABELS = {
    "fecha_inicio": "Fecha desde",
    "fecha_fin": "Fecha hasta",
    "desde": "Fecha desde",
    "hasta": "Fecha hasta",
    "clasificacion": "Clasificación filtro",
    "sucursal": "Sucursal",
    "unidad": "Unidad",
    "agente": "Agente",
    "gestion": "Gestión",
    "nombre": "Nombre",
    "cedula": "Cédula",
    "encuesta_id": "Encuesta ID",
    "respuesta_id": "Respuesta ID",
    "banco": "Banco",
    "tipo_ahorro": "Tipo ahorro",
    "tipo_credito": "Tipo crédito",
    "tipo_vehiculo": "Tipo vehículo",
    "forma_pago": "Forma pago",
    "numero_contrato": "Contrato",
}


def _slugify_filename_part(value, max_length=60):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text[:max_length] or "SinDato"


def _get_param(request, *keys):
    for key in keys:
        value = request.GET.get(key)
        if value:
            return value
    return ""


def _build_export_filename(request):
    match = getattr(request, "resolver_match", None)
    url_name = getattr(match, "url_name", "") or ""
    kwargs = getattr(match, "kwargs", {}) or {}

    title = EXPORT_TITLE_MAP.get(url_name) or _slugify_filename_part(url_name or "Reporte")
    parts = [title]

    respuesta_id = kwargs.get("respuesta_id")
    if respuesta_id:
        parts.append(f"ID_{respuesta_id}")

    desde = _get_param(request, *FILTER_PARAM_ALIASES[0])
    hasta = _get_param(request, *FILTER_PARAM_ALIASES[1])
    if desde or hasta:
        parts.append(f"Rango_{_slugify_filename_part(desde or 'Inicio')}_a_{_slugify_filename_part(hasta or 'Hoy')}")
    else:
        parts.append("Rango_Todos")

    clasificacion = request.GET.get("clasificacion", "")
    if clasificacion:
        parts.append(f"Clasificacion_{_slugify_filename_part(clasificacion)}")
    else:
        parts.append("Clasificacion_Todas")

    for key in EXTRA_FILTER_KEYS[1:]:
        value = request.GET.get(key, "")
        if value:
            parts.append(f"{key.title()}_{_slugify_filename_part(value, max_length=30)}")

    generated_at = localtime().strftime("%Y-%m-%d_%H-%M")
    parts.append(f"Generado_{generated_at}")

    return f"{'_'.join(parts)}.xlsx"


def _build_excel_filter_summary(request):
    filters = []
    for key, values in request.GET.lists():
        if key in IGNORED_FILTER_KEYS:
            continue
        cleaned_values = [str(value).strip() for value in values if str(value).strip()]
        if not cleaned_values:
            continue
        label = FILTER_LABELS.get(key, key.replace("_", " ").title())
        filters.append(f"{label}: {', '.join(cleaned_values)}")

    if not filters:
        return "Filtros aplicados: Sin filtros"
    return "Filtros aplicados: " + " | ".join(filters)


def _build_excel_period_summary(request):
    desde = _get_param(request, "fecha_inicio", "desde")
    hasta = _get_param(request, "fecha_fin", "hasta")
    if not desde and not hasta:
        return "Periodo: Todos los registros"
    return f"Periodo: {desde or 'Inicio'} a {hasta or 'Hoy'}"


def _inject_excel_header_rows(response, request):
    try:
        workbook = load_workbook(filename=BytesIO(response.content))
    except Exception:
        return response

    worksheet = workbook.active
    worksheet.insert_rows(1, amount=3)

    url_name = getattr(getattr(request, "resolver_match", None), "url_name", "") or ""
    title = EXPORT_TITLE_MAP.get(url_name, "Reporte")
    generated_at = localtime().strftime("%d/%m/%Y %H:%M")

    worksheet["A1"] = f"Reporte: {title}"
    worksheet["A2"] = f"{_build_excel_period_summary(request)} | {_build_excel_filter_summary(request)} | Generado: {generated_at}"
    worksheet["A3"] = "Clasificación: CONFIDENCIAL"

    max_col = max(worksheet.max_column, 8)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    worksheet["A2"].font = Font(bold=True, size=10)
    worksheet["A3"].font = Font(bold=True, size=10)

    worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    worksheet["A2"].fill = PatternFill("solid", fgColor="D9E2F3")
    worksheet["A3"].fill = PatternFill("solid", fgColor="FFF2CC")

    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 38
    worksheet.row_dimensions[3].height = 22

    output = BytesIO()
    workbook.save(output)
    response.content = output.getvalue()
    response["Content-Length"] = str(len(response.content))
    return response


class LimpiarCachePermisosMiddleware:
    """
    Limpia el caché de permisos de Django en cada request.
    Necesario cuando se asignan grupos con sesión activa.
    """
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        if request.user.is_authenticated:
            for attr in ("_perm_cache", "_user_perm_cache", "_group_perm_cache"):
                try:
                    delattr(request.user, attr)
                except AttributeError:
                    pass
        return self.get_response(request)
    
class ContentSecurityPolicyMiddleware:
    """
    Agrega la cabecera Content-Security-Policy a cada respuesta.
    """
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        response = self.get_response(request)
        response["Content-Security-Policy"] = (
            "default-src 'self'; "
            "img-src 'self' data: blob: https:; "
            "style-src 'self' 'unsafe-inline' https:; "
            "script-src 'self' 'unsafe-inline' 'unsafe-eval' https:; "
            "font-src 'self' data: https:; "
            "connect-src 'self' https:; "
            "frame-ancestors 'self'; "
            "object-src 'none'; "
            "base-uri 'self';"
        )
        response["X-Content-Type-Options"] = "nosniff"
        response["X-XSS-Protection"] = "1; mode=block"
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        response["Permissions-Policy"] = (
            "geolocation=(), microphone=(), camera=(), payment=(), usb=()"
        )
        return response


class FriendlyNotFoundMiddleware:
    """
    Muestra la pagina 404 amigable incluso cuando DEBUG=True.
    """

    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        try:
            response = self.get_response(request)
        except Http404:
            return render(request, "errors/404.html", status=404)

        if response.status_code == 404 and "text/html" in response.get("Content-Type", ""):
            return render(request, "errors/404.html", status=404)

        return response


class ExcelDownloadFilenameMiddleware:
    """
    Homologa el nombre final de los archivos Excel descargados.
    """

    XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        response = self.get_response(request)

        content_type = response.get("Content-Type", "")
        disposition = response.get("Content-Disposition", "")
        if self.XLSX_CONTENT_TYPE not in content_type:
            return response
        if "attachment" not in disposition.lower():
            return response

        response = _inject_excel_header_rows(response, request)
        filename = _build_export_filename(request)
        response["Content-Disposition"] = (
            f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"
        )
        return response
